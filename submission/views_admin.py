from django.shortcuts import render
from submission.models import (
    UserProfile,
    Submission,
    Schedule,
    Stamp,
    ScoringItem,
    DiscussionBonus,
    ExperimentCompletion,
    ExperimentProgress,
    ExperimentTaskConfig,
    ExperimentEquipmentConfig,
    Course,
    CourseOffering,
    Enrollment,
)
from attendance.models import AttendanceRecord
from datetime import time, timedelta, date as dt_date, datetime
from zoneinfo import ZoneInfo
from django.core.files.storage import default_storage
import json
import csv
import io
import os
import zipfile
import re
import unicodedata
import math
import fitz
from submission.decorators import role_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from collections import Counter, defaultdict
from django.contrib.auth.models import Group, Permission, User
from django.utils import timezone
from urllib.parse import unquote
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse
from openpyxl import load_workbook
import xlrd

JST = ZoneInfo("Asia/Tokyo")
ABSENCE_CUTOFF_TIME = time(21, 0)
LATE_CHECKIN_TIME = time(13, 40)
SYSTEM_SCORING_DEFS = [
    {'code': 'late', 'label': '遅刻', 'category': 'pre'},
    {'code': 'late', 'label': '遅刻', 'category': 'main'},
    {'code': 'absence', 'label': '欠席', 'category': 'main'},
    {'code': 'lab_time', 'label': '実験時間', 'category': 'pre'},
    {'code': 'discussion', 'label': 'ディスカッション', 'category': 'main'},
]
SCHEDULE_DATE_RE = re.compile(
    r'(?P<month>\d{1,2})\s*月\s*(?P<day>\d{1,2})\s*日(?:\s*[（(]\s*(?P<weekday>[月火水木金土日])\s*[)）])?'
)
GROUP_ASSIGNMENT_SESSION_KEY = 'group_assignment_preview_v1'
GROUP_ASSIGNMENT_DEFAULT_DAYS = ('火', '木')
GROUP_ASSIGNMENT_GROUPS_PER_DAY = 20
GROUP_ASSIGNMENT_REASON_KEYWORDS = (
    '教職', '教養', '教育', '専門'
)
GROUP_ASSIGNMENT_JP_FONT = '/usr/share/fonts/opentype/ipafont-gothic/ipag.ttf'
GROUP_ASSIGNMENT_DEFAULT_CONSTRAINTS = {
    'group_count': GROUP_ASSIGNMENT_GROUPS_PER_DAY,
    'ideal_group_size': 3,
    'separate_repeaters': False,
    'forbid_two_females': False,
    'forbid_mixed_two_person_group': False,
    'use_liberal_arts_credits_priority': False,
    'balance_gpa': False,
}


def _weekday_label(dt):
    # 0=Mon ... 6=Sun
    return ['月', '火', '水', '木', '金', '土', '日'][dt.weekday()]


def _system_item_weight(offering_id, code):
    if not offering_id:
        return 0.0
    offering = CourseOffering.objects.select_related('course').filter(id=offering_id).first()
    if not offering:
        return 0.0
    for category in ('pre', 'main'):
        specific = ScoringItem.objects.filter(
            category=category,
            course_offering_id=offering_id,
            code=code
        ).order_by('order').first()
        if specific:
            return float(specific.weight)
        common = ScoringItem.objects.filter(
            category=category,
            course_id=offering.course_id,
            course_offering__isnull=True,
            code=code
        ).order_by('order').first()
        if common:
            return float(common.weight)
    return 0.0


def _absence_penalty_weight(offering_id):
    return abs(_system_item_weight(offering_id, 'absence'))


def _discussion_bonus_weight(offering_id):
    return abs(_system_item_weight(offering_id, 'discussion'))


def _collect_requested_groups(request):
    raw_values = request.GET.getlist('experiment_group')
    if not raw_values:
        single = request.GET.get('experiment_group')
        raw_values = [single] if single else []
    groups = []
    seen = set()
    for raw in raw_values:
        for token in str(raw or '').split(','):
            value = token.strip()
            if not value:
                continue
            if value.isdigit():
                value = value.zfill(2)
            if value in seen:
                continue
            groups.append(value)
            seen.add(value)
    return groups


def _ensure_course_system_items(course):
    for definition in SYSTEM_SCORING_DEFS:
        code = definition['code']
        category = definition['category']
        label = definition['label']
        item = ScoringItem.objects.filter(
            course=course,
            course_offering__isnull=True,
            category=category,
            code=code
        ).first()
        if item:
            continue
        ScoringItem.objects.create(
            course=course,
            course_offering=None,
            category=category,
            label=label,
            code=code,
            is_system=True,
            show_in_grading_form=False,
            order=0,
            weight=0,
        )


def _sum_score_details(submissions):
    totals = {}
    order = []
    for sub in submissions:
        for detail in sub.score_details or []:
            key = detail.get('code') or detail.get('label') or ''
            if not key:
                continue
            if key not in totals:
                totals[key] = {
                    'label': detail.get('label') or key,
                    'weight': detail.get('weight', 1),
                    'value': 0,
                }
                order.append(key)
            totals[key]['value'] += detail.get('value', 0)
            if totals[key].get('weight') in (None, '') and detail.get('weight') is not None:
                totals[key]['weight'] = detail.get('weight')
    return [totals[k] for k in order]


def _aggregate_score_details(student_id, experiment_number, offering_id=None):
    qs = Submission.objects.filter(
        student_id=student_id,
        experiment_number=experiment_number,
        score_details__isnull=False
    )
    if offering_id:
        qs = qs.filter(course_offering_id=offering_id)
    pre_qs = qs.filter(report_type='prep').order_by('submitted_at', 'id')
    main_qs = qs.filter(report_type='main').order_by('submitted_at', 'id')
    return {
        'pre': _sum_score_details(pre_qs),
        'main': _sum_score_details(main_qs),
    }


def _normalize_task_list(values):
    if values is None:
        return []
    if isinstance(values, str):
        values = values.replace('\r', '').replace(',', '\n').split('\n')
    normalized = []
    seen = set()
    for value in values:
        token = str(value).strip()
        if not token or token in seen:
            continue
        normalized.append(token)
        seen.add(token)
    return normalized


def _normalize_equipment_item_list(values):
    if values is None:
        return []
    if isinstance(values, str):
        values = values.replace('\r', '').replace(',', '\n').split('\n')
    normalized = []
    seen = set()
    for value in values:
        token = str(value).strip()
        if not token or token in seen:
            continue
        normalized.append(token)
        seen.add(token)
    return normalized


def _ga_normalize_text(value):
    return unicodedata.normalize('NFKC', str(value or '')).strip()


def _ga_normalize_email(value):
    return _ga_normalize_text(value).lower()


def _ga_normalize_name(value):
    normalized = _ga_normalize_text(value).replace('　', ' ')
    return ''.join(normalized.split()).lower()


def _ga_extract_student_id(value):
    normalized = _ga_normalize_text(value)
    digits = ''.join(ch for ch in normalized if ch.isdigit())
    if not digits:
        return ''
    return digits[-4:].zfill(4)


def _ga_parse_grade_level(value):
    digits = ''.join(ch for ch in _ga_normalize_text(value) if ch.isdigit())
    return int(digits) if digits else None


def _ga_parse_float(value):
    normalized = _ga_normalize_text(value).replace(',', '')
    if not normalized:
        return None
    try:
        return float(normalized)
    except (TypeError, ValueError):
        return None


def _ga_parse_checkbox(value):
    return str(value or '').strip().lower() in {'1', 'true', 'on', 'yes'}


def _ga_build_constraint_profile(raw_profile):
    profile = dict(GROUP_ASSIGNMENT_DEFAULT_CONSTRAINTS)
    raw_profile = raw_profile or {}
    try:
        group_count = int(raw_profile.get('group_count') or profile['group_count'])
    except (TypeError, ValueError):
        group_count = profile['group_count']
    try:
        ideal_group_size = int(raw_profile.get('ideal_group_size') or profile['ideal_group_size'])
    except (TypeError, ValueError):
        ideal_group_size = profile['ideal_group_size']
    group_count = max(1, group_count)
    ideal_group_size = max(2, min(5, ideal_group_size))
    profile.update({
        'group_count': group_count,
        'ideal_group_size': ideal_group_size,
        'separate_repeaters': _ga_parse_checkbox(raw_profile.get('separate_repeaters')),
        'forbid_two_females': _ga_parse_checkbox(raw_profile.get('forbid_two_females')),
        'forbid_mixed_two_person_group': _ga_parse_checkbox(raw_profile.get('forbid_mixed_two_person_group')),
        'use_liberal_arts_credits_priority': _ga_parse_checkbox(raw_profile.get('use_liberal_arts_credits_priority')),
        'balance_gpa': _ga_parse_checkbox(raw_profile.get('balance_gpa')),
    })
    return profile


def _ga_day_char_from_text(value):
    normalized = _ga_normalize_text(value)
    if not normalized or 'どちら' in normalized:
        return ''
    for day in ('月', '火', '水', '木', '金', '土', '日'):
        if normalized.startswith(day):
            return day
    return ''


def _ga_day_label(day):
    day = _ga_normalize_text(day)
    if not day:
        return '未設定'
    if day.endswith('曜'):
        return day
    return f'{day}曜'


def _ga_reason_priority(preferred_day, reason):
    if not preferred_day:
        return 0
    normalized = _ga_normalize_text(reason)
    if not normalized:
        return 1
    if any(keyword in normalized for keyword in GROUP_ASSIGNMENT_REASON_KEYWORDS):
        return 3
    return 2


def _ga_credit_assignment_priority(student, use_priority=False):
    if not use_priority:
        return 0
    credits = student.get('liberal_arts_credits')
    if credits is None:
        return float('-inf')
    return -credits


def _ga_credit_rebalance_priority(student, use_priority=False):
    if not use_priority:
        return 0
    credits = student.get('liberal_arts_credits')
    if credits is None:
        return float('inf')
    return -credits


def _ga_read_csv_dicts(uploaded_file):
    raw = uploaded_file.read()
    uploaded_file.seek(0)
    for encoding in ('utf-8-sig', 'cp932', 'utf-8', 'shift_jis'):
        try:
            text = raw.decode(encoding)
            reader = csv.DictReader(io.StringIO(text))
            return [dict(row) for row in reader]
        except UnicodeDecodeError:
            continue
    raise ValueError(f'CSVの文字コードを判別できません: {uploaded_file.name}')


def _ga_iter_workbook_rows(uploaded_file):
    filename = (uploaded_file.name or '').lower()
    if filename.endswith('.xlsx'):
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
        for sheet in workbook.worksheets:
            rows = [[cell if cell is not None else '' for cell in row] for row in sheet.iter_rows(values_only=True)]
            yield sheet.title, rows
        return
    if filename.endswith('.xls'):
        data = uploaded_file.read()
        uploaded_file.seek(0)
        workbook = xlrd.open_workbook(file_contents=data)
        for sheet in workbook.sheets():
            rows = [
                [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]
                for row_idx in range(sheet.nrows)
            ]
            yield sheet.name, rows
        return
    raise ValueError(f'未対応のExcel形式です: {uploaded_file.name}')


def _ga_find_header_mapping(rows, required_headers):
    for row_index, row in enumerate(rows[:20]):
        normalized_row = [_ga_normalize_text(cell) for cell in row]
        if all(header in normalized_row for header in required_headers):
            return row_index, {header: normalized_row.index(header) for header in required_headers}
    raise ValueError(f'必要な列が見つかりません: {", ".join(required_headers)}')


def _ga_compact_header_cell(value):
    return _ga_normalize_text(value).replace('\n', '').replace('/', '').replace(' ', '')


def _ga_parse_participants(uploaded_file):
    rows = _ga_read_csv_dicts(uploaded_file)
    participants = []
    seen_emails = set()
    for row in rows:
        email = _ga_normalize_email(row.get('メールアドレス'))
        if not email or email in seen_emails:
            continue
        name = _ga_normalize_text(row.get('名') or row.get('氏名') or row.get('名前'))
        participants.append({
            'email': email,
            'display_name': name or email,
        })
        seen_emails.add(email)
    if not participants:
        raise ValueError('履修予定者ファイルに有効なデータがありません。')
    return participants


def _ga_parse_existing_assignments(uploaded_file):
    rows = _ga_read_csv_dicts(uploaded_file)
    assignments = []
    for row in rows:
        normalized = {(_ga_normalize_text(key).replace('\ufeff', '').strip()): value for key, value in (row or {}).items()}
        email = _ga_normalize_email(normalized.get('メールアドレス'))
        student_id = _ga_extract_student_id(normalized.get('学生番号'))
        name = _ga_normalize_text(normalized.get('名前') or normalized.get('氏名'))
        day = _ga_day_char_from_text(normalized.get('曜日'))
        group_no_raw = _ga_normalize_text(normalized.get('班'))
        if not any([email, student_id, name]) or not day:
            continue
        assignments.append({
            'email': email,
            'student_id': student_id,
            'name': name,
            'name_key': _ga_normalize_name(name),
            'day': day,
            'group_no': group_no_raw,
        })
    return assignments


def _ga_parse_survey(uploaded_file):
    rows = _ga_read_csv_dicts(uploaded_file)
    survey_map = {}
    for index, row in enumerate(rows):
        email = _ga_normalize_email(row.get('ユーザー名') or row.get('ユーザ名') or row.get('メールアドレス'))
        if not email:
            continue
        timestamp_raw = _ga_normalize_text(row.get('タイムスタンプ'))
        try:
            sort_key = datetime.strptime(timestamp_raw, '%Y/%m/%d %H:%M:%S')
        except ValueError:
            sort_key = index
        survey_map[email] = {
            'timestamp': timestamp_raw,
            'sort_key': sort_key,
            'student_id': _ga_extract_student_id(row.get('学生番号')),
            'name': _ga_normalize_text(row.get('氏名')),
            'preferred_day': _ga_day_char_from_text(row.get('希望曜日')),
            'preferred_day_raw': _ga_normalize_text(row.get('希望曜日')),
            'reason': _ga_normalize_text(row.get('希望理由')),
        }
    return survey_map


def _ga_parse_roster(uploaded_file):
    sheets = list(_ga_iter_workbook_rows(uploaded_file))
    for _, rows in sheets:
        try:
            header_row, mapping = _ga_find_header_mapping(rows, ['学生番号', '学生氏名', '学年', '性別'])
        except ValueError:
            continue
        records = []
        for row in rows[header_row + 1:]:
            student_id = _ga_extract_student_id(row[mapping['学生番号']] if mapping['学生番号'] < len(row) else '')
            name = _ga_normalize_text(row[mapping['学生氏名']] if mapping['学生氏名'] < len(row) else '')
            if not student_id and not name:
                continue
            grade_level = _ga_parse_grade_level(row[mapping['学年']] if mapping['学年'] < len(row) else '')
            sex = _ga_normalize_text(row[mapping['性別']] if mapping['性別'] < len(row) else '')
            records.append({
                'student_id': student_id,
                'name': name,
                'name_key': _ga_normalize_name(name),
                'grade_level': grade_level,
                'sex': sex,
            })
        if records:
            return records
    raise ValueError('名簿ファイルから必要な列（学生番号, 学生氏名, 学年, 性別）を読み取れません。')


def _ga_parse_grade_rows(uploaded_file):
    records = []
    for sheet_name, rows in _ga_iter_workbook_rows(uploaded_file):
        mapping = None
        header_row = None
        for row_index in range(min(len(rows) - 1, 20)):
            top = rows[row_index]
            bottom = rows[row_index + 1]
            width = max(len(top), len(bottom))
            compact_pairs = []
            top_only = []
            bottom_only = []
            for col_index in range(width):
                top_value = top[col_index] if col_index < len(top) else ''
                bottom_value = bottom[col_index] if col_index < len(bottom) else ''
                top_compact = _ga_compact_header_cell(top_value)
                bottom_compact = _ga_compact_header_cell(bottom_value)
                compact_pairs.append(_ga_compact_header_cell(f'{top_value}{bottom_value}'))
                top_only.append(top_compact)
                bottom_only.append(bottom_compact)

            if '学生番号' not in compact_pairs or '学生氏名' not in compact_pairs:
                continue

            student_id_index = compact_pairs.index('学生番号')
            student_name_index = compact_pairs.index('学生氏名')

            total_gpa_index = None
            total_block_start = None
            liberal_arts_credit_index = None
            liberal_arts_block_start = None
            for col_index in range(width):
                if top_only[col_index] == '合計':
                    total_block_start = col_index
                    break
            if total_block_start is not None:
                for col_index in range(total_block_start, width):
                    if bottom_only[col_index] == 'GPA':
                        total_gpa_index = col_index
            for col_index in range(width):
                if top_only[col_index] == '教養計':
                    liberal_arts_block_start = col_index
                    break
            if liberal_arts_block_start is not None:
                for col_index in range(liberal_arts_block_start, width):
                    if bottom_only[col_index] == '取得単位':
                        liberal_arts_credit_index = col_index

            fallback_gpa_index = compact_pairs.index('GPA') if 'GPA' in compact_pairs else None
            gpa_index = total_gpa_index if total_gpa_index is not None else fallback_gpa_index

            if gpa_index is None:
                continue

            mapping = {
                '学生番号': student_id_index,
                '学生氏名': student_name_index,
                'GPA': gpa_index,
                '教養計取得単位': liberal_arts_credit_index,
            }
            header_row = row_index + 1
            break
        if mapping is None or header_row is None:
            continue
        for row in rows[header_row + 1:]:
            student_id = _ga_extract_student_id(row[mapping['学生番号']] if mapping['学生番号'] < len(row) else '')
            name = _ga_normalize_text(row[mapping['学生氏名']] if mapping['学生氏名'] < len(row) else '')
            gpa_source = row[mapping['GPA']] if mapping['GPA'] < len(row) else ''
            gpa = _ga_parse_float(gpa_source)
            liberal_arts_credit_source = ''
            liberal_arts_credit = None
            credit_index = mapping.get('教養計取得単位')
            if credit_index is not None and credit_index < len(row):
                liberal_arts_credit_source = row[credit_index]
                liberal_arts_credit = _ga_parse_float(liberal_arts_credit_source)
            if not student_id and not name:
                continue
            records.append({
                'student_id': student_id,
                'name': name,
                'name_key': _ga_normalize_name(name),
                'gpa': gpa,
                'gpa_display': _ga_normalize_text(gpa_source),
                'liberal_arts_credits': liberal_arts_credit,
                'liberal_arts_credits_display': _ga_normalize_text(liberal_arts_credit_source),
                'sheet_name': sheet_name,
            })
    if not records:
        raise ValueError('成績ファイルから必要な列（学生番号, 学生氏名, 合計配下のGPA）を読み取れません。')
    return records


def _ga_build_unique_name_map(records):
    grouped = defaultdict(list)
    for record in records:
        name_key = record.get('name_key') or _ga_normalize_name(record.get('name'))
        if not name_key:
            continue
        grouped[name_key].append(record)
    return {
        name_key: values[0]
        for name_key, values in grouped.items()
        if len(values) == 1
    }


def _ga_student_sort_value(student_id):
    if student_id and student_id.isdigit():
        return (0, int(student_id))
    return (1, student_id or '')


def _ga_select_days(_merged_rows):
    return list(GROUP_ASSIGNMENT_DEFAULT_DAYS)


def _ga_desired_day_counts(total_students, day_labels):
    base = total_students // len(day_labels)
    remainder = total_students % len(day_labels)
    return {
        day: base + (1 if index < remainder else 0)
        for index, day in enumerate(day_labels)
    }


def _ga_choose_day(student, counts, desired_counts, day_labels):
    preferred = student.get('preferred_day')
    if preferred in day_labels:
        other = [day for day in day_labels if day != preferred][0]
        if counts[preferred] < desired_counts[preferred]:
            return preferred
        if student.get('reason_priority', 0) >= 3 and counts[preferred] <= desired_counts[preferred] + 1:
            return preferred
        if counts[other] < desired_counts[other]:
            return other
        return preferred if counts[preferred] <= counts[other] else other
    ordered = sorted(day_labels, key=lambda day: (counts[day] - desired_counts[day], counts[day]))
    return ordered[0]


def _ga_rebalance_days(assignments, day_labels, desired_counts, constraint_profile=None):
    constraint_profile = constraint_profile or GROUP_ASSIGNMENT_DEFAULT_CONSTRAINTS
    day_a, day_b = day_labels
    while abs(len(assignments[day_a]) - len(assignments[day_b])) > 1:
        high_day = day_a if len(assignments[day_a]) > len(assignments[day_b]) else day_b
        low_day = day_b if high_day == day_a else day_a
        movable = sorted(
            [item for item in assignments[high_day] if not item.get('day_fixed')],
            key=lambda item: (
                0 if not item.get('preferred_day') else 1,
                item.get('reason_priority', 0),
                _ga_credit_rebalance_priority(item, constraint_profile.get('use_liberal_arts_credits_priority')),
                0 if item.get('preferred_day') == low_day else 1,
                0 if item.get('is_repeater') else 1,
            )
        )
        if not movable:
            break
        student = movable[0]
        assignments[high_day].remove(student)
        assignments[low_day].append(student)
        student['assigned_day'] = low_day
        student['assigned_day_label'] = _ga_day_label(low_day)


def _ga_assign_days(merged_rows, day_labels, constraint_profile=None):
    constraint_profile = constraint_profile or GROUP_ASSIGNMENT_DEFAULT_CONSTRAINTS
    desired_counts = _ga_desired_day_counts(len(merged_rows), day_labels)
    assignments = {day: [] for day in day_labels}
    ordered_students = sorted(
        merged_rows,
        key=lambda item: (
            1 if item.get('is_repeater') else 0,
            item.get('reason_priority', 0),
            _ga_credit_assignment_priority(item, constraint_profile.get('use_liberal_arts_credits_priority')),
            1 if item.get('preferred_day') else 0,
            1 if item.get('sex') == '女' else 0,
            1 if item.get('gpa') is not None else 0,
        ),
        reverse=True,
    )
    counts = {day: 0 for day in day_labels}
    for student in ordered_students:
        chosen_day = _ga_choose_day(student, counts, desired_counts, day_labels)
        assignments[chosen_day].append(student)
        counts[chosen_day] += 1
        student['assigned_day'] = chosen_day
        student['assigned_day_label'] = _ga_day_label(chosen_day)
    _ga_rebalance_days(assignments, day_labels, desired_counts, constraint_profile=constraint_profile)
    for day in day_labels:
        assignments[day].sort(key=lambda item: _ga_student_sort_value(item.get('student_id')))
    return assignments, desired_counts


def _ga_assign_days_with_existing(merged_rows, day_labels, existing_assignments, constraint_profile=None):
    constraint_profile = constraint_profile or GROUP_ASSIGNMENT_DEFAULT_CONSTRAINTS
    desired_counts = _ga_desired_day_counts(len(merged_rows), day_labels)
    assignments = {day: [] for day in day_labels}
    counts = {day: 0 for day in day_labels}
    existing_by_email = {item['email']: item for item in existing_assignments if item.get('email')}
    existing_by_student_id = {item['student_id']: item for item in existing_assignments if item.get('student_id')}
    existing_by_name = _ga_build_unique_name_map(existing_assignments)
    pending_students = []

    for student in merged_rows:
        existing = None
        if student.get('email'):
            existing = existing_by_email.get(student['email'])
        if not existing and student.get('student_id'):
            existing = existing_by_student_id.get(student['student_id'])
        if not existing:
            existing = existing_by_name.get(_ga_normalize_name(student.get('name')))

        fixed_day = existing.get('day') if existing else ''
        if fixed_day in day_labels:
            student['assigned_day'] = fixed_day
            student['assigned_day_label'] = _ga_day_label(fixed_day)
            student['day_fixed'] = True
            assignments[fixed_day].append(student)
            counts[fixed_day] += 1
        else:
            student['day_fixed'] = False
            pending_students.append(student)

    ordered_students = sorted(
        pending_students,
        key=lambda item: (
            1 if item.get('is_repeater') else 0,
            item.get('reason_priority', 0),
            _ga_credit_assignment_priority(item, constraint_profile.get('use_liberal_arts_credits_priority')),
            1 if item.get('preferred_day') else 0,
            1 if item.get('sex') == '女' else 0,
            1 if item.get('gpa') is not None else 0,
        ),
        reverse=True,
    )
    for student in ordered_students:
        chosen_day = _ga_choose_day(student, counts, desired_counts, day_labels)
        assignments[chosen_day].append(student)
        counts[chosen_day] += 1
        student['assigned_day'] = chosen_day
        student['assigned_day_label'] = _ga_day_label(chosen_day)
        student['day_fixed'] = False

    _ga_rebalance_days(assignments, day_labels, desired_counts, constraint_profile=constraint_profile)
    for day in day_labels:
        assignments[day].sort(key=lambda item: _ga_student_sort_value(item.get('student_id')))
    return assignments, desired_counts


def _ga_build_group_sizes(student_count, group_count, minimum_size=2, maximum_size=5, ideal_size=3):
    if group_count <= 0:
        return []
    minimum_total = group_count * minimum_size
    maximum_total = group_count * maximum_size
    if student_count < minimum_total or student_count > maximum_total:
        raise ValueError(
            f'{group_count}班で {student_count} 名を割り当てできません。'
            f'各班 {minimum_size}〜{maximum_size} 名の条件を満たしてください。'
        )
    sizes = [ideal_size] * group_count
    diff = student_count - (ideal_size * group_count)
    if diff > 0:
        while diff > 0:
            moved = False
            for index in range(group_count):
                if diff <= 0:
                    break
                if sizes[index] < maximum_size:
                    sizes[index] += 1
                    diff -= 1
                    moved = True
            if not moved:
                break
    elif diff < 0:
        while diff < 0:
            moved = False
            for index in range(group_count):
                if diff >= 0:
                    break
                if sizes[index] > minimum_size:
                    sizes[index] -= 1
                    diff += 1
                    moved = True
            if not moved:
                break
    return sizes


def _ga_group_sizes_support_gender_rules(group_sizes, female_count, allow_multiple_females=False):
    if allow_multiple_females:
        return True
    if female_count <= 0:
        return True
    eligible_groups = sum(1 for size in group_sizes if size >= 3)
    return female_count <= eligible_groups


def _ga_choose_repeater_group_count(repeater_count, regular_count, total_group_count, female_count=0, maximum_size=5, allow_multiple_females=False, ideal_size=3):
    if repeater_count <= 0:
        return 0
    best_group_count = None
    best_penalty = None
    best_feasible_group_count = None
    best_feasible_penalty = None
    for repeater_group_count in range(1, total_group_count):
        regular_group_count = total_group_count - repeater_group_count
        if regular_count > 0 and regular_group_count <= 0:
            continue
        repeater_ok = (2 * repeater_group_count) <= repeater_count <= (maximum_size * repeater_group_count)
        regular_ok = True if regular_count == 0 else (2 * regular_group_count) <= regular_count <= (maximum_size * regular_group_count)
        if not repeater_ok or not regular_ok:
            continue
        penalty = abs(repeater_count - (ideal_size * repeater_group_count)) + abs(regular_count - (ideal_size * regular_group_count))
        if best_penalty is None or penalty < best_penalty:
            best_penalty = penalty
            best_group_count = repeater_group_count
        repeater_group_sizes = _ga_build_group_sizes(repeater_count, repeater_group_count, maximum_size=maximum_size)
        if not _ga_group_sizes_support_gender_rules(repeater_group_sizes, female_count, allow_multiple_females=allow_multiple_females):
            continue
        if best_feasible_penalty is None or penalty < best_feasible_penalty:
            best_feasible_penalty = penalty
            best_feasible_group_count = repeater_group_count
    if best_feasible_group_count is not None:
        return best_feasible_group_count
    if best_group_count is not None:
        return best_group_count
    return max(1, min(total_group_count - 1, round(repeater_count / ideal_size) or 1))


def _ga_order_students_for_grouping(students, target_gpa):
    gpa_students = [student for student in students if student.get('gpa') is not None]
    no_gpa_students = [student for student in students if student.get('gpa') is None]
    gpa_students.sort(key=lambda item: item.get('gpa', 0), reverse=True)

    woven = []
    left = 0
    right = len(gpa_students) - 1
    take_high = True
    while left <= right:
        if take_high:
            woven.append(gpa_students[left])
            left += 1
        else:
            woven.append(gpa_students[right])
            right -= 1
        take_high = not take_high

    no_gpa_students.sort(
        key=lambda item: (
            1 if item.get('sex') == '女' else 0,
            item.get('reason_priority', 0),
        ),
        reverse=True,
    )
    return woven + no_gpa_students


def _ga_select_best_group(student, groups, target_gpa, allow_multiple_females=False, allow_mixed_two_person_group=False):
    def is_valid_candidate(group):
        if len(group['students']) >= group['target_size']:
            return False
        if not allow_multiple_females and student.get('sex') == '女' and group['female_count'] > 0:
            return False
        if (
            not allow_mixed_two_person_group
            and
            group['target_size'] == 2
            and len(group['students']) == 1
            and student.get('sex')
            and group['students'][0].get('sex')
            and student.get('sex') != group['students'][0].get('sex')
        ):
            return False
        return True

    def candidate_score(group):
        score = len(group['students']) * 10
        if target_gpa is not None and student.get('gpa') is not None:
            new_count = group['gpa_count'] + 1
            new_total = group['gpa_total'] + student['gpa']
            new_avg = new_total / new_count
            score += abs(new_avg - target_gpa) * 100
        score += abs((len(group['students']) + 1) - group['target_size']) * 5
        score += group['female_count'] * 3
        return score

    candidates = [group for group in groups if len(group['students']) < group['target_size']]
    valid_candidates = [group for group in candidates if is_valid_candidate(group)]
    target_candidates = valid_candidates or candidates
    if not target_candidates:
        return groups[0]

    best_group = None
    best_score = None
    for group in target_candidates:
        score = candidate_score(group)
        if best_score is None or score < best_score:
            best_score = score
            best_group = group
    return best_group or groups[0]


def _ga_assign_students_to_groups(students, group_labels, group_sizes, target_gpa=None, allow_multiple_females=False, allow_mixed_two_person_group=False):
    groups = []
    for label, size in zip(group_labels, group_sizes):
        groups.append({
            'group_no': label,
            'target_size': size,
            'students': [],
            'female_count': 0,
            'male_count': 0,
            'gpa_total': 0.0,
            'gpa_count': 0,
        })
    ordered_students = _ga_order_students_for_grouping(students, target_gpa)
    for student in ordered_students:
        group = _ga_select_best_group(
            student,
            groups,
            target_gpa,
            allow_multiple_females=allow_multiple_females,
            allow_mixed_two_person_group=allow_mixed_two_person_group,
        )
        group['students'].append(student)
        if student.get('sex') == '女':
            group['female_count'] += 1
        if student.get('sex') == '男':
            group['male_count'] += 1
        if student.get('gpa') is not None:
            group['gpa_total'] += student['gpa']
            group['gpa_count'] += 1
        student['assigned_group'] = group['group_no']
    for group in groups:
        group['students'].sort(key=lambda item: _ga_student_sort_value(item.get('student_id')))
        group['gpa_average'] = round(group['gpa_total'] / group['gpa_count'], 2) if group['gpa_count'] else None
    return groups


def _ga_recalculate_group_metrics(group):
    group['female_count'] = sum(1 for student in group['students'] if student.get('sex') == '女')
    group['male_count'] = sum(1 for student in group['students'] if student.get('sex') == '男')
    group['gpa_total'] = sum(student.get('gpa') or 0 for student in group['students'] if student.get('gpa') is not None)
    group['gpa_count'] = sum(1 for student in group['students'] if student.get('gpa') is not None)
    group['count'] = len(group['students'])
    group['gpa_average'] = round(group['gpa_total'] / group['gpa_count'], 2) if group['gpa_count'] else None


def _ga_has_mixed_two_person_group(group):
    if group['target_size'] != 2 or len(group['students']) != 2:
        return False
    sexes = {student.get('sex') for student in group['students'] if student.get('sex')}
    return len(sexes) >= 2


def _ga_group_gender_valid(group, allow_multiple_females=False, allow_mixed_two_person_group=False):
    if not allow_multiple_females and group['female_count'] > 1:
        return False
    if not allow_mixed_two_person_group and _ga_has_mixed_two_person_group(group):
        return False
    return True


def _ga_try_swap_students(group_a, group_b, student_a, student_b, allow_multiple_females=False, allow_mixed_two_person_group=False):
    group_a['students'].remove(student_a)
    group_b['students'].remove(student_b)
    group_a['students'].append(student_b)
    group_b['students'].append(student_a)
    _ga_recalculate_group_metrics(group_a)
    _ga_recalculate_group_metrics(group_b)
    if _ga_group_gender_valid(group_a, allow_multiple_females=allow_multiple_females, allow_mixed_two_person_group=allow_mixed_two_person_group) and _ga_group_gender_valid(group_b, allow_multiple_females=allow_multiple_females, allow_mixed_two_person_group=allow_mixed_two_person_group):
        return True
    group_a['students'].remove(student_b)
    group_b['students'].remove(student_a)
    group_a['students'].append(student_a)
    group_b['students'].append(student_b)
    _ga_recalculate_group_metrics(group_a)
    _ga_recalculate_group_metrics(group_b)
    return False


def _ga_fix_mixed_two_person_groups(groups, allow_multiple_females=False, allow_mixed_two_person_group=False):
    if allow_mixed_two_person_group:
        return
    changed = True
    while changed:
        changed = False
        mixed_groups = [group for group in groups if _ga_has_mixed_two_person_group(group)]
        if not mixed_groups:
            return
        for group in mixed_groups:
            students = list(group['students'])
            resolved = False
            for student_to_replace in students:
                remaining_student = next(student for student in students if student is not student_to_replace)
                needed_sex = remaining_student.get('sex')
                if not needed_sex:
                    continue
                for other_group in groups:
                    if other_group is group:
                        continue
                    for candidate in list(other_group['students']):
                        if candidate.get('sex') != needed_sex:
                            continue
                        if _ga_try_swap_students(
                            group,
                            other_group,
                            student_to_replace,
                            candidate,
                            allow_multiple_females=allow_multiple_females,
                            allow_mixed_two_person_group=allow_mixed_two_person_group,
                        ):
                            changed = True
                            resolved = True
                            break
                    if resolved:
                        break
                if resolved:
                    break


def _ga_fix_female_collisions(groups, allow_multiple_females=False, allow_mixed_two_person_group=False):
    if allow_multiple_females:
        return
    changed = True
    while changed:
        changed = False
        violating_groups = [group for group in groups if group['female_count'] > 1]
        if not violating_groups:
            return
        for group in violating_groups:
            female_students = [student for student in group['students'] if student.get('sex') == '女']
            resolved = False
            for student_to_move in female_students[1:]:
                for other_group in groups:
                    if other_group is group:
                        continue
                    if other_group['female_count'] > 0:
                        continue
                    for candidate in list(other_group['students']):
                        if candidate.get('sex') == '女':
                            continue
                        if _ga_try_swap_students(
                            group,
                            other_group,
                            student_to_move,
                            candidate,
                            allow_multiple_females=allow_multiple_females,
                            allow_mixed_two_person_group=allow_mixed_two_person_group,
                        ):
                            changed = True
                            resolved = True
                            break
                    if resolved:
                        break
                if resolved:
                    break


def _ga_fix_two_person_female_groups(groups, allow_multiple_females=False, allow_mixed_two_person_group=False):
    if allow_multiple_females:
        return
    changed = True
    while changed:
        changed = False
        target_groups = [
            group for group in groups
            if group['target_size'] == 2 and len(group['students']) == 2 and group['female_count'] == 2
        ]
        if not target_groups:
            return
        for group in target_groups:
            female_students = [student for student in group['students'] if student.get('sex') == '女']
            donor_groups = [
                other for other in groups
                if other is not group and other['target_size'] >= 3 and other['female_count'] == 0 and any(s.get('sex') == '男' for s in other['students'])
            ]
            resolved = False
            for idx_a in range(len(donor_groups)):
                for idx_b in range(idx_a + 1, len(donor_groups)):
                    donor_a = donor_groups[idx_a]
                    donor_b = donor_groups[idx_b]
                    male_a = next((student for student in donor_a['students'] if student.get('sex') == '男'), None)
                    male_b = next((student for student in donor_b['students'] if student.get('sex') == '男'), None)
                    if not male_a or not male_b:
                        continue
                    female_a, female_b = female_students[0], female_students[1]
                    group['students'].remove(female_a)
                    group['students'].remove(female_b)
                    donor_a['students'].remove(male_a)
                    donor_b['students'].remove(male_b)
                    group['students'].extend([male_a, male_b])
                    donor_a['students'].append(female_a)
                    donor_b['students'].append(female_b)
                    _ga_recalculate_group_metrics(group)
                    _ga_recalculate_group_metrics(donor_a)
                    _ga_recalculate_group_metrics(donor_b)
                    valid = (
                        _ga_group_gender_valid(group, allow_multiple_females=allow_multiple_females, allow_mixed_two_person_group=allow_mixed_two_person_group)
                        and _ga_group_gender_valid(donor_a, allow_multiple_females=allow_multiple_females, allow_mixed_two_person_group=allow_mixed_two_person_group)
                        and _ga_group_gender_valid(donor_b, allow_multiple_females=allow_multiple_females, allow_mixed_two_person_group=allow_mixed_two_person_group)
                    )
                    if valid:
                        changed = True
                        resolved = True
                        break
                    group['students'].remove(male_a)
                    group['students'].remove(male_b)
                    donor_a['students'].remove(female_a)
                    donor_b['students'].remove(female_b)
                    group['students'].extend([female_a, female_b])
                    donor_a['students'].append(male_a)
                    donor_b['students'].append(male_b)
                    _ga_recalculate_group_metrics(group)
                    _ga_recalculate_group_metrics(donor_a)
                    _ga_recalculate_group_metrics(donor_b)
                if resolved:
                    break


def _ga_group_balance_score(groups, target_gpa=None):
    numbered_groups = sorted(groups, key=lambda item: int(item['group_no']))
    averages = [group['gpa_average'] for group in numbered_groups if group.get('gpa_average') is not None]
    if not averages:
        return (0.0, 0.0, 0.0)
    target_penalty = 0.0
    if target_gpa is not None:
        target_penalty = sum((avg - target_gpa) ** 2 for avg in averages)
    adjacent_penalty = 0.0
    previous_avg = None
    for group in numbered_groups:
        avg = group.get('gpa_average')
        if avg is None:
            continue
        if previous_avg is not None:
            adjacent_penalty += (avg - previous_avg) ** 2
        previous_avg = avg
    range_penalty = (max(averages) - min(averages)) ** 2 if len(averages) >= 2 else 0.0
    return (target_penalty, adjacent_penalty, range_penalty)


def _ga_spread_two_person_group_labels(groups):
    if len(groups) < 2:
        return
    numbered_groups = sorted(groups, key=lambda item: int(item['group_no']))
    two_person_groups = [group for group in numbered_groups if group.get('count', len(group.get('students', []))) == 2]
    if len(two_person_groups) <= 1:
        return
    other_groups = [group for group in numbered_groups if group not in two_person_groups]
    total = len(numbered_groups)
    positions = []
    for index in range(len(two_person_groups)):
        position = round(((index + 0.5) * total / len(two_person_groups)) - 0.5)
        position = max(0, min(total - 1, position))
        while position in positions and position < total - 1:
            position += 1
        while position in positions and position > 0:
            position -= 1
        positions.append(position)
    positions = sorted(set(positions))
    while len(positions) < len(two_person_groups):
        for candidate in range(total):
            if candidate not in positions:
                positions.append(candidate)
                if len(positions) == len(two_person_groups):
                    break
    positions = sorted(positions[:len(two_person_groups)])

    reordered = [None] * total
    two_iter = iter(two_person_groups)
    other_iter = iter(other_groups)
    for index in range(total):
        reordered[index] = next(two_iter) if index in positions else next(other_iter)

    for index, group in enumerate(reordered, start=1):
        group['group_no'] = str(index).zfill(2)
        for student in group['students']:
            student['assigned_group'] = group['group_no']


def _ga_balance_group_gpas(groups, max_diff=0.5, allow_multiple_females=False, allow_mixed_two_person_group=False):
    if len(groups) < 2:
        return
    def current_diff():
        averages = [group['gpa_average'] for group in groups if group.get('gpa_average') is not None]
        if len(averages) < 2:
            return 0
        return max(averages) - min(averages)

    def current_key():
        diff = current_diff()
        target_gpa = None
        values = [group['gpa_average'] for group in groups if group.get('gpa_average') is not None]
        if values:
            target_gpa = sum(values) / len(values)
        target_penalty, adjacent_penalty, range_penalty = _ga_group_balance_score(groups, target_gpa=target_gpa)
        return (
            max(0, diff - max_diff),
            range_penalty,
            adjacent_penalty,
            target_penalty,
            diff,
        )

    for _ in range(120):
        diff = current_diff()
        if diff <= max_diff:
            # Even if the range is already acceptable, continue only if we can smooth adjacent valleys.
            pass
        ranked = sorted([group for group in groups if group.get('gpa_average') is not None], key=lambda item: item['gpa_average'])
        if len(ranked) < 2:
            return
        baseline_key = current_key()
        best_swap = None
        best_key = baseline_key
        candidate_groups = ranked
        for idx_a in range(len(candidate_groups)):
            for idx_b in range(idx_a + 1, len(candidate_groups)):
                group_a = candidate_groups[idx_a]
                group_b = candidate_groups[idx_b]
                for student_a in group_a['students']:
                    if student_a.get('gpa') is None:
                        continue
                    for student_b in group_b['students']:
                        if student_b.get('gpa') is None:
                            continue
                        group_a['students'].remove(student_a)
                        group_b['students'].remove(student_b)
                        group_a['students'].append(student_b)
                        group_b['students'].append(student_a)
                        _ga_recalculate_group_metrics(group_a)
                        _ga_recalculate_group_metrics(group_b)
                        valid = (
                            _ga_group_gender_valid(group_a, allow_multiple_females=allow_multiple_females, allow_mixed_two_person_group=allow_mixed_two_person_group)
                            and _ga_group_gender_valid(group_b, allow_multiple_females=allow_multiple_females, allow_mixed_two_person_group=allow_mixed_two_person_group)
                        )
                        candidate_key = current_key()
                        if valid and candidate_key < best_key:
                            best_key = candidate_key
                            best_swap = (group_a, group_b, student_a, student_b)
                        group_a['students'].remove(student_b)
                        group_b['students'].remove(student_a)
                        group_a['students'].append(student_a)
                        group_b['students'].append(student_b)
                        _ga_recalculate_group_metrics(group_a)
                        _ga_recalculate_group_metrics(group_b)
        if best_swap is None:
            return
        group_a, group_b, student_a, student_b = best_swap
        group_a['students'].remove(student_a)
        group_b['students'].remove(student_b)
        group_a['students'].append(student_b)
        group_b['students'].append(student_a)
        _ga_recalculate_group_metrics(group_a)
        _ga_recalculate_group_metrics(group_b)


def _ga_build_day_group_plan(day, day_students, max_group_size=5, max_gpa_diff=0.5, allow_multiple_females=False, constraint_profile=None):
    constraint_profile = _ga_build_constraint_profile(constraint_profile)
    group_count = constraint_profile['group_count']
    ideal_group_size = constraint_profile['ideal_group_size']
    separate_repeaters = constraint_profile['separate_repeaters']
    forbid_two_females = constraint_profile['forbid_two_females']
    forbid_mixed_two_person_group = constraint_profile['forbid_mixed_two_person_group']
    balance_gpa = constraint_profile['balance_gpa']
    allow_multiple_females = allow_multiple_females or not forbid_two_females
    allow_mixed_two_person_group = not forbid_mixed_two_person_group

    if separate_repeaters:
        repeaters = [student for student in day_students if student.get('is_repeater')]
        regulars = [student for student in day_students if not student.get('is_repeater')]
    else:
        repeaters = []
        regulars = list(day_students)
    warnings = []
    total_group_count = group_count
    student_count = len(day_students)
    if student_count < total_group_count * 2 or student_count > total_group_count * max_group_size:
        raise ValueError(
            f'{_ga_day_label(day)}の人数 {student_count} 名では、'
            f'{total_group_count}班を各班2〜{max_group_size}名で構成できません。'
        )

    repeater_group_count = _ga_choose_repeater_group_count(
        len(repeaters),
        len(regulars),
        total_group_count,
        female_count=sum(1 for student in repeaters if student.get('sex') == '女'),
        maximum_size=max_group_size,
        allow_multiple_females=allow_multiple_females,
        ideal_size=ideal_group_size,
    )
    regular_group_count = total_group_count - repeater_group_count
    if len(repeaters) == 0:
        regular_group_count = total_group_count
    if len(regulars) == 0:
        repeater_group_count = total_group_count
        regular_group_count = 0

    regular_group_labels = [str(index).zfill(2) for index in range(1, regular_group_count + 1)]
    repeater_group_labels = [str(index).zfill(2) for index in range(regular_group_count + 1, total_group_count + 1)]
    regular_group_sizes = _ga_build_group_sizes(
        len(regulars),
        regular_group_count,
        maximum_size=max_group_size,
        ideal_size=ideal_group_size,
    ) if regular_group_count else []
    repeater_group_sizes = _ga_build_group_sizes(
        len(repeaters),
        repeater_group_count,
        maximum_size=max_group_size,
        ideal_size=ideal_group_size,
    ) if repeater_group_count else []

    if any(size == 2 for size in regular_group_sizes + repeater_group_sizes):
        warnings.append(f'{_ga_day_label(day)}は人数条件の都合で2名班が含まれます。')
    if any(size >= 4 for size in regular_group_sizes + repeater_group_sizes):
        warnings.append(f'{_ga_day_label(day)}は人数条件の都合で4名以上の班が含まれます。')

    regular_gpas = [student['gpa'] for student in regulars if student.get('gpa') is not None]
    regular_target_gpa = sum(regular_gpas) / len(regular_gpas) if regular_gpas else None
    regular_groups = _ga_assign_students_to_groups(
        regulars,
        regular_group_labels,
        regular_group_sizes,
        target_gpa=regular_target_gpa,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    ) if regular_group_count else []
    repeater_groups = _ga_assign_students_to_groups(
        repeaters,
        repeater_group_labels,
        repeater_group_sizes,
        target_gpa=None,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    ) if repeater_group_count else []
    _ga_fix_female_collisions(
        regular_groups,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    )
    _ga_fix_female_collisions(
        repeater_groups,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    )
    _ga_fix_two_person_female_groups(
        regular_groups,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    )
    _ga_fix_two_person_female_groups(
        repeater_groups,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    )
    _ga_fix_mixed_two_person_groups(
        regular_groups,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    )
    _ga_fix_mixed_two_person_groups(
        repeater_groups,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    )
    if regular_groups and balance_gpa:
        _ga_balance_group_gpas(
            regular_groups,
            max_diff=max_gpa_diff,
            allow_multiple_females=allow_multiple_females,
            allow_mixed_two_person_group=allow_mixed_two_person_group,
        )
    _ga_fix_female_collisions(
        regular_groups,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    )
    _ga_fix_female_collisions(
        repeater_groups,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    )
    _ga_fix_two_person_female_groups(
        regular_groups,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    )
    _ga_fix_two_person_female_groups(
        repeater_groups,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    )
    _ga_fix_mixed_two_person_groups(
        regular_groups,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    )
    _ga_fix_mixed_two_person_groups(
        repeater_groups,
        allow_multiple_females=allow_multiple_females,
        allow_mixed_two_person_group=allow_mixed_two_person_group,
    )

    all_groups = regular_groups + repeater_groups
    _ga_spread_two_person_group_labels(all_groups)
    all_groups.sort(key=lambda item: item['group_no'])
    for group in all_groups:
        _ga_recalculate_group_metrics(group)
        if forbid_two_females and group['female_count'] > 1:
            warnings.append(f'{_ga_day_label(day)} {group["group_no"]}班 は女子2名以上になっています。')
        if forbid_mixed_two_person_group and _ga_has_mixed_two_person_group(group):
            warnings.append(f'{_ga_day_label(day)} {group["group_no"]}班 は男女2名班になっています。')
    regular_gpa_values = [group['gpa_average'] for group in regular_groups if group.get('gpa_average') is not None]
    gpa_diff_exceeded = (
        balance_gpa
        and len(regular_gpa_values) >= 2
        and (max(regular_gpa_values) - min(regular_gpa_values)) >= max_gpa_diff
    )
    if gpa_diff_exceeded:
        warnings.append(f'{_ga_day_label(day)}の通常班でGPA平均差が{max_gpa_diff:.2f}以上残っています。')
    return {
        'day': day,
        'day_label': _ga_day_label(day),
        'groups': all_groups,
        'warnings': warnings,
        'student_count': len(day_students),
        'repeater_count': len(repeaters),
        'constraints': {
            'max_group_size': max_group_size,
            'max_gpa_diff': max_gpa_diff,
            'allow_multiple_females': allow_multiple_females,
            'allow_mixed_two_person_group': allow_mixed_two_person_group,
            'gpa_diff_exceeded': gpa_diff_exceeded,
        },
    }


def _ga_build_day_group_plan_with_relaxation(day, day_students, diff_mode=False, constraint_profile=None):
    constraint_profile = _ga_build_constraint_profile(constraint_profile)
    if not diff_mode:
        plan = _ga_build_day_group_plan(day, day_students, constraint_profile=constraint_profile)
        plan['relaxations'] = []
        return plan

    base_max_group_size = 5
    required_max_group_size = max(base_max_group_size, math.ceil(len(day_students) / constraint_profile['group_count']))
    profiles = [{
        'max_gpa_diff': 0.50,
        'max_group_size': base_max_group_size,
        'allow_multiple_females': not constraint_profile['forbid_two_females'],
        'relaxations': [],
    }]
    if constraint_profile['balance_gpa']:
        profiles.extend([
            {
                'max_gpa_diff': 0.75,
                'max_group_size': base_max_group_size,
                'allow_multiple_females': not constraint_profile['forbid_two_females'],
                'relaxations': ['GPA平均差許容を0.75に緩和'],
            },
            {
                'max_gpa_diff': 1.00,
                'max_group_size': base_max_group_size,
                'allow_multiple_females': not constraint_profile['forbid_two_females'],
                'relaxations': ['GPA平均差許容を1.00に緩和'],
            },
        ])
    if required_max_group_size > base_max_group_size:
        previous_relaxations = list(profiles[-1]['relaxations'])
        profiles.append({
            'max_gpa_diff': profiles[-1]['max_gpa_diff'],
            'max_group_size': required_max_group_size,
            'allow_multiple_females': profiles[-1]['allow_multiple_females'],
            'relaxations': previous_relaxations + [f'班員人数上限を{required_max_group_size}名に緩和'],
        })
    if constraint_profile['forbid_two_females']:
        profiles.append({
            'max_gpa_diff': profiles[-1]['max_gpa_diff'],
            'max_group_size': max(required_max_group_size, profiles[-1]['max_group_size']),
            'allow_multiple_females': True,
            'relaxations': profiles[-1]['relaxations'] + ['女子2名同班禁止を撤廃'],
        })

    last_error = None
    for profile in profiles:
        try:
            plan = _ga_build_day_group_plan(
                day,
                day_students,
                max_group_size=profile['max_group_size'],
                max_gpa_diff=profile['max_gpa_diff'],
                allow_multiple_females=profile['allow_multiple_females'],
                constraint_profile=constraint_profile,
            )
            if plan['constraints'].get('gpa_diff_exceeded') and constraint_profile['balance_gpa'] and not profile['allow_multiple_females']:
                continue
            plan['relaxations'] = profile['relaxations']
            for relaxation in profile['relaxations']:
                plan['warnings'].append(f'{_ga_day_label(day)}は差分割当のため {relaxation} しました。')
            return plan
        except ValueError as exc:
            last_error = exc
            continue
    if last_error is not None:
        raise last_error
    raise ValueError(f'{_ga_day_label(day)}の差分割当案を作成できませんでした。')


def _ga_build_preview_data(participants_file, survey_file, roster_file=None, grades_file=None, target_grade=None, existing_assignment_file=None, constraint_profile=None):
    constraint_profile = _ga_build_constraint_profile(constraint_profile)
    participants = _ga_parse_participants(participants_file)
    survey_map = _ga_parse_survey(survey_file)
    needs_roster = (
        constraint_profile['separate_repeaters']
        or constraint_profile['forbid_two_females']
        or constraint_profile['forbid_mixed_two_person_group']
    )
    needs_grades = (
        constraint_profile['use_liberal_arts_credits_priority']
        or constraint_profile['balance_gpa']
    )
    roster_records = _ga_parse_roster(roster_file) if needs_roster and roster_file else []
    grade_records = _ga_parse_grade_rows(grades_file) if needs_grades and grades_file else []
    existing_assignments = _ga_parse_existing_assignments(existing_assignment_file) if existing_assignment_file else []

    roster_by_student_id = {record['student_id']: record for record in roster_records if record.get('student_id')}
    roster_by_name = _ga_build_unique_name_map(roster_records)
    grade_by_student_id = {record['student_id']: record for record in grade_records if record.get('student_id')}
    grade_by_name = _ga_build_unique_name_map(grade_records)

    merged_rows = []
    warnings = []
    for participant in participants:
        email = participant['email']
        survey = survey_map.get(email)
        source_name = participant['display_name']
        student_id = survey.get('student_id', '') if survey else ''
        roster = roster_by_student_id.get(student_id) if student_id else None
        if not roster:
            roster = roster_by_name.get(_ga_normalize_name(source_name)) or roster_by_name.get(_ga_normalize_name(participant['display_name']))
        if roster and not student_id:
            student_id = roster.get('student_id', '')
        display_name = roster.get('name') if roster and roster.get('name') else source_name
        if roster and roster.get('name') and _ga_normalize_name(roster.get('name')) != _ga_normalize_name(participant.get('display_name')):
            display_name = participant.get('display_name')
        grade = grade_by_student_id.get(student_id) if student_id else None
        if not grade:
            grade = grade_by_name.get(_ga_normalize_name(source_name)) or grade_by_name.get(_ga_normalize_name(participant['display_name']))

        grade_level = roster.get('grade_level') if roster else None
        is_repeater = bool(
            constraint_profile['separate_repeaters']
            and grade_level is not None
            and target_grade is not None
            and grade_level != target_grade
        )
        preferred_day = survey.get('preferred_day', '') if survey else ''
        reason = survey.get('reason', '') if survey else ''
        row = {
            'email': email,
            'name': display_name or participant['display_name'],
            'course_name': participant['display_name'],
            'student_id': student_id,
            'preferred_day': preferred_day,
            'preferred_day_label': _ga_day_label(preferred_day) if preferred_day else 'どちらでもよい',
            'preferred_reason': reason,
            'reason_priority': _ga_reason_priority(preferred_day, reason),
            'sex': roster.get('sex', '') if roster else '',
            'grade_level': grade_level,
            'is_repeater': is_repeater,
            'gpa': grade.get('gpa') if grade else None,
            'gpa_display': grade.get('gpa_display') if grade else '',
            'liberal_arts_credits': grade.get('liberal_arts_credits') if grade else None,
            'liberal_arts_credits_display': grade.get('liberal_arts_credits_display') if grade else '',
            'has_survey': survey is not None,
            'has_roster': roster is not None,
            'has_grade': grade is not None,
            'assigned_day': '',
            'assigned_day_label': '',
            'assigned_group': '',
            'day_fixed': False,
        }
        if not row['has_survey']:
            warnings.append(f'{row["name"]}（{email}）は希望調査の回答が見つかりませんでした。')
        if needs_roster and not row['has_roster']:
            warnings.append(f'{row["name"]}（{email}）は名簿情報の突合に失敗しました。')
        if needs_grades and not row['has_grade'] and not is_repeater:
            warnings.append(f'{row["name"]}（{email}）は成績情報の突合に失敗しました。')
        merged_rows.append(row)

    merged_rows.sort(key=lambda item: _ga_student_sort_value(item.get('student_id')))
    day_labels = _ga_select_days(merged_rows)
    diff_mode = bool(existing_assignments)
    if diff_mode:
        assignments, desired_counts = _ga_assign_days_with_existing(
            merged_rows,
            day_labels,
            existing_assignments,
            constraint_profile=constraint_profile,
        )
        warnings.append('既存班分けCSVを参照し、既存学生の曜日を固定した差分割当モードで処理しました。')
    else:
        assignments, desired_counts = _ga_assign_days(merged_rows, day_labels, constraint_profile=constraint_profile)
    day_panels = []
    group_plans = []
    for day in day_labels:
        day_students = assignments[day]
        day_panels.append({
            'day': day,
            'day_label': _ga_day_label(day),
            'count': len(day_students),
            'students': [
                {
                    **student,
                    'gpa_display': student.get('gpa_display') or ('' if student['gpa'] is None else str(student['gpa'])),
                }
                for student in day_students
            ],
        })
        group_plans.append(
            _ga_build_day_group_plan_with_relaxation(
                day,
                day_students,
                diff_mode=diff_mode,
                constraint_profile=constraint_profile,
            )
        )
    for plan in group_plans:
        warnings.extend(plan.get('warnings', []))

    return {
        'target_grade': target_grade,
        'generated_at': timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S'),
        'merged_rows': merged_rows,
        'day_panels': day_panels,
        'group_plans': group_plans,
        'warnings': list(dict.fromkeys(warnings)),
        'day_labels': day_labels,
        'desired_counts': desired_counts,
        'approved': False,
        'diff_mode': diff_mode,
        'constraint_profile': constraint_profile,
        'uses_roster': needs_roster,
        'uses_grades': needs_grades,
    }


def _ga_result_rows(preview):
    rows = []
    for row in preview.get('merged_rows', []):
        rows.append({
            'name': row.get('name', ''),
            'email': row.get('email', ''),
            'student_id': row.get('student_id', ''),
            'day': row.get('assigned_day', ''),
            'day_label': _ga_day_label(row.get('assigned_day', '')),
            'group_no': row.get('assigned_group', ''),
        })
    rows.sort(key=lambda item: _ga_student_sort_value(item.get('student_id')))
    return rows


def _ga_generate_bulk_csv_response(preview):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="group_assignment_bulk_import.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(['名前', 'メールアドレス', '学生番号', '曜日', '班'])
    for row in _ga_result_rows(preview):
        writer.writerow([
            row['name'],
            row['email'],
            row['student_id'],
            row['day'],
            row['group_no'],
        ])
    return response


def _ga_generate_pdf_response(preview):
    rows = _ga_result_rows(preview)
    doc = fitz.open()
    title = '情報工学実験Ⅱ班分け表'
    output_date = timezone.localdate().strftime('%Y/%m/%d')
    page_width = fitz.paper_size('a4')[0]
    page_height = fitz.paper_size('a4')[1]
    page_margin = 26
    top_band_height = 34
    title_y = 42
    date_y = 42
    table_top = 78
    rows_per_column = 58
    row_height = 11.0
    gutter = 14
    column_total_width = (page_width - (page_margin * 2) - gutter) / 2
    headers = ('学生番号', '氏名', '曜日', '班番号')
    column_ratios = (0.24, 0.44, 0.14, 0.18)
    header_fill = (0.84, 0.89, 0.94)
    title_fill = (0.92, 0.95, 0.98)
    body_fill = (1.0, 1.0, 1.0)
    line_color = (0.64, 0.70, 0.76)
    title_color = (0.24, 0.34, 0.47)
    text_color = (0.10, 0.10, 0.10)
    title_font_size = 18
    date_font_size = 11
    header_font_size = 10
    body_font_size = 8.8
    day_font_size = 8.4
    group_font_size = 8.8
    title_rect = fitz.Rect(page_margin, 20, page_width - page_margin, 54)
    date_rect = fitz.Rect(page_width - 150, 20, page_width - page_margin, 54)
    left_table_x = page_margin
    right_table_x = page_margin + column_total_width + gutter

    def _column_widths():
        return [column_total_width * ratio for ratio in column_ratios]

    def _cell_rect(base_x, row_index, column_index, header=False):
        y = table_top + (row_index * row_height)
        x = base_x + sum(_column_widths()[:column_index])
        width = _column_widths()[column_index]
        height = row_height if not header else row_height + 1
        return fitz.Rect(x, y, x + width, y + height)

    def _draw_textbox(page_obj, rect, text, fontsize, align=1, color=text_color, bold=False):
        fontname = 'jpfont'
        current_size = fontsize
        while current_size >= 6.0:
            remain = page_obj.insert_textbox(
                rect,
                str(text or ''),
                fontname=fontname,
                fontsize=current_size,
                color=color,
                align=align,
            )
            if remain >= 0:
                return
            current_size -= 0.4
        page_obj.insert_textbox(
            rect,
            str(text or ''),
            fontname=fontname,
            fontsize=6.0,
            color=color,
            align=align,
        )

    def _draw_table(page_obj, base_x, table_rows):
        widths = _column_widths()
        header_y0 = table_top
        header_y1 = table_top + row_height + 1
        x = base_x
        for index, header in enumerate(headers):
            rect = fitz.Rect(x, header_y0, x + widths[index], header_y1)
            page_obj.draw_rect(rect, color=line_color, fill=header_fill, width=0.6)
            _draw_textbox(page_obj, rect + (2, 1, -2, -1), header, header_font_size, align=1)
            x += widths[index]

        for row_index in range(rows_per_column):
            y0 = table_top + ((row_index + 1) * row_height)
            y1 = y0 + row_height
            x = base_x
            source_row = table_rows[row_index] if row_index < len(table_rows) else None
            values = (
                source_row['student_id'] if source_row else '',
                source_row['name'] if source_row else '',
                source_row['day_label'] if source_row else '',
                source_row['group_no'] if source_row else '',
            )
            for column_index, value in enumerate(values):
                rect = fitz.Rect(x, y0, x + widths[column_index], y1)
                page_obj.draw_rect(rect, color=line_color, fill=body_fill, width=0.5)
                if source_row:
                    if column_index == 1:
                        _draw_textbox(page_obj, rect + (2, 1, -2, -1), value, body_font_size, align=1)
                    elif column_index == 2:
                        _draw_textbox(page_obj, rect + (1, 1, -1, -1), value, day_font_size, align=1)
                    elif column_index == 3:
                        _draw_textbox(page_obj, rect + (1, 1, -1, -1), value, group_font_size, align=1)
                    else:
                        _draw_textbox(page_obj, rect + (1, 1, -1, -1), value, body_font_size, align=1)
                x += widths[column_index]

    def _draw_page(page_obj, page_rows):
        page_obj.insert_font(fontname='jpfont', fontfile=GROUP_ASSIGNMENT_JP_FONT)
        page_obj.draw_rect(
            fitz.Rect(page_margin, 20, page_width - page_margin, 54),
            color=line_color,
            fill=title_fill,
            width=0.6,
        )
        _draw_textbox(page_obj, title_rect, title, title_font_size, align=1, color=title_color)
        _draw_textbox(page_obj, date_rect, output_date, date_font_size, align=2, color=title_color)
        left_rows = page_rows[:rows_per_column]
        right_rows = page_rows[rows_per_column:(rows_per_column * 2)]
        _draw_table(page_obj, left_table_x, left_rows)
        _draw_table(page_obj, right_table_x, right_rows)

    rows_per_page = rows_per_column * 2
    for page_index in range(max(1, math.ceil(len(rows) / rows_per_page))):
        page = doc.new_page(width=page_width, height=page_height)
        page_rows = rows[page_index * rows_per_page:(page_index + 1) * rows_per_page]
        _draw_page(page, page_rows)

    output = io.BytesIO()
    doc.save(output)
    doc.close()
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="group_assignment.pdf"'
    return response


def _ga_preview_from_session(request):
    preview = request.session.get(GROUP_ASSIGNMENT_SESSION_KEY)
    if not preview:
        raise ValueError('班分け案が生成されていません。')
    return preview

@role_required('admin')
def admin_dashboard(request):
    is_admin = False
    if hasattr(request.user, "userprofile") and request.user.userprofile.role == "admin":
        is_admin = True
    enrollments = list(
        Enrollment.objects.filter(user=request.user).select_related('course_offering__course')
    )
    offerings_data = []
    for enr in enrollments:
        offerings_data.append({
            'id': enr.course_offering_id,
            'course_id': enr.course_offering.course_id,
            'course_code': enr.course_offering.course.code,
            'course_name': enr.course_offering.course.name,
            'year': enr.course_offering.year,
            'meeting_days': enr.course_offering.course.meeting_days,
            'experiment_numbers': enr.course_offering.course.experiment_numbers,
        })
    if not offerings_data and is_admin:
        # 管理者がEnrollment未設定の場合は全開講を選択肢に入れる
        for off in CourseOffering.objects.select_related('course'):
            offerings_data.append({
                'id': off.id,
                'course_id': off.course_id,
                'course_code': off.course.code,
                'course_name': off.course.name,
                'year': off.year,
                'meeting_days': off.course.meeting_days,
                'experiment_numbers': off.course.experiment_numbers,
            })
    default_offering_id = None
    if offerings_data:
        latest = max(offerings_data, key=lambda o: (o['year'], o['id']))
        default_offering_id = latest['id']
    return render(request, 'submission/admin_dashboard.html', {
        'is_admin': 'true' if is_admin else 'false',
        'offerings_json': json.dumps(offerings_data, ensure_ascii=False),
        'default_offering_id': default_offering_id,
    })


@role_required('admin')
def course_management(request):
    return render(request, 'submission/course_management.html', {})

@role_required('admin')
def admin_get_submissions_api(request):
    # 本レポートのみ抽出
    day = request.GET.get('experiment_day')
    group = request.GET.get('experiment_group')
    exp_no = request.GET.get('experiment_number')
    offering_id = request.GET.get('offering_id')
    base_qs = Submission.objects.filter(report_type='main', accepted=False).select_related('student', 'student__userprofile')
    if offering_id:
        base_qs = base_qs.filter(course_offering_id=offering_id)

    # (student_id, experiment_number)で未受付レポートをカウント
    count_map = Counter((sub.student_id, sub.experiment_number) for sub in base_qs)

    # 3回提出されているものを自動で受付
    for (student_id, experiment_number), cnt in count_map.items():
        comp_qs = ExperimentCompletion.objects.filter(student=student_id, experiment_number=experiment_number)
        if offering_id:
            comp_qs = comp_qs.filter(course_offering_id=offering_id)
        comp_status = comp_qs.values_list('completed', flat=True)
        completed = comp_status[0] if comp_status else False

        progress_qs = ExperimentProgress.objects.filter(
            student_id=student_id,
            experiment_number=experiment_number,
        )
        if offering_id:
            progress_qs = progress_qs.filter(course_offering_id=offering_id)
        else:
            progress_qs = progress_qs.filter(course_offering__isnull=True)
        has_any_progress = progress_qs.exists()

        if cnt >= 3 and (completed or has_any_progress):
            Submission.objects.filter(
                report_type='main', graded=False, accepted=False,
                student_id=student_id, experiment_number=experiment_number,
                course_offering_id=offering_id if offering_id else None
            ).update(graded=True,accepted=True)
    
    qs = base_qs.filter(graded=False, accepted=False)
    if day:
        qs = qs.filter(student__userprofile__experiment_day=day)
    if group:
        qs = qs.filter(student__userprofile__experiment_group=group)
    if exp_no:
        qs = qs.filter(experiment_number=exp_no)
    
    # 各実験ごとのstudent+experiment_numberで「本レポートの提出回数」を算出
    all_main = Submission.objects.filter(report_type='main')
    if offering_id:
        all_main = all_main.filter(course_offering_id=offering_id)
    submit_count_map = Counter((sub.student_id, sub.experiment_number) for sub in all_main)
    
    detail_cache = {}
    submissions = []
    for sub in qs:
        up = getattr(sub.student, 'userprofile', None)
        detail_offering_id = offering_id or sub.course_offering_id
        cache_key = (sub.student_id, sub.experiment_number, detail_offering_id)
        if cache_key not in detail_cache:
            detail_cache[cache_key] = _aggregate_score_details(
                sub.student_id, sub.experiment_number, detail_offering_id
            )
        details = detail_cache[cache_key]
        submit_count = submit_count_map[(sub.student_id, sub.experiment_number)]  # 本レポート提出回数
        submissions.append({
            'id': sub.id,
            'experiment_day': up.experiment_day if up else "",
            'experiment_group': up.experiment_group if up else "",
            'experiment_number': sub.experiment_number,
            'full_name': up.full_name if up else "",
            'file': sub.file.url if sub.file else "",  # 既存互換
            'file_url': sub.file.url if sub.file else "",
            'file_name': sub.file.name.split('/')[-1] if sub.file else "",
            'score': (
                sum(detail.get("value", 0) * detail.get("weight", 1) for detail in sub.score_details)
                if sub.score_details else "0"
            ),
            "score_details": sub.score_details if sub.score_details else "",
            'pre_score_details': details['pre'],
            'main_score_details': details['main'],
            'submission_count': submit_count,
        })
    return JsonResponse({'submissions': submissions})


@role_required('admin')
def admin_get_accepted_submissions_api(request):
    day = request.GET.get('experiment_day')
    group = request.GET.get('experiment_group')
    exp_no = request.GET.get('experiment_number')
    student_id = request.GET.get('student_id')
    offering_id = request.GET.get('offering_id')
    qs = Submission.objects.filter(report_type='main', accepted=True).select_related('student', 'student__userprofile')
    if offering_id:
        qs = qs.filter(course_offering_id=offering_id)
    if day:
        qs = qs.filter(student__userprofile__experiment_day=day)
    if group:
        qs = qs.filter(student__userprofile__experiment_group=group)
    if exp_no:
        qs = qs.filter(experiment_number=exp_no)
    if student_id:
        qs = qs.filter(student__userprofile__student_id__icontains=student_id)

    detail_cache = {}
    submissions = []
    for sub in qs:
        up = getattr(sub.student, 'userprofile', None)
        detail_offering_id = offering_id or sub.course_offering_id
        cache_key = (sub.student_id, sub.experiment_number, detail_offering_id)
        if cache_key not in detail_cache:
            detail_cache[cache_key] = _aggregate_score_details(
                sub.student_id, sub.experiment_number, detail_offering_id
            )
        details = detail_cache[cache_key]
        submissions.append({
            'id': sub.id,
            'experiment_day': up.experiment_day if up else "",
            'experiment_group': up.experiment_group if up else "",
            'experiment_number': sub.experiment_number,
            'full_name': up.full_name if up else "",
            'student_id': up.student_id if up else "",
            'file': sub.file.url if sub.file else "",
            'file_url': sub.file.url if sub.file else "",
            'file_name': sub.file.name.split('/')[-1] if sub.file else "",
            'score': (
                sum(detail.get("value", 0) * detail.get("weight", 1) for detail in sub.score_details)
                if sub.score_details else "0"
            ),
            "score_details": sub.score_details if sub.score_details else "",
            'pre_score_details': details['pre'],
            'main_score_details': details['main'],
        })
    return JsonResponse({'submissions': submissions})


@role_required('admin')
@require_POST
def admin_return_submission(request):
    try:
        data = json.loads(request.body)
        submission_id = data.get('submission_id')
        sub = Submission.objects.get(id=submission_id, report_type='main')
        # ファイル物理削除
        if sub.file and os.path.isfile(sub.file.path):
            os.remove(sub.file.path)
        if sub.graded_file and os.path.isfile(sub.graded_file.path):
            os.remove(sub.graded_file.path)
        sub.delete()
        return JsonResponse({'status': 'success'})
    except Submission.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': '提出物が見つかりません'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# ---------------------------
# Course / Offering / Enrollment API
# ---------------------------
@role_required('admin')
def admin_course_data_api(request):
    courses = list(Course.objects.all().values('id', 'name', 'code', 'meeting_days', 'experiment_numbers'))
    for c in courses:
        c['experiment_numbers'] = c.get('experiment_numbers') or []
    offerings = []
    for off in CourseOffering.objects.select_related('course'):
        offerings.append({
            'id': off.id,
            'course_id': off.course_id,
            'course_code': off.course.code,
            'course_name': off.course.name,
            'year': off.year,
            'meeting_days': off.course.meeting_days,
            'experiment_numbers': off.course.experiment_numbers,
        })
    enrollments = []
    for enr in Enrollment.objects.exclude(role='student').select_related('user', 'course_offering', 'course_offering__course'):
        up = getattr(enr.user, 'userprofile', None)
        enrollments.append({
            'id': enr.id,
            'user_id': enr.user_id,
            'full_name': up.full_name if up else enr.user.username,
            'student_id': up.student_id if up else '',
            'email': enr.user.email,
            'course_offering_id': enr.course_offering_id,
            'course_code': enr.course_offering.course.code,
            'course_name': enr.course_offering.course.name,
            'year': enr.course_offering.year,
            'role': enr.role,
            'experiment_day': enr.experiment_day,
            'experiment_group': enr.experiment_group,
        })
    users = []
    for u in User.objects.filter(userprofile__role__in=['admin', 'teacher', 'course-teacher', 'non-editing teacher']):
        up = getattr(u, 'userprofile', None)
        users.append({
            'id': u.id,
            'full_name': up.full_name if up else u.username,
            'student_id': up.student_id if up else '',
            'email': u.email,
        })
    task_configs = []
    for cfg in ExperimentTaskConfig.objects.select_related('course_offering__course').order_by(
        'course_offering__year', 'course_offering__course__code', 'experiment_number'
    ):
        task_list = cfg.task_list if isinstance(cfg.task_list, list) else []
        task_configs.append({
            'id': cfg.id,
            'course_offering_id': cfg.course_offering_id,
            'course_code': cfg.course_offering.course.code,
            'course_name': cfg.course_offering.course.name,
            'year': cfg.course_offering.year,
            'experiment_number': cfg.experiment_number,
            'task_list': [str(task).strip() for task in task_list if str(task).strip()],
        })
    equipment_configs = []
    for cfg in ExperimentEquipmentConfig.objects.select_related('course_offering__course').order_by(
        'course_offering__year', 'course_offering__course__code', 'experiment_number'
    ):
        items = cfg.items_json if isinstance(cfg.items_json, list) else []
        equipment_configs.append({
            'id': cfg.id,
            'course_offering_id': cfg.course_offering_id,
            'course_code': cfg.course_offering.course.code,
            'course_name': cfg.course_offering.course.name,
            'year': cfg.course_offering.year,
            'experiment_number': cfg.experiment_number,
            'items_json': [str(item).strip() for item in items if str(item).strip()],
        })
    return JsonResponse({
        'courses': courses,
        'offerings': offerings,
        'enrollments': enrollments,
        'users': users,
        'task_configs': task_configs,
        'equipment_configs': equipment_configs,
    })


@role_required('admin')
@require_POST
def admin_add_course(request):
    data = json.loads(request.body)
    name = data.get('name')
    code = data.get('code')
    meeting_days = data.get('meeting_days', [])
    experiment_numbers = data.get('experiment_numbers', [])
    if not name or not code:
        return JsonResponse({'status': 'error', 'message': 'name and code are required'}, status=400)
    course, created = Course.objects.get_or_create(
        code=code,
        defaults={
            'name': name,
            'meeting_days': meeting_days,
            'experiment_numbers': experiment_numbers,
        }
    )
    if not created:
        return JsonResponse({'status': 'error', 'message': 'code already exists'}, status=400)
    _ensure_course_system_items(course)
    return JsonResponse({'status': 'success', 'course': {
        'id': course.id,
        'name': course.name,
        'code': course.code,
        'meeting_days': course.meeting_days,
        'experiment_numbers': course.experiment_numbers,
    }})


@role_required('admin')
@require_POST
def admin_update_course(request, course_id):
    try:
        data = json.loads(request.body)
        name = data.get('name')
        code = data.get('code')
        meeting_days = data.get('meeting_days', [])
        experiment_numbers = data.get('experiment_numbers', [])
        if not name or not code:
            return JsonResponse({'status': 'error', 'message': 'name and code are required'}, status=400)
        if Course.objects.filter(code=code).exclude(id=course_id).exists():
            return JsonResponse({'status': 'error', 'message': 'code already exists'}, status=400)
        course = Course.objects.get(id=course_id)
        course.name = name
        course.code = code
        course.meeting_days = meeting_days
        course.experiment_numbers = experiment_numbers
        course.save()
        return JsonResponse({'status': 'success', 'course': {
            'id': course.id,
            'name': course.name,
            'code': course.code,
            'meeting_days': course.meeting_days,
            'experiment_numbers': course.experiment_numbers,
        }})
    except Course.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@role_required('admin')
@require_POST
def admin_delete_course(request, course_id):
    try:
        Course.objects.get(id=course_id).delete()
        return JsonResponse({'status': 'success'})
    except Course.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'not found'}, status=404)


@role_required('admin')
@require_POST
def admin_add_offering(request):
    data = json.loads(request.body)
    course_id = data.get('course_id')
    year = data.get('year')
    if not course_id or not year:
        return JsonResponse({'status': 'error', 'message': 'course_id and year are required'}, status=400)
    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'course not found'}, status=404)
    off, created = CourseOffering.objects.get_or_create(course=course, year=year)
    if not created:
        return JsonResponse({'status': 'error', 'message': 'offering already exists'}, status=400)
    return JsonResponse({'status': 'success', 'offering': {
        'id': off.id,
        'course_id': course.id,
        'course_code': course.code,
        'course_name': course.name,
        'year': off.year,
        'meeting_days': course.meeting_days,
        'experiment_numbers': course.experiment_numbers,
    }})


@role_required('admin')
@require_POST
def admin_delete_offering(request, offering_id):
    try:
        CourseOffering.objects.get(id=offering_id).delete()
        return JsonResponse({'status': 'success'})
    except CourseOffering.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'not found'}, status=404)


@role_required('admin')
@require_POST
def admin_add_enrollment(request):
    data = json.loads(request.body)
    user_id = data.get('user_id')
    offering_id = data.get('offering_id')
    role = data.get('role')
    exp_day = data.get('experiment_day', '')
    exp_group = data.get('experiment_group', '')
    if not (user_id and offering_id and role):
        return JsonResponse({'status': 'error', 'message': 'user_id, offering_id, role are required'}, status=400)
    try:
        user = User.objects.get(id=user_id)
        offering = CourseOffering.objects.get(id=offering_id)
        enr, created = Enrollment.objects.get_or_create(
            user=user, course_offering=offering, role=role,
            defaults={'experiment_day': exp_day, 'experiment_group': exp_group}
        )
        if not created:
            return JsonResponse({'status': 'error', 'message': 'already enrolled'}, status=400)
        return JsonResponse({'status': 'success', 'enrollment': {
            'id': enr.id,
            'user_id': user.id,
            'full_name': getattr(user, 'userprofile', None).full_name if hasattr(user, 'userprofile') else user.username,
            'student_id': getattr(user, 'userprofile', None).student_id if hasattr(user, 'userprofile') else '',
            'email': user.email,
            'course_offering_id': offering.id,
            'course_code': offering.course.code,
            'course_name': offering.course.name,
            'year': offering.year,
            'role': enr.role,
            'experiment_day': enr.experiment_day,
            'experiment_group': enr.experiment_group,
        }})
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'user not found'}, status=404)
    except CourseOffering.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'offering not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@role_required('admin')
@require_POST
def admin_delete_enrollment(request, enrollment_id):
    try:
        Enrollment.objects.get(id=enrollment_id).delete()
        return JsonResponse({'status': 'success'})
    except Enrollment.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'not found'}, status=404)


@role_required('admin')
@require_POST
def admin_add_task_config(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON形式が不正です'}, status=400)
    offering_id = data.get('offering_id')
    experiment_number = str(data.get('experiment_number', '')).strip()
    task_list = _normalize_task_list(data.get('task_list', []))
    if not offering_id or not experiment_number:
        return JsonResponse({'status': 'error', 'message': 'offering_id と experiment_number は必須です'}, status=400)
    if not task_list:
        return JsonResponse({'status': 'error', 'message': 'task_list は1件以上必要です'}, status=400)
    offering = CourseOffering.objects.select_related('course').filter(id=offering_id).first()
    if not offering:
        return JsonResponse({'status': 'error', 'message': 'offering not found'}, status=404)
    cfg, created = ExperimentTaskConfig.objects.get_or_create(
        course_offering=offering,
        experiment_number=experiment_number,
        defaults={'task_list': task_list}
    )
    if not created:
        cfg.task_list = task_list
        cfg.save(update_fields=['task_list'])
    return JsonResponse({
        'status': 'success',
        'task_config': {
            'id': cfg.id,
            'course_offering_id': cfg.course_offering_id,
            'course_code': offering.course.code,
            'course_name': offering.course.name,
            'year': offering.year,
            'experiment_number': cfg.experiment_number,
            'task_list': cfg.task_list,
        }
    })


@role_required('admin')
@require_POST
def admin_update_task_config(request, task_config_id):
    try:
        cfg = ExperimentTaskConfig.objects.select_related('course_offering__course').get(id=task_config_id)
    except ExperimentTaskConfig.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'not found'}, status=404)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON形式が不正です'}, status=400)

    offering_id = data.get('offering_id')
    experiment_number = str(data.get('experiment_number', cfg.experiment_number)).strip()
    task_list = _normalize_task_list(data.get('task_list', cfg.task_list))
    if not offering_id or not experiment_number:
        return JsonResponse({'status': 'error', 'message': 'offering_id と experiment_number は必須です'}, status=400)
    if not task_list:
        return JsonResponse({'status': 'error', 'message': 'task_list は1件以上必要です'}, status=400)
    offering = CourseOffering.objects.select_related('course').filter(id=offering_id).first()
    if not offering:
        return JsonResponse({'status': 'error', 'message': 'offering not found'}, status=404)
    duplicate = ExperimentTaskConfig.objects.filter(
        course_offering=offering,
        experiment_number=experiment_number
    ).exclude(id=cfg.id).exists()
    if duplicate:
        return JsonResponse({'status': 'error', 'message': '同じ科目/年度・実験番号の設定が既にあります'}, status=400)
    cfg.course_offering = offering
    cfg.experiment_number = experiment_number
    cfg.task_list = task_list
    cfg.save(update_fields=['course_offering', 'experiment_number', 'task_list'])
    return JsonResponse({
        'status': 'success',
        'task_config': {
            'id': cfg.id,
            'course_offering_id': cfg.course_offering_id,
            'course_code': offering.course.code,
            'course_name': offering.course.name,
            'year': offering.year,
            'experiment_number': cfg.experiment_number,
            'task_list': cfg.task_list,
        }
    })


@role_required('admin')
@require_POST
def admin_delete_task_config(request, task_config_id):
    try:
        ExperimentTaskConfig.objects.get(id=task_config_id).delete()
        return JsonResponse({'status': 'success'})
    except ExperimentTaskConfig.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'not found'}, status=404)


@role_required('admin')
@require_POST
def admin_copy_task_configs(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON形式が不正です'}, status=400)

    target_offering_id = data.get('target_offering_id')
    source_offering_id = data.get('source_offering_id')
    if not target_offering_id or not source_offering_id:
        return JsonResponse({'status': 'error', 'message': 'target_offering_id と source_offering_id は必須です'}, status=400)
    if str(target_offering_id) == str(source_offering_id):
        return JsonResponse({'status': 'error', 'message': '同じ年度はコピー元に選択できません'}, status=400)

    target = CourseOffering.objects.select_related('course').filter(id=target_offering_id).first()
    source = CourseOffering.objects.select_related('course').filter(id=source_offering_id).first()
    if not target or not source:
        return JsonResponse({'status': 'error', 'message': 'offering not found'}, status=404)
    if target.course_id != source.course_id:
        return JsonResponse({'status': 'error', 'message': '同一科目の年度のみコピーできます'}, status=400)
    if int(source.year) >= int(target.year):
        return JsonResponse({'status': 'error', 'message': 'コピー元は過去年度のみ選択できます'}, status=400)

    source_configs = list(
        ExperimentTaskConfig.objects.filter(course_offering=source).order_by('experiment_number', 'id')
    )
    existing_numbers = set(
        ExperimentTaskConfig.objects.filter(course_offering=target).values_list('experiment_number', flat=True)
    )

    to_create = []
    skipped_numbers = []
    for cfg in source_configs:
        exp_no = str(cfg.experiment_number or '').strip()
        if not exp_no or exp_no in existing_numbers:
            if exp_no:
                skipped_numbers.append(exp_no)
            continue
        to_create.append(ExperimentTaskConfig(
            course_offering=target,
            experiment_number=exp_no,
            task_list=_normalize_task_list(cfg.task_list),
        ))
        existing_numbers.add(exp_no)

    if to_create:
        ExperimentTaskConfig.objects.bulk_create(to_create)

    return JsonResponse({
        'status': 'success',
        'created_count': len(to_create),
        'skipped_count': len(skipped_numbers),
        'total_count': len(source_configs),
        'source_year': source.year,
        'target_year': target.year,
    })


@role_required('admin')
@require_POST
def admin_add_equipment_config(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON形式が不正です'}, status=400)

    offering_id = data.get('offering_id')
    experiment_number = str(data.get('experiment_number', '')).strip()
    items = _normalize_equipment_item_list(data.get('items_json', []))
    if not offering_id or not experiment_number:
        return JsonResponse({'status': 'error', 'message': 'offering_id と experiment_number は必須です'}, status=400)
    if not items:
        return JsonResponse({'status': 'error', 'message': '器具チェック項目は1件以上必要です'}, status=400)

    offering = CourseOffering.objects.select_related('course').filter(id=offering_id).first()
    if not offering:
        return JsonResponse({'status': 'error', 'message': 'offering not found'}, status=404)

    cfg, created = ExperimentEquipmentConfig.objects.get_or_create(
        course_offering=offering,
        experiment_number=experiment_number,
        defaults={'items_json': items}
    )
    if not created:
        cfg.items_json = items
        cfg.save(update_fields=['items_json', 'updated_at'])

    return JsonResponse({
        'status': 'success',
        'equipment_config': {
            'id': cfg.id,
            'course_offering_id': cfg.course_offering_id,
            'course_code': offering.course.code,
            'course_name': offering.course.name,
            'year': offering.year,
            'experiment_number': cfg.experiment_number,
            'items_json': cfg.items_json,
        }
    })


@role_required('admin')
@require_POST
def admin_update_equipment_config(request, equipment_config_id):
    try:
        cfg = ExperimentEquipmentConfig.objects.select_related('course_offering__course').get(id=equipment_config_id)
    except ExperimentEquipmentConfig.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'not found'}, status=404)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON形式が不正です'}, status=400)

    offering_id = data.get('offering_id')
    experiment_number = str(data.get('experiment_number', cfg.experiment_number)).strip()
    items = _normalize_equipment_item_list(data.get('items_json', cfg.items_json))
    if not offering_id or not experiment_number:
        return JsonResponse({'status': 'error', 'message': 'offering_id と experiment_number は必須です'}, status=400)
    if not items:
        return JsonResponse({'status': 'error', 'message': '器具チェック項目は1件以上必要です'}, status=400)

    offering = CourseOffering.objects.select_related('course').filter(id=offering_id).first()
    if not offering:
        return JsonResponse({'status': 'error', 'message': 'offering not found'}, status=404)

    duplicate = ExperimentEquipmentConfig.objects.filter(
        course_offering=offering,
        experiment_number=experiment_number
    ).exclude(id=cfg.id).exists()
    if duplicate:
        return JsonResponse({'status': 'error', 'message': '同じ科目/年度・実験番号の設定が既にあります'}, status=400)

    cfg.course_offering = offering
    cfg.experiment_number = experiment_number
    cfg.items_json = items
    cfg.save(update_fields=['course_offering', 'experiment_number', 'items_json', 'updated_at'])

    return JsonResponse({
        'status': 'success',
        'equipment_config': {
            'id': cfg.id,
            'course_offering_id': cfg.course_offering_id,
            'course_code': offering.course.code,
            'course_name': offering.course.name,
            'year': offering.year,
            'experiment_number': cfg.experiment_number,
            'items_json': cfg.items_json,
        }
    })


@role_required('admin')
@require_POST
def admin_delete_equipment_config(request, equipment_config_id):
    try:
        ExperimentEquipmentConfig.objects.get(id=equipment_config_id).delete()
        return JsonResponse({'status': 'success'})
    except ExperimentEquipmentConfig.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'not found'}, status=404)


@role_required('admin')
@require_POST
def admin_copy_equipment_configs(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON形式が不正です'}, status=400)

    target_offering_id = data.get('target_offering_id')
    source_offering_id = data.get('source_offering_id')
    if not target_offering_id or not source_offering_id:
        return JsonResponse({'status': 'error', 'message': 'target_offering_id と source_offering_id は必須です'}, status=400)
    if str(target_offering_id) == str(source_offering_id):
        return JsonResponse({'status': 'error', 'message': '同じ年度はコピー元に選択できません'}, status=400)

    target = CourseOffering.objects.select_related('course').filter(id=target_offering_id).first()
    source = CourseOffering.objects.select_related('course').filter(id=source_offering_id).first()
    if not target or not source:
        return JsonResponse({'status': 'error', 'message': 'offering not found'}, status=404)
    if target.course_id != source.course_id:
        return JsonResponse({'status': 'error', 'message': '同一科目の年度のみコピーできます'}, status=400)
    if int(source.year) >= int(target.year):
        return JsonResponse({'status': 'error', 'message': 'コピー元は過去年度のみ選択できます'}, status=400)

    source_configs = list(
        ExperimentEquipmentConfig.objects.filter(course_offering=source).order_by('experiment_number', 'id')
    )
    existing_numbers = set(
        ExperimentEquipmentConfig.objects.filter(course_offering=target).values_list('experiment_number', flat=True)
    )

    to_create = []
    skipped_numbers = []
    for cfg in source_configs:
        exp_no = str(cfg.experiment_number or '').strip()
        if not exp_no or exp_no in existing_numbers:
            if exp_no:
                skipped_numbers.append(exp_no)
            continue
        to_create.append(ExperimentEquipmentConfig(
            course_offering=target,
            experiment_number=exp_no,
            items_json=_normalize_equipment_item_list(cfg.items_json),
        ))
        existing_numbers.add(exp_no)

    if to_create:
        ExperimentEquipmentConfig.objects.bulk_create(to_create)

    return JsonResponse({
        'status': 'success',
        'created_count': len(to_create),
        'skipped_count': len(skipped_numbers),
        'total_count': len(source_configs),
        'source_year': source.year,
        'target_year': target.year,
    })


def get_students_api(request):
    student_id = request.GET.get('student_id')
    day = request.GET.get('experiment_day')
    groups = _collect_requested_groups(request)
    offering_id = request.GET.get('offering_id')
    qs = UserProfile.objects.filter(role='student')
    enr_map = {}
    if offering_id:
        enr_qs = Enrollment.objects.filter(role='student', course_offering_id=offering_id)
        if day:
            enr_qs = enr_qs.filter(experiment_day=day)
        if groups:
            enr_qs = enr_qs.filter(experiment_group__in=groups)
        user_ids = enr_qs.values_list('user_id', 'experiment_day', 'experiment_group')
        enr_map = {u: {'experiment_day': d, 'experiment_group': g} for u, d, g in user_ids}
        qs = qs.filter(user_id__in=enr_map.keys())
    else:
        if day:
            qs = qs.filter(experiment_day=day)
        if groups:
            qs = qs.filter(experiment_group__in=groups)
    if student_id:
        qs = qs.filter(student_id__icontains=student_id)
    students = []
    for up in qs:
        override = enr_map.get(up.user_id, {})
        students.append({
            'id': up.id,
            'full_name': up.full_name,
            'student_id': up.student_id,
            'user__email': up.user.email,
            'experiment_day': override.get('experiment_day', up.experiment_day),
            'experiment_group': override.get('experiment_group', up.experiment_group),
            'photo': up.photo.url if up.photo else ''
        })
    return JsonResponse({'students_json': students})

def get_summary_api(request):
    student_id = request.GET.get('student_id')
    offering_id = request.GET.get('offering_id')
    experiment_numbers = []
    students = UserProfile.objects.filter(role='student')

    # 科目/年度を選択している場合、その科目の実験番号とEnrollmentで対象学生を絞る
    if offering_id:
        try:
            off = CourseOffering.objects.select_related('course').get(id=offering_id)
            experiment_numbers = off.course.experiment_numbers or []
        except CourseOffering.DoesNotExist:
            experiment_numbers = []
        user_ids = Enrollment.objects.filter(role='student', course_offering_id=offering_id).values_list('user_id', flat=True)
        students = students.filter(user_id__in=user_ids)
    if not experiment_numbers:
        experiment_numbers = [x[0] for x in Submission.EXPERIMENT_NUMBER_CHOICES]
    if student_id:
        students = students.filter(student_id__icontains=student_id)
    results = []
    for item in students:
        user = item.user
        # 受付済みレポートのみ
        accepted_qs = Submission.objects.filter(
            student=user,
            report_type='main',
            accepted=True
        )
        if offering_id:
            accepted_qs = accepted_qs.filter(course_offering_id=offering_id)
        accepted_reports = accepted_qs.values_list('experiment_number', flat=True)
        accepted_set = set(accepted_reports)
        missing_set = set(experiment_numbers) - accepted_set
        results.append({
            'name': item.full_name,
            'student_id': item.student_id,
            'submitted': len(accepted_set),
            'missing': len(missing_set),
            'accepted_numbers': list(accepted_set),
            'missing_numbers': list(missing_set),
        })
    return JsonResponse({'submission_summary': results})

def _normalize_schedule_pdf_line(text):
    normalized = unicodedata.normalize("NFKC", str(text or ""))
    normalized = normalized.replace('\u3000', ' ')
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    return normalized


def _extract_schedule_pdf_lines(pdf_bytes):
    lines = []
    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        for page in doc:
            text = page.get_text("text") or ""
            for raw_line in text.splitlines():
                line = _normalize_schedule_pdf_line(raw_line)
                if line:
                    lines.append(line)
    return lines


def _resolve_schedule_date(offering_year, month, day):
    year = int(offering_year)
    if month <= 3:
        year += 1
    return dt_date(year, month, day)


def _extract_face_to_face_dates(lines, offering_year):
    extracted = []
    for idx, line in enumerate(lines):
        if '対面授業' not in line:
            continue
        window_indexes = [idx, idx - 1, idx - 2, idx - 3, idx - 4, idx + 1, idx + 2]
        found = False
        for target_idx in window_indexes:
            if target_idx < 0 or target_idx >= len(lines):
                continue
            source_line = lines[target_idx]
            for match in SCHEDULE_DATE_RE.finditer(source_line):
                month = int(match.group('month'))
                day = int(match.group('day'))
                weekday_hint = (match.group('weekday') or '').strip()
                try:
                    date_obj = _resolve_schedule_date(offering_year, month, day)
                except ValueError:
                    continue
                extracted.append({
                    'date': date_obj,
                    'weekday_hint': weekday_hint,
                    'source_line': source_line,
                    'marker_line': line,
                })
                found = True
                break
            if found:
                break
        if not found:
            extracted.append({
                'date': None,
                'weekday_hint': '',
                'source_line': '',
                'marker_line': line,
            })
    return extracted


def _build_schedule_pdf_preview(offering, pdf_bytes):
    lines = _extract_schedule_pdf_lines(pdf_bytes)
    candidates = _extract_face_to_face_dates(lines, offering.year)
    existing_dates = set(
        Schedule.objects.filter(course_offering=offering).values_list('date', flat=True)
    )
    meeting_days = {
        str(day).strip()
        for day in (offering.course.meeting_days or [])
        if str(day).strip()
    }
    if not meeting_days:
        meeting_days = {'月', '火', '水', '木', '金', '土', '日'}

    registerable = []
    duplicate_dates = []
    weekday_mismatch = []
    parse_errors = []
    seen_date_keys = set()

    for item in candidates:
        date_obj = item.get('date')
        if not date_obj:
            parse_errors.append({
                'marker_line': item.get('marker_line', ''),
                'message': '対面授業の近傍で日付を抽出できませんでした',
            })
            continue

        date_key = date_obj.isoformat()
        if date_key in seen_date_keys:
            continue
        seen_date_keys.add(date_key)

        weekday = _weekday_label(date_obj)
        base_payload = {
            'date': date_key,
            'weekday': weekday,
            'weekday_hint': item.get('weekday_hint', ''),
            'source_line': item.get('source_line', ''),
        }

        weekday_hint = item.get('weekday_hint')
        if weekday_hint and weekday_hint != weekday:
            parse_errors.append({
                'marker_line': item.get('marker_line', ''),
                'message': f'PDF内曜日({weekday_hint})と計算曜日({weekday})が不一致です',
                'date': date_key,
            })

        if weekday not in meeting_days:
            weekday_mismatch.append(base_payload)
            continue
        if date_obj in existing_dates:
            duplicate_dates.append(base_payload)
            continue
        registerable.append(base_payload)

    registerable.sort(key=lambda x: x['date'])
    duplicate_dates.sort(key=lambda x: x['date'])
    weekday_mismatch.sort(key=lambda x: x['date'])

    return {
        'registerable_dates': registerable,
        'duplicate_dates': duplicate_dates,
        'weekday_mismatch_dates': weekday_mismatch,
        'parse_errors': parse_errors,
        'meeting_days': sorted(meeting_days),
        'found_face_to_face_count': len([x for x in candidates if x.get('date')]),
    }

def get_schedule_api(request):
    offering_id = request.GET.get('offering_id')
    schedule_qs = Schedule.objects.all()
    if offering_id:
        schedule_qs = schedule_qs.filter(course_offering_id=offering_id)
    schedule_qs = schedule_qs.values('id', 'date', 'course_offering_id')
    schedule = [
        {
            'id': s['id'],
            'date': s['date'].strftime('%Y-%m-%d'),
            'course_offering_id': s.get('course_offering_id'),
        }
        for s in schedule_qs
    ]
    return JsonResponse({'schedule_json': schedule})


@role_required('admin')
@require_POST
def admin_schedule_pdf_preview_api(request):
    offering_id = request.POST.get('offering_id')
    pdf_file = request.FILES.get('pdf')

    if not offering_id:
        return JsonResponse({'status': 'error', 'message': 'offering_id は必須です'}, status=400)
    if not pdf_file:
        return JsonResponse({'status': 'error', 'message': 'PDFファイルは必須です'}, status=400)

    offering = CourseOffering.objects.select_related('course').filter(id=offering_id).first()
    if not offering:
        return JsonResponse({'status': 'error', 'message': '科目/年度が不正です'}, status=400)

    try:
        pdf_bytes = pdf_file.read()
        if not pdf_bytes:
            return JsonResponse({'status': 'error', 'message': 'PDFファイルが空です'}, status=400)
        preview = _build_schedule_pdf_preview(offering, pdf_bytes)
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': f'PDF解析に失敗しました: {exc}'}, status=400)

    return JsonResponse({
        'status': 'success',
        'preview': preview,
    })


@role_required('admin')
@require_POST
def admin_schedule_pdf_commit_api(request):
    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON形式が不正です'}, status=400)

    offering_id = data.get('offering_id')
    dates = data.get('dates') or []
    if not offering_id:
        return JsonResponse({'status': 'error', 'message': 'offering_id は必須です'}, status=400)
    if not isinstance(dates, list) or not dates:
        return JsonResponse({'status': 'error', 'message': 'dates は1件以上必要です'}, status=400)

    offering = CourseOffering.objects.select_related('course').filter(id=offering_id).first()
    if not offering:
        return JsonResponse({'status': 'error', 'message': '科目/年度が不正です'}, status=400)

    meeting_days = {
        str(day).strip()
        for day in (offering.course.meeting_days or [])
        if str(day).strip()
    }
    if not meeting_days:
        meeting_days = {'月', '火', '水', '木', '金', '土', '日'}

    existing_dates = set(
        Schedule.objects.filter(course_offering=offering).values_list('date', flat=True)
    )

    created = []
    skipped_duplicate = []
    skipped_weekday = []
    skipped_invalid = []
    to_create = []

    for token in dates:
        token_str = str(token).strip()
        if not token_str:
            continue
        try:
            date_obj = dt_date.fromisoformat(token_str)
        except ValueError:
            skipped_invalid.append(token_str)
            continue
        weekday = _weekday_label(date_obj)
        if weekday not in meeting_days:
            skipped_weekday.append({'date': token_str, 'weekday': weekday})
            continue
        if date_obj in existing_dates:
            skipped_duplicate.append({'date': token_str, 'weekday': weekday})
            continue
        to_create.append(Schedule(date=date_obj, course_offering=offering))
        existing_dates.add(date_obj)
        created.append({'date': token_str, 'weekday': weekday})

    if to_create:
        Schedule.objects.bulk_create(to_create)

    return JsonResponse({
        'status': 'success',
        'created_dates': created,
        'skipped_duplicate_dates': skipped_duplicate,
        'skipped_weekday_mismatch_dates': skipped_weekday,
        'skipped_invalid_dates': skipped_invalid,
    })

@role_required('admin')
def add_schedule_api(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            date = data.get('date')
            offering_id = data.get('offering_id')
            # バリデーション: 日付必須
            if not date:
                return JsonResponse({'status': 'error', 'message': '日付は必須です'}, status=400)
            course_offering = None
            if offering_id:
                course_offering = CourseOffering.objects.filter(id=offering_id).first()
            # 同一科目/年度で同一日付が既に登録されている場合はエラー
            if course_offering and Schedule.objects.filter(course_offering=course_offering, date=date).exists():
                return JsonResponse({'status': 'error', 'message': '同じ日付が既に登録されています'}, status=400)
            s = Schedule.objects.create(date=date, course_offering=course_offering)
            s.refresh_from_db()
            return JsonResponse({'status': 'success', 'schedule': {'id': s.id, 'date': s.date.strftime('%Y-%m-%d'), 'course_offering_id': s.course_offering_id}})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'POSTでリクエストしてください'}, status=400)

@role_required('admin')
def update_schedule_api(request, schedule_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            date = data.get('date')
            offering_id = data.get('offering_id')
            if not date:
                return JsonResponse({'status': 'error', 'message': '日付は必須です'}, status=400)
            s = Schedule.objects.get(id=schedule_id)
            s.date = date
            if offering_id:
                course_offering = CourseOffering.objects.filter(id=offering_id).first()
                s.course_offering = course_offering
            # 更新時も重複チェック（自身は除外）
            if s.course_offering and Schedule.objects.filter(course_offering=s.course_offering, date=date).exclude(id=schedule_id).exists():
                return JsonResponse({'status': 'error', 'message': '同じ日付が既に登録されています'}, status=400)
            s.save()
            s.refresh_from_db()
            return JsonResponse({'status': 'success', 'schedule': {'id': s.id, 'date': s.date.strftime('%Y-%m-%d'), 'course_offering_id': s.course_offering_id}})
        except Schedule.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Scheduleが見つかりません'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'POSTでリクエストしてください'}, status=400)

@role_required('admin')
def delete_schedule_api(request, schedule_id):
    if request.method == 'POST':
        try:
            s = Schedule.objects.get(id=schedule_id)
            s.delete()
            return JsonResponse({'status': 'success'})
        except Schedule.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Scheduleが見つかりません'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'POSTでリクエストしてください'}, status=400)

@role_required('admin')
def scoring_items(request):
    system_codes = {'late', 'absence', 'lab_time', 'discussion'}
    courses_data = []
    for course in Course.objects.prefetch_related('offerings').all():
        offerings = list(course.offerings.all().order_by('-year', '-id'))
        courses_data.append({
            'id': course.id,
            'course_code': course.code,
            'course_name': course.name,
            'offerings': [
                {'id': off.id, 'year': off.year}
                for off in offerings
            ],
        })

    selected_course = None
    selected_offering_id = None
    course_param = request.GET.get('course_id')
    offering_param = request.GET.get('offering_id')

    if course_param:
        try:
            selected_course = Course.objects.get(id=int(course_param))
        except (Course.DoesNotExist, TypeError, ValueError):
            selected_course = None

    if offering_param and offering_param != 'common':
        try:
            candidate = CourseOffering.objects.select_related('course').get(id=int(offering_param))
        except (CourseOffering.DoesNotExist, TypeError, ValueError):
            candidate = None
        if candidate and (selected_course is None or candidate.course_id == selected_course.id):
            selected_course = candidate.course
            selected_offering_id = candidate.id

    if selected_course is None and courses_data:
        latest_offering = CourseOffering.objects.select_related('course').order_by('-year', '-id').first()
        if latest_offering:
            selected_course = latest_offering.course
            selected_offering_id = latest_offering.id
    elif selected_course and offering_param == 'common':
        selected_offering_id = None

    course_offering = None
    if selected_offering_id:
        course_offering = CourseOffering.objects.filter(id=selected_offering_id).first()

    if selected_course:
        _ensure_course_system_items(selected_course)

    if request.method == 'POST':
        data = json.loads(request.body)
        course_id = data.get('course_id')
        offering_id = data.get('offering_id')
        course = None
        course_offering = None
        if course_id:
            course = Course.objects.filter(id=course_id).first()
        if offering_id and offering_id != 'common':
            course_offering = CourseOffering.objects.filter(id=offering_id, course=course).first()
            if not course_offering:
                return JsonResponse({'status': 'error', 'message': '科目/年度が不正です'}, status=400)
        if not course:
            return JsonResponse({'status': 'error', 'message': '科目が不正です'}, status=400)

        def _normalize(items):
            normalized = []
            labels = []
            codes = []
            for item in items:
                label = (item.get('label') or '').strip()
                if not label:
                    continue
                code = (item.get('code') or '').strip() or None
                labels.append(label)
                if code:
                    codes.append(code)
                normalized.append({
                    'label': label,
                    'weight': item.get('weight', 1),
                    'code': code,
                    'is_system': bool(item.get('is_system')),
                    'show_in_grading_form': bool(item.get('show_in_grading_form', True)),
                })
            return normalized, labels, codes

        def _dup_labels(labels):
            seen = set()
            dupes = []
            for label in labels:
                if label in seen and label not in dupes:
                    dupes.append(label)
                seen.add(label)
            return dupes

        pre_items, pre_labels, pre_codes = _normalize(data.get('pre', []))
        main_items, main_labels, main_codes = _normalize(data.get('main', []))

        dup_pre = _dup_labels(pre_labels)
        if dup_pre:
            return JsonResponse(
                {'status': 'error', 'message': f'予習レポートに重複ラベルがあります: {", ".join(dup_pre)}'},
                status=400
            )
        dup_main = _dup_labels(main_labels)
        if dup_main:
            return JsonResponse(
                {'status': 'error', 'message': f'本レポートに重複ラベルがあります: {", ".join(dup_main)}'},
                status=400
            )

        if len(pre_codes) != len(set(pre_codes)):
            return JsonResponse(
                {'status': 'error', 'message': '予習レポートに重複コードがあります'},
                status=400
            )
        if len(main_codes) != len(set(main_codes)):
            return JsonResponse(
                {'status': 'error', 'message': '本レポートに重複コードがあります'},
                status=400
            )

        existing_system_pre = list(
            ScoringItem.objects.filter(
                category='pre',
                course=course,
                course_offering=course_offering,
                is_system=True,
            )
        )
        existing_system_main = list(
            ScoringItem.objects.filter(
                category='main',
                course=course,
                course_offering=course_offering,
                is_system=True,
            )
        )
        payload_pre_codes = {item.get('code') for item in pre_items if item.get('code')}
        payload_main_codes = {item.get('code') for item in main_items if item.get('code')}
        for item in existing_system_pre:
            if item.code and item.code not in payload_pre_codes:
                pre_items.append({
                    'label': item.label,
                    'weight': item.weight,
                    'code': item.code,
                    'is_system': True,
                    'show_in_grading_form': item.show_in_grading_form,
                })
        for item in existing_system_main:
            if item.code and item.code not in payload_main_codes:
                main_items.append({
                    'label': item.label,
                    'weight': item.weight,
                    'code': item.code,
                    'is_system': True,
                    'show_in_grading_form': item.show_in_grading_form,
                })

        ScoringItem.objects.filter(
            category='pre', course=course, course_offering=course_offering
        ).delete()
        ScoringItem.objects.filter(
            category='main', course=course, course_offering=course_offering
        ).delete()
        for idx, item in enumerate(pre_items):
            code = item.get('code') or None
            is_system = bool(item.get('is_system')) or (code in system_codes)
            ScoringItem.objects.create(
                category='pre',
                label=item.get('label', ''),
                weight=item.get('weight', 1),  # ← getでデフォルト値
                order=idx,
                course_offering=course_offering,
                course=course,
                code=code,
                is_system=is_system,
                show_in_grading_form=bool(item.get('show_in_grading_form', True)),
            )
        for idx, item in enumerate(main_items):
            code = item.get('code') or None
            is_system = bool(item.get('is_system')) or (code in system_codes)
            ScoringItem.objects.create(
                category='main',
                label=item.get('label', ''),
                weight=item.get('weight', 1),  # ← getでデフォルト値
                order=idx,
                course_offering=course_offering,
                course=course,
                code=code,
                is_system=is_system,
                show_in_grading_form=bool(item.get('show_in_grading_form', True)),
            )
        return JsonResponse({'status': 'ok'})

    pre_qs = ScoringItem.objects.none()
    main_qs = ScoringItem.objects.none()
    if selected_course:
        if course_offering:
            pre_qs = ScoringItem.objects.filter(category='pre', course_offering=course_offering).order_by('order')
            main_qs = ScoringItem.objects.filter(category='main', course_offering=course_offering).order_by('order')
        else:
            pre_qs = ScoringItem.objects.filter(
                category='pre', course=selected_course, course_offering__isnull=True
            ).order_by('order')
            main_qs = ScoringItem.objects.filter(
                category='main', course=selected_course, course_offering__isnull=True
            ).order_by('order')

    pre = list(pre_qs.values('label', 'weight', 'code', 'is_system', 'show_in_grading_form'))
    main = list(main_qs.values('label', 'weight', 'code', 'is_system', 'show_in_grading_form'))
    for x in pre:
        x['weight'] = int(x['weight'])
        x['code'] = x.get('code') or ''
    for x in main:
        x['weight'] = int(x['weight'])
        x['code'] = x.get('code') or ''
    return render(request, 'submission/scoring_items.html', {
        'pre': json.dumps(pre, ensure_ascii=False),
        'main': json.dumps(main, ensure_ascii=False),
        'courses_json': json.dumps(courses_data, ensure_ascii=False),
        'selected_course_id': selected_course.id if selected_course else '',
        'selected_offering_id': selected_offering_id if selected_offering_id else 'common',
    })

@role_required('admin')
def stamps_view(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        text = data.get('text', '')
        stamp = Stamp.objects.create(text=text)
        return JsonResponse({'status': 'ok', 'stamp': {'id': stamp.id, 'text': stamp.text}})
    stamps = list(Stamp.objects.all().values('id', 'text'))
    return render(request, 'submission/stamps.html', {
        'stamps': json.dumps(stamps, ensure_ascii=False)
    })

@role_required('admin')
def delete_stamp_api(request, stamp_id):
    if request.method == 'POST':
        try:
            stamp = Stamp.objects.get(id=stamp_id)
            stamp.delete()
            return JsonResponse({'status': 'success'})
        except Stamp.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Stampが見つかりません'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'POSTでリクエストしてください'}, status=400)

@require_POST
@role_required('admin')
def accept_submission(request):
    data = json.loads(request.body)
    submission_id = data.get("submission_id")
    from .models import Submission
    sub = Submission.objects.get(id=submission_id)
    sub.accepted = True
    sub.graded = True
    sub.save()
    return JsonResponse({"status": "ok"})

@role_required('admin')
def api_student_reports(request):
    student_id = request.GET.get('student_id')
    offering_id = request.GET.get('offering_id')
    if not student_id:
        return JsonResponse({
            'reports': [],
            'full_name': '',
            'attendance_logs': [],
            'absence_count': 0,
            'experiment_logs': [],
            'discussion_bonus_rows': [],
            'discussion_total_count': 0,
            'discussion_can_edit': True,
        })
    try:
        profile = UserProfile.objects.get(id=student_id)
    except UserProfile.DoesNotExist:
        return JsonResponse({
            'reports': [],
            'full_name': '',
            'attendance_logs': [],
            'absence_count': 0,
            'experiment_logs': [],
            'discussion_bonus_rows': [],
            'discussion_total_count': 0,
            'discussion_can_edit': True,
        })
    full_name = profile.full_name

    attendance_logs = []
    absence_count = 0
    experiment_logs = []
    discussion_bonus_rows = []
    discussion_total_count = 0
    if offering_id:
        try:
            offering_id_int = int(offering_id)
        except (TypeError, ValueError):
            offering_id_int = None
        enrollment = Enrollment.objects.filter(
            user=profile.user,
            course_offering_id=offering_id_int,
            role='student'
        ).first()
        student_day = enrollment.experiment_day if enrollment else profile.experiment_day

        now_local = timezone.localtime(timezone.now(), JST)
        cutoff_date = now_local.date()
        if now_local.time() < ABSENCE_CUTOFF_TIME:
            cutoff_date = cutoff_date - timedelta(days=1)
        schedule_dates = set()
        if student_day:
            for sched in Schedule.objects.filter(
                course_offering_id=offering_id_int,
                date__lte=cutoff_date
            ):
                if _weekday_label(sched.date) == student_day:
                    schedule_dates.add(sched.date)

        attendance_dates = set(
            AttendanceRecord.objects.filter(
                user=profile.user,
                course_offering_id=offering_id_int,
                date__in=schedule_dates
            ).values_list('date', flat=True)
        )
        absence_count = len(schedule_dates - attendance_dates)

        records = AttendanceRecord.objects.filter(
            user=profile.user,
            course_offering_id=offering_id_int
        ).order_by('-date')
        for record in records:
            date_str = record.date.strftime('%Y-%m-%d')
            if record.check_in:
                attendance_logs.append({
                    'date': date_str,
                    'status': '入室',
                    'time': timezone.localtime(record.check_in, JST).strftime('%H:%M')
                })
            if record.check_out:
                attendance_logs.append({
                    'date': date_str,
                    'status': '退室',
                    'time': timezone.localtime(record.check_out, JST).strftime('%H:%M')
                })

        if offering_id_int:
            offering = CourseOffering.objects.select_related('course').filter(id=offering_id_int).first()
            configured_numbers = (offering.course.experiment_numbers if offering else []) or []
            config_map = {}
            for cfg in ExperimentTaskConfig.objects.filter(course_offering_id=offering_id_int):
                config_map[cfg.experiment_number] = _normalize_task_list(cfg.task_list)
            progress_map = {}
            for exp_no, task_no in ExperimentProgress.objects.filter(
                student=profile.user,
                course_offering_id=offering_id_int
            ).values_list('experiment_number', 'task_no'):
                progress_map.setdefault(exp_no, set()).add(str(task_no))
            completion_map = dict(
                ExperimentCompletion.objects.filter(
                    student=profile.user,
                    course_offering_id=offering_id_int
                ).values_list('experiment_number', 'completed')
            )

            # configured順を維持しつつ重複を除外
            all_numbers = []
            seen_numbers = set()
            for exp_no in configured_numbers:
                if exp_no in seen_numbers:
                    continue
                all_numbers.append(exp_no)
                seen_numbers.add(exp_no)
            # 差集合はOR全体に対して計算する（演算子優先順位バグ対策）
            extras = sorted(
                (set(progress_map.keys()) | set(completion_map.keys()) | set(config_map.keys()))
                - set(all_numbers)
            )
            all_numbers.extend(extras)
            for exp_no in all_numbers:
                task_list = _normalize_task_list(config_map.get(exp_no, []))
                completed_set = progress_map.get(exp_no, set())
                ordered_done = [task for task in task_list if task in completed_set]
                ordered_done.extend(sorted(completed_set - set(task_list)))
                if task_list:
                    status = '完了' if set(task_list).issubset(completed_set) else '未完了'
                else:
                    status = '完了' if completion_map.get(exp_no) or completed_set else '未完了'
                experiment_logs.append({
                    'experiment_number': exp_no,
                    'status': status,
                    'completed_tasks': ', '.join(ordered_done) if ordered_done else '-',
                })

            discussion_counts = {
                exp_no: count
                for exp_no, count in DiscussionBonus.objects.filter(
                    student=profile.user,
                    course_offering_id=offering_id_int
                ).values_list('experiment_number', 'count')
            }
            ordered_numbers = []
            seen_numbers = set()
            for exp_no in configured_numbers:
                if exp_no in seen_numbers:
                    continue
                ordered_numbers.append(exp_no)
                seen_numbers.add(exp_no)
            extra_discussion_numbers = sorted(set(discussion_counts.keys()) - set(ordered_numbers))
            ordered_numbers.extend(extra_discussion_numbers)
            discussion_bonus_rows = [
                {
                    'experiment_number': exp_no,
                    'count': int(discussion_counts.get(exp_no, 0) or 0),
                }
                for exp_no in ordered_numbers
            ]
            discussion_total_count = sum(row['count'] for row in discussion_bonus_rows)

    qs = Submission.objects.filter(student__userprofile__id=student_id)
    if offering_id:
        qs = qs.filter(course_offering_id=offering_id)
    qs = qs.order_by('-submitted_at')
    data = []
    for items in qs:
        data.append({
            "file": items.file.url if items.file else "",
            "experiment_number": items.experiment_number,
            "report_type": '予' if items.report_type == 'prep' else '本' ,
            "submitted_at": timezone.localtime(items.submitted_at).strftime('%Y-%m-%d %H:%M'),
        })
    return JsonResponse({
        'reports': data,
        'full_name': full_name,
        'attendance_logs': attendance_logs,
        'absence_count': absence_count,
        'experiment_logs': experiment_logs,
        'discussion_bonus_rows': discussion_bonus_rows,
        'discussion_total_count': discussion_total_count,
        'discussion_can_edit': True,
    })

@role_required('admin')
def user_list_view(request):
    # teacher 以上のみアクセス可
    if not request.user.is_staff:
        return render(request, 'submission/permission_denied.html')

    offerings_qs = CourseOffering.objects.select_related('course')
    offerings_data = [
        {
            'id': o.id,
            'course_id': o.course_id,
            'course_code': o.course.code,
            'course_name': o.course.name,
            'year': o.year,
            'meeting_days': o.course.meeting_days,
        }
        for o in offerings_qs
    ]

    user_data = []
    for user in User.objects.all():
        try:
            profile = user.userprofile
            enrollments = list(
                Enrollment.objects
                .filter(user=user)
                .select_related('course_offering__course')
            )
            last_login = (
                timezone.localtime(user.last_login).strftime("%Y-%m-%d %H:%M")
                if user.last_login else "未ログイン"
            )
            can_view_attendance = user.user_permissions.filter(
                codename='view_attendancerecord',
                content_type__app_label='attendance'
            ).exists()
            is_attendance_only = user.groups.filter(name='attendance_only').exists()

            # Enrollmentごとに行を作成。紐付けが無い場合は空で1行表示。
            if enrollments:
                for enrollment in enrollments:
                    course_offering = enrollment.course_offering
                    group = ""
                    exp_day = enrollment.experiment_day or profile.experiment_day
                    exp_group = enrollment.experiment_group or profile.experiment_group
                    if exp_day and exp_group:
                        group = f"{exp_day}-{str(exp_group).zfill(2)}"
                    user_data.append({
                        'id': user.id,
                        'row_key': f"{user.id}-enr-{enrollment.id}",
                        'enrollment_id': enrollment.id,
                        'name': profile.full_name,
                        'email': user.email,
                        'student_id': profile.student_id,
                        'role': enrollment.role,
                        'group': group,
                        'offering_id': course_offering.id if course_offering else None,
                        'course_id': course_offering.course_id if course_offering else None,
                        'year': course_offering.year if course_offering else None,
                        'last_login': last_login,
                        'can_view_attendance': can_view_attendance,
                        'is_attendance_only': is_attendance_only,
                    })
            else:
                group = ""
                if profile.experiment_day and profile.experiment_group:
                    group = f"{profile.experiment_day}-{str(profile.experiment_group).zfill(2)}"
                user_data.append({
                    'id': user.id,
                    'row_key': f"{user.id}-no-enrollment",
                    'enrollment_id': None,
                    'name': profile.full_name,
                    'email': user.email,
                    'student_id': profile.student_id,
                    'role': profile.role,
                    'group': group,
                    'offering_id': None,
                    'course_id': None,
                    'year': None,
                    'last_login': last_login,
                    'can_view_attendance': can_view_attendance,
                    'is_attendance_only': is_attendance_only,
                })
        except UserProfile.DoesNotExist:
            continue

    context = {
            'users': user_data,
            'users_json': json.dumps(user_data, ensure_ascii=False),
            'offerings_json': json.dumps(offerings_data, ensure_ascii=False),
        }

    return render(request, 'submission/user_list.html', context)


@role_required('admin')
def update_user_role(request, user_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            new_role = data.get('role')

            user = User.objects.get(id=user_id)
            profile = user.userprofile

            profile.role = new_role
            user.is_superuser = new_role == 'admin'
            user.is_staff = new_role in ['teacher', 'course-teacher', 'admin']

            profile.save()
            user.save()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@role_required('admin')
def update_group_view(request, user_id):
    if request.method == 'POST':
        data = json.loads(request.body)
        try:
            user = User.objects.get(id=user_id)
            profile = user.userprofile
            profile.experiment_day = data['experiment_day']
            profile.experiment_group = data['experiment_group']
            profile.save()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@role_required('admin')
def update_attendance_permission(request, user_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            allow = data.get('allow', False)
            user = User.objects.get(id=user_id)
            perm = Permission.objects.get(codename='view_attendancerecord')
            if allow:
                user.user_permissions.add(perm)
            else:
                user.user_permissions.remove(perm)
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
        
@role_required('admin')
def update_attendance_only(request, user_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body or '{}')
            enable = bool(data.get('enable'))
            user = User.objects.get(id=user_id)
            group, _ = Group.objects.get_or_create(name='attendance_only')
            view_perm = Permission.objects.get(codename='view_attendancerecord')
            change_perm = Permission.objects.get(codename='change_attendancerecord')
            group.permissions.add(view_perm, change_perm)
            if enable:
                user.groups.add(group)
            else:
                user.groups.remove(group)
                user.user_permissions.remove(change_perm)
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@role_required('admin')
def delete_user_view(request, user_id):
    if request.method == 'POST':
        try:
            user = User.objects.get(id=user_id)
            user.delete()
            return JsonResponse({'status': 'success'})
        except User.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'User not found'}, status=404)

@role_required('admin')
def create_user_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            offering_id = data.get('offering_id')
            if not offering_id:
                return JsonResponse({'status': 'error', 'message': 'offering_id is required'}, status=400)
            try:
                offering = CourseOffering.objects.get(id=offering_id)
            except CourseOffering.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'offering not found'}, status=400)

            user = User.objects.filter(username=data['email']).first()
            if not user:
                password = data.get('password') or '0000'
                user = User.objects.create_user(
                    username=data['email'],
                    email=data['email'],
                    password=password
                )
                profile = UserProfile.objects.create(
                    user=user,
                    full_name=data['full_name'],
                    email=data['email'],
                    student_id=data.get('student_id', '') or '',
                    experiment_day=data.get('experiment_day', ''),
                    experiment_group=data.get('experiment_group', ''),
                    role='student'
                )
            else:
                profile = getattr(user, 'userprofile', None)
                if not profile:
                    profile = UserProfile.objects.create(
                        user=user,
                        full_name=data.get('full_name', user.username),
                        email=user.email,
                        student_id=data.get('student_id', ''),
                        experiment_day=data.get('experiment_day', ''),
                        experiment_group=data.get('experiment_group', ''),
                        role='student'
                    )
            enr, created = Enrollment.objects.get_or_create(
                user=user,
                course_offering=offering,
                role='student',
                defaults={
                    'experiment_day': data.get('experiment_day', '') or profile.experiment_day,
                    'experiment_group': data.get('experiment_group', '') or profile.experiment_group,
                }
            )
            if not created:
                return JsonResponse({'status': 'error', 'message': 'このユーザは既に当該科目/年度に登録されています。'}, status=400)
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@role_required('admin')
def update_user_view(request, user_id):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=400)

    try:
        data = json.loads(request.body)

        user = User.objects.get(id=user_id)
        profile = user.userprofile

        new_email = data.get('email', '').strip()
        if not new_email:
            return JsonResponse({'status': 'error', 'message': 'メールアドレスは必須です。'}, status=400)

        # 他ユーザとの重複チェック
        if User.objects.filter(username=new_email).exclude(id=user_id).exists():
            return JsonResponse({'status': 'error', 'message': 'このメールアドレスは既に使用されています。'}, status=400)

        new_role = data.get('role', profile.role)

        user.username = new_email
        user.email = new_email
        user.is_superuser = new_role == 'admin'
        user.is_staff = new_role in ['teacher', 'course-teacher', 'admin']

        profile.email = new_email
        profile.full_name = data.get('full_name', profile.full_name)
        profile.student_id = data.get('student_id', profile.student_id)
        profile.experiment_day = data.get('experiment_day', profile.experiment_day)
        profile.experiment_group = data.get('experiment_group', profile.experiment_group)
        profile.role = new_role

        profile.save()
        user.save()

        # 既存の他ロールのEnrollmentは削除（ロールは単一）
        Enrollment.objects.filter(user=user).exclude(role=new_role).delete()

        offering_id = data.get('offering_id')
        if offering_id:
            try:
                offering = CourseOffering.objects.get(id=offering_id)
            except CourseOffering.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'offering not found'}, status=400)

            enrollment = Enrollment.objects.filter(user=user, role=new_role).first()
            if enrollment:
                enrollment.course_offering = offering
                enrollment.experiment_day = profile.experiment_day
                enrollment.experiment_group = profile.experiment_group
                enrollment.save()
            else:
                Enrollment.objects.create(
                    user=user,
                    course_offering=offering,
                    role=new_role,
                    experiment_day=profile.experiment_day,
                    experiment_group=profile.experiment_group,
                )
        else:
            # 科目・年度が未選択の場合は当該ロールの履修情報を削除する
            Enrollment.objects.filter(user=user, role=new_role).delete()

        return JsonResponse({'status': 'success'})
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@role_required('admin')
def bulk_create_users(request):
    """Create multiple users from uploaded CSV file.

    Expected CSV columns: 名前, メールアドレス, 学生番号, 曜日, 班
    Password will be set to 学生番号.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=400)

    csv_file = request.FILES.get('file')
    if not csv_file:
        return JsonResponse({'status': 'error', 'message': 'CSVファイルが必要です'}, status=400)

    offering_id = request.POST.get('offering_id')
    offering = None
    if offering_id:
        try:
            offering = CourseOffering.objects.get(id=offering_id)
        except CourseOffering.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'offering not found'}, status=400)

    created = 0
    skipped = 0
    duplicates = []
    try:
        decoded = csv_file.read().decode('utf-8-sig').splitlines()
        reader = csv.DictReader(decoded)
        required_fields = ['名前', 'メールアドレス', '学生番号', '曜日', '班']
        normalized_fieldnames = []
        if reader.fieldnames:
            normalized_fieldnames = [
                str(field or '').replace('\ufeff', '').strip()
                for field in reader.fieldnames
            ]
            reader.fieldnames = normalized_fieldnames
        if not normalized_fieldnames or any(f not in normalized_fieldnames for f in required_fields):
            return JsonResponse({
                'status': 'error',
                'message': 'CSVのカラムは「名前，メールアドレス，学生番号，曜日，班」にしてください'
            }, status=400)
        # 既存の当該科目/年度Enrollmentをキャッシュ（メールで判定／ロール問わず重複禁止）
        existing_emails = set()
        if offering:
            for enr in Enrollment.objects.filter(course_offering=offering).select_related('user__userprofile'):
                email_val = (enr.user.username or "").lower()
                if email_val:
                    existing_emails.add(email_val)
        for row in reader:
            normalized_row = {
                str(key or '').replace('\ufeff', '').strip(): value
                for key, value in row.items()
            }
            email = (normalized_row.get('メールアドレス') or '').strip()
            if not email:
                skipped += 1
                continue
            full_name = (normalized_row.get('名前', '') or '').strip()
            student_id_val = (normalized_row.get('学生番号', '') or '').strip()
            day_val = (normalized_row.get('曜日', '') or '').strip()
            group_val = (normalized_row.get('班', normalized_row.get('班番号', '')) or '').strip()
            email_lower = email.lower()
            # まず、選択中科目/年度で同メールのEnrollmentが既にある場合は重複扱い（ロール問わず）
            if offering and Enrollment.objects.filter(
                user__username__iexact=email,
                course_offering=offering,
            ).exists():
                duplicates.append({'名前': full_name, 'メールアドレス': email, '学生番号': student_id_val})
                skipped += 1
                existing_emails.add(email_lower)
                continue
            # 当該科目/年度で既に登録済みなら重複としてスキップ（メールのみ判定）
            if offering and (email_lower in existing_emails):
                duplicates.append({'名前': full_name, 'メールアドレス': email, '学生番号': student_id_val})
                skipped += 1
                continue
            if User.objects.filter(username=email).exists():
                if offering and not Enrollment.objects.filter(user__username=email, course_offering=offering).exists():
                    user = User.objects.get(username=email)
                    Enrollment.objects.create(
                        user=user,
                        course_offering=offering,
                        role='student',
                        experiment_day=day_val,
                        experiment_group=group_val,
                    )
                    created += 1
                    existing_emails.add(email_lower)
                else:
                    skipped += 1
                continue
            user = User.objects.create_user(
                username=email,
                email=email,
                password=student_id_val
            )
            profile = UserProfile.objects.create(
                user=user,
                full_name=full_name,
                email=email,
                student_id=student_id_val,
                experiment_day=day_val,
                experiment_group=group_val,
                role='student'
            )
            if offering:
                Enrollment.objects.get_or_create(
                    user=user,
                    course_offering=offering,
                    role='student',
                    defaults={
                        'experiment_day': profile.experiment_day,
                        'experiment_group': profile.experiment_group,
                    }
                )
            created += 1
            # 新規作成分も重複判定セットに追加
            existing_emails.add(email_lower)
        return JsonResponse({'status': 'success', 'created': created, 'skipped': skipped, 'duplicates': duplicates})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@role_required('admin')
def bulk_user_template_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="user_bulk_import_template.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(['名前', 'メールアドレス', '学生番号', '曜日', '班'])
    return response


@role_required('admin')
def group_assignment_builder(request):
    return render(request, 'submission/group_assignment_builder.html', {})


@role_required('admin')
def group_assignment_preview_api(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=400)

    participants_file = request.FILES.get('participants_file')
    survey_file = request.FILES.get('survey_file')
    roster_file = request.FILES.get('roster_file')
    grades_file = request.FILES.get('grades_file')
    existing_assignment_file = request.FILES.get('existing_assignment_file')
    target_grade_raw = request.POST.get('target_grade', '').strip()
    group_count_raw = request.POST.get('group_count', '').strip()
    ideal_group_size_raw = request.POST.get('ideal_group_size', '').strip()

    if not group_count_raw or not ideal_group_size_raw:
        return JsonResponse({'status': 'error', 'message': '班数と基本班人数は必須です。'}, status=400)
    constraint_profile = _ga_build_constraint_profile({
        'group_count': group_count_raw,
        'ideal_group_size': ideal_group_size_raw,
        'separate_repeaters': request.POST.get('separate_repeaters'),
        'forbid_two_females': request.POST.get('forbid_two_females'),
        'forbid_mixed_two_person_group': request.POST.get('forbid_mixed_two_person_group'),
        'use_liberal_arts_credits_priority': request.POST.get('use_liberal_arts_credits_priority'),
        'balance_gpa': request.POST.get('balance_gpa'),
    })

    if not participants_file or not survey_file:
        return JsonResponse({'status': 'error', 'message': '履修予定者ファイルとGoogle Form回答一覧ファイルは必須です。'}, status=400)

    needs_roster = (
        constraint_profile['separate_repeaters']
        or constraint_profile['forbid_two_females']
        or constraint_profile['forbid_mixed_two_person_group']
    )
    needs_grades = (
        constraint_profile['use_liberal_arts_credits_priority']
        or constraint_profile['balance_gpa']
    )
    if needs_roster and not roster_file:
        return JsonResponse({'status': 'error', 'message': '選択した制約のため名簿ファイルが必要です。'}, status=400)
    if needs_grades and not grades_file:
        return JsonResponse({'status': 'error', 'message': '選択した制約のため成績ファイルが必要です。'}, status=400)

    target_grade = None
    if constraint_profile['separate_repeaters']:
        if not target_grade_raw:
            return JsonResponse({'status': 'error', 'message': '再履修生分離を使う場合は対象学年が必要です。'}, status=400)
        try:
            target_grade = int(target_grade_raw)
        except ValueError:
            return JsonResponse({'status': 'error', 'message': '対象学年は数値で指定してください。'}, status=400)

    try:
        preview = _ga_build_preview_data(
            participants_file=participants_file,
            survey_file=survey_file,
            roster_file=roster_file,
            grades_file=grades_file,
            target_grade=target_grade,
            existing_assignment_file=existing_assignment_file,
            constraint_profile=constraint_profile,
        )
        request.session[GROUP_ASSIGNMENT_SESSION_KEY] = preview
        request.session.modified = True
        return JsonResponse({
            'status': 'success',
            'preview': preview,
            'downloads': {
                'csv': reverse('group_assignment_download_csv'),
                'pdf': reverse('group_assignment_download_pdf'),
            }
        })
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)


@role_required('admin')
def group_assignment_finalize_api(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=400)
    try:
        preview = _ga_preview_from_session(request)
        preview['approved'] = True
        request.session[GROUP_ASSIGNMENT_SESSION_KEY] = preview
        request.session.modified = True
        return JsonResponse({
            'status': 'success',
            'downloads': {
                'csv': reverse('group_assignment_download_csv'),
                'pdf': reverse('group_assignment_download_pdf'),
            }
        })
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)


@role_required('admin')
def group_assignment_download_csv(request):
    try:
        preview = _ga_preview_from_session(request)
        if not preview.get('approved'):
            return JsonResponse({'status': 'error', 'message': '承認後にダウンロードしてください。'}, status=400)
        return _ga_generate_bulk_csv_response(preview)
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)


@role_required('admin')
def group_assignment_download_pdf(request):
    try:
        preview = _ga_preview_from_session(request)
        if not preview.get('approved'):
            return JsonResponse({'status': 'error', 'message': '承認後にダウンロードしてください。'}, status=400)
        return _ga_generate_pdf_response(preview)
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)


@role_required('admin')
def upload_student_photo(request, student_id):
    """Receive uploaded photo and save to UserProfile"""
    if request.method == 'POST':
        try:
            profile = UserProfile.objects.get(id=student_id)
            photo = request.FILES.get('photo')
            if not photo:
                return JsonResponse({'status': 'error', 'message': 'photo required'}, status=400)
            filename = f"{profile.student_id}_{profile.full_name}.png"
            path = default_storage.save(f"student_photos/{filename}", photo)
            profile.photo.name = path
            profile.save()
            return JsonResponse({'status': 'success', 'photo_url': profile.photo.url})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'POST required'}, status=400)

@role_required('admin')
def final_score_list_view(request):
    offerings = list(CourseOffering.objects.select_related('course'))
    offering_options = [
        {
            'id': o.id,
            'course_id': o.course_id,
            'course_code': o.course.code,
            'course_name': o.course.name,
            'year': o.year,
        }
        for o in offerings
    ]
    default_offering_id = None
    if offerings:
        latest = max(offerings, key=lambda o: (o.year, o.id))
        default_offering_id = latest.id

    offering_id = request.GET.get('offering_id') or default_offering_id
    experiment_numbers = [n[0] for n in Submission.EXPERIMENT_NUMBER_CHOICES]
    if offering_id:
        try:
            off = CourseOffering.objects.select_related('course').get(id=offering_id)
            experiment_numbers = off.course.experiment_numbers or experiment_numbers
        except CourseOffering.DoesNotExist:
            pass
    student_data = _build_final_score_rows(experiment_numbers, offering_id)
    context = {
        'students_json': json.dumps(student_data, ensure_ascii=False),
        'students': student_data,
        'experiment_numbers': json.dumps(experiment_numbers, ensure_ascii=False),
        'offerings_json': json.dumps(offering_options, ensure_ascii=False),
        'default_offering_id': default_offering_id,
    }
    return render(request, 'submission/final_score_list.html', context)


@role_required('admin')
def final_score_list_csv(request):
    """Download final scores as CSV (現在の表示条件に合わせる)."""
    offering_id = request.GET.get('offering_id')
    day = request.GET.get('day') or None
    group = request.GET.get('group') or None
    experiment_numbers = [n[0] for n in Submission.EXPERIMENT_NUMBER_CHOICES]
    if offering_id:
        try:
            off = CourseOffering.objects.select_related('course').get(id=offering_id)
            experiment_numbers = off.course.experiment_numbers or experiment_numbers
        except CourseOffering.DoesNotExist:
            pass
    student_data = _build_final_score_rows(experiment_numbers, offering_id, day=day, group=group)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="final_scores.csv"'

    # Add BOM for Excel compatibility
    response.write('\ufeff')
    writer = csv.writer(response)
    header = ['名前', '学生番号', '曜日', '班番号'] + experiment_numbers + ['遅刻回数', '欠席回数', '減点', 'ディスカッション', '実施項目数', '最終成績']
    writer.writerow(header)

    for row_data in student_data:
        row = [
            row_data['name'],
            row_data['student_id'],
            row_data['experiment_day'],
            row_data['experiment_group'],
        ]
        for ex in experiment_numbers:
            row.append(row_data.get(ex, ''))
        row.append(row_data.get('late_count', 0))
        row.append(row_data.get('absence_count', 0))
        row.append(row_data.get('score_details_total', ''))
        row.append(row_data.get('discussion_count_total', 0))
        row.append(row_data.get('completed_task_count', 0))
        row.append(row_data.get('final_grade', ''))
        writer.writerow(row)

    return response


def _build_final_score_rows(experiment_numbers, offering_id, day=None, group=None):
    # 実験番号の順序を維持して重複を除外
    unique_experiment_numbers = []
    seen_experiment_numbers = set()
    for ex in experiment_numbers:
        if ex in seen_experiment_numbers:
            continue
        unique_experiment_numbers.append(ex)
        seen_experiment_numbers.add(ex)

    students_qs = UserProfile.objects.filter(role='student').select_related('user')
    enrollment_map = {}
    if offering_id:
        enr_qs = Enrollment.objects.filter(
            course_offering_id=offering_id,
            role='student',
        )
        if day:
            enr_qs = enr_qs.filter(experiment_day=day)
        if group:
            enr_qs = enr_qs.filter(experiment_group=group)
        enrollment_map = {
            e.user_id: e
            for e in enr_qs.only('user_id', 'experiment_day', 'experiment_group')
        }
        students_qs = students_qs.filter(user_id__in=enrollment_map.keys())
    else:
        if day:
            students_qs = students_qs.filter(experiment_day=day)
        if group:
            students_qs = students_qs.filter(experiment_group=group)

    absence_penalty_weight = _absence_penalty_weight(offering_id)
    discussion_bonus_weight = _discussion_bonus_weight(offering_id)
    schedule_by_day = {}
    attendance_map = {}
    late_attendance_map = {}
    if offering_id and students_qs.exists():
        now_local = timezone.localtime(timezone.now(), JST)
        cutoff_date = now_local.date()
        if now_local.time() < ABSENCE_CUTOFF_TIME:
            cutoff_date = cutoff_date - timedelta(days=1)
        schedule_qs = Schedule.objects.filter(
            course_offering_id=offering_id,
            date__lte=cutoff_date
        )
        schedule_dates_all = set()
        for sched in schedule_qs:
            label = _weekday_label(sched.date)
            schedule_by_day.setdefault(label, set()).add(sched.date)
            schedule_dates_all.add(sched.date)

        if schedule_dates_all:
            attendance_qs = AttendanceRecord.objects.filter(
                course_offering_id=offering_id,
                date__in=schedule_dates_all,
                user_id__in=list(enrollment_map.keys()) if enrollment_map else list(students_qs.values_list('user_id', flat=True))
            ).values_list('user_id', 'date', 'check_in')
            for user_id, att_date, check_in in attendance_qs:
                attendance_map.setdefault(user_id, set()).add(att_date)
                if check_in and timezone.localtime(check_in, JST).time() > LATE_CHECKIN_TIME:
                    late_attendance_map.setdefault(user_id, set()).add(att_date)
    student_data = []
    experiment_count = len(unique_experiment_numbers) if unique_experiment_numbers else 0

    # 実施項目集計用の事前ロード（選択科目/年度かつ対象実験のみ）
    progress_map = {}
    completion_map = {}
    task_config_map = {}
    discussion_count_map = {}
    if offering_id:
        target_user_ids = list(students_qs.values_list('user_id', flat=True))
        progress_qs = ExperimentProgress.objects.filter(
            course_offering_id=offering_id,
            student_id__in=target_user_ids,
            experiment_number__in=unique_experiment_numbers,
        ).values_list('student_id', 'experiment_number', 'task_no')
        for student_id, exp_no, task_no in progress_qs:
            progress_map.setdefault((student_id, exp_no), set()).add(str(task_no))

        completion_qs = ExperimentCompletion.objects.filter(
            course_offering_id=offering_id,
            student_id__in=target_user_ids,
            experiment_number__in=unique_experiment_numbers,
        ).values_list('student_id', 'experiment_number', 'completed')
        for student_id, exp_no, completed in completion_qs:
            completion_map[(student_id, exp_no)] = bool(completed)

        task_cfg_qs = ExperimentTaskConfig.objects.filter(
            course_offering_id=offering_id,
            experiment_number__in=unique_experiment_numbers,
        ).values_list('experiment_number', 'task_list')
        for exp_no, task_list in task_cfg_qs:
            task_config_map[exp_no] = _normalize_task_list(task_list)

        discussion_qs = DiscussionBonus.objects.filter(
            course_offering_id=offering_id,
            student_id__in=target_user_ids,
            experiment_number__in=unique_experiment_numbers,
        ).values_list('student_id', 'experiment_number', 'count')
        for student_id, exp_no, count in discussion_qs:
            discussion_count_map[(student_id, exp_no)] = int(count or 0)

    for up in students_qs:
        enr = enrollment_map.get(up.user_id)
        record = {
            'user_profile_id': up.id,
            'user_id': up.user_id,
            'name': up.full_name,
            'student_id': up.student_id,
            'experiment_day': enr.experiment_day if enr else up.experiment_day,
            'experiment_group': enr.experiment_group if enr else up.experiment_group,
        }
        total_final_score = 0.0
        score_details_total = 0.0
        completed_task_count = 0
        discussion_count_total = 0
        experiment_logs = []
        for ex in unique_experiment_numbers:
            sub_qs = Submission.objects.filter(
                student=up.user,
                experiment_number=ex,
                report_type='main',
                final_evaluated=True,
                accepted=True,
            )
            if offering_id:
                sub_qs = sub_qs.filter(course_offering_id=offering_id)
            sub = sub_qs.order_by('-submitted_at').first()
            record[ex] = float(sub.final_score) if sub and sub.final_score is not None else ''
            if sub and sub.final_score is not None:
                try:
                    total_final_score += float(sub.final_score)
                except Exception:
                    pass

            details_qs = Submission.objects.filter(
                student=up.user,
                experiment_number=ex,
                score_details__isnull=False,
            )
            if offering_id:
                details_qs = details_qs.filter(course_offering_id=offering_id)
            for s in details_qs:
                for item in (s.score_details or []):
                    if not isinstance(item, dict):
                        continue
                    value = item.get('value', 0)
                    weight = item.get('weight', 1)
                    try:
                        score_details_total += float(value) * float(weight)
                    except Exception:
                        continue

            completed_set = progress_map.get((up.user_id, ex), set())
            task_list = task_config_map.get(ex, [])
            ordered_done = [task for task in task_list if task in completed_set]
            ordered_done.extend(sorted(completed_set - set(task_list)))
            completed_task_count += len(completed_set)
            if task_list:
                status = '完了' if set(task_list).issubset(completed_set) else '未完了'
            else:
                status = '完了' if completion_map.get((up.user_id, ex), False) or completed_set else '未完了'
            experiment_logs.append({
                'experiment_number': ex,
                'status': status,
                'completed_tasks': ', '.join(ordered_done) if ordered_done else '-',
            })
            discussion_count_total += discussion_count_map.get((up.user_id, ex), 0)

        absence_count = 0
        late_count = 0
        if offering_id:
            day_label = record['experiment_day']
            target_dates = schedule_by_day.get(day_label, set())
            if target_dates:
                attended_dates = attendance_map.get(up.user_id, set())
                absence_count = len(target_dates - attended_dates)
                late_dates = late_attendance_map.get(up.user_id, set())
                late_count = len(target_dates & late_dates)
        absence_penalty = round(absence_count * absence_penalty_weight, 2)
        discussion_bonus_total = round(discussion_count_total * discussion_bonus_weight, 2)
        score_details_total = round(abs(score_details_total), 2)
        absence_penalty = round(absence_penalty, 2)
        total_final_score = round(total_final_score, 2)

        final_grade = ''
        score_details_avg = 0.0
        grade_divisor = 10
        if grade_divisor:
            score_details_avg = score_details_total / grade_divisor
            final_grade = (total_final_score - score_details_avg - absence_penalty + discussion_bonus_total) / grade_divisor
            final_grade = round(final_grade, 2)
            score_details_avg = round(score_details_avg, 2)

        record['late_count'] = late_count
        record['absence_count'] = absence_count
        record['score_details_total'] = score_details_total
        record['absence_penalty'] = absence_penalty
        record['discussion_count_total'] = discussion_count_total
        record['discussion_bonus_total'] = discussion_bonus_total
        record['final_score_total'] = total_final_score
        record['score_details_avg'] = score_details_avg
        record['experiment_count'] = experiment_count
        record['grade_divisor'] = grade_divisor
        record['completed_task_count'] = completed_task_count
        record['experiment_logs'] = experiment_logs
        record['final_grade'] = final_grade
        student_data.append(record)
    return student_data


@role_required('admin')
def final_score_data_api(request):
    offering_id = request.GET.get('offering_id')
    experiment_numbers = [n[0] for n in Submission.EXPERIMENT_NUMBER_CHOICES]
    if offering_id:
        try:
            off = CourseOffering.objects.select_related('course').get(id=offering_id)
            experiment_numbers = off.course.experiment_numbers or experiment_numbers
        except CourseOffering.DoesNotExist:
            pass
    data = _build_final_score_rows(experiment_numbers, offering_id)
    return JsonResponse({'students': data, 'experiment_numbers': experiment_numbers})


@role_required('admin')
def final_score_detail_api(request):
    offering_id = request.GET.get('offering_id')
    user_profile_id = request.GET.get('user_profile_id')
    experiment_number = request.GET.get('experiment_number')
    if not (offering_id and user_profile_id and experiment_number):
        return JsonResponse({'status': 'error', 'message': 'offering_id, user_profile_id, experiment_number are required'}, status=400)
    try:
        offering = CourseOffering.objects.select_related('course').get(id=offering_id)
        up = UserProfile.objects.select_related('user').get(id=user_profile_id, role='student')
    except (CourseOffering.DoesNotExist, UserProfile.DoesNotExist):
        return JsonResponse({'status': 'error', 'message': 'not found'}, status=404)

    if not Enrollment.objects.filter(user=up.user, course_offering=offering).exists():
        return JsonResponse({'status': 'error', 'message': 'not enrolled'}, status=403)
    enr = Enrollment.objects.filter(user=up.user, course_offering=offering, role='student').first() or Enrollment.objects.filter(user=up.user, course_offering=offering).first()

    subs = (
        Submission.objects
        .filter(student=up.user, course_offering=offering, experiment_number=experiment_number)
        .order_by('submitted_at')
    )

    def _to_float(x, default=0.0):
        try:
            return float(x)
        except Exception:
            return default

    submissions = []
    for s in subs:
        details = s.score_details or []
        total = 0.0
        normalized = []
        if isinstance(details, list):
            for item in details:
                if not isinstance(item, dict):
                    continue
                label = item.get('label') or item.get('key') or ''
                value = _to_float(item.get('value', 0))
                weight = _to_float(item.get('weight', 1), default=1.0)
                subtotal = value * weight
                total += subtotal
                normalized.append({
                    'label': label,
                    'value': value,
                    'weight': weight,
                    'subtotal': subtotal,
                })
        submissions.append({
            'id': s.id,
            'report_type': s.report_type,
            'submitted_at': timezone.localtime(s.submitted_at).strftime('%Y-%m-%d %H:%M'),
            'total_score': total,
            'final_score': float(s.final_score) if s.final_score is not None else None,
            'final_evaluated': bool(s.final_evaluated),
            'details': normalized,
        })

    return JsonResponse({
        'status': 'success',
        'student': {
            'name': up.full_name,
            'student_id': up.student_id,
            'experiment_day': enr.experiment_day if enr else getattr(up, 'experiment_day', ''),
            'experiment_group': enr.experiment_group if enr else getattr(up, 'experiment_group', ''),
        },
        'course': {
            'course_code': offering.course.code,
            'course_name': offering.course.name,
            'year': offering.year,
        },
        'experiment_number': experiment_number,
        'submissions': submissions,
    })


@role_required('admin')
def download_accepted_reports(request):
    """Download all accepted reports grouped by experiment number as a zip."""
    experiment_numbers = [n[0] for n in Submission.EXPERIMENT_NUMBER_CHOICES]
    offering_id = request.GET.get('offering_id')
    day = request.GET.get('day') or None
    group = request.GET.get('group') or None

    memfile = io.BytesIO()
    with zipfile.ZipFile(memfile, 'w') as zf:
        for ex in experiment_numbers:
            submissions = Submission.objects.filter(experiment_number=ex, accepted=True)
            if offering_id:
                submissions = submissions.filter(course_offering_id=offering_id)
            if day:
                submissions = submissions.filter(student__userprofile__experiment_day=day)
            if group:
                submissions = submissions.filter(student__userprofile__experiment_group=group)
            for sub in submissions:
                if sub.file and default_storage.exists(sub.file.name):
                    filename_raw = os.path.basename(sub.file.name)
                    filename = unquote(filename_raw)
                    student_id = getattr(sub.student.userprofile, 'student_id', sub.student.username)
                    arcname = f"{ex}/{student_id}_{filename}"
                    with default_storage.open(sub.file.name, 'rb') as f:
                        zf.writestr(arcname, f.read())

    memfile.seek(0)
    response = HttpResponse(memfile.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="accepted_reports.zip"'
    return response
