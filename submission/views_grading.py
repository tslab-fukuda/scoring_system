from django.shortcuts import render, get_object_or_404, redirect
from submission.models import UserProfile, Submission, Schedule
from django.contrib.auth.decorators import login_required
from submission.decorators import role_required
from submission.models import (
    ScoringItem,
    CourseOffering,
    Enrollment,
    SubmissionTextIndex,
    SimilarityJob,
    ExperimentProgress,
    ExperimentTaskConfig,
    FinalRubric,
)
from submission.models import Stamp
from django.http import JsonResponse, HttpResponse
from django.http import FileResponse, Http404
from django.conf import settings
from django.core.files import File
from django.core.files.storage import default_storage
from django.utils import timezone
from datetime import timedelta

import os
import io
import json
import base64
import tempfile
import fitz  # PyMuPDF
import re
import unicodedata
import hashlib
from functools import lru_cache
from difflib import SequenceMatcher
from decimal import Decimal
from PIL import Image
from openpyxl import Workbook, load_workbook

from submission.final_rubric_utils import (
    build_rubric_editor_payload,
    build_rubric_state_for_submission,
    get_copy_source_candidates,
    get_copy_source_rubric,
    get_active_default_final_rubric,
    get_active_final_rubric,
    get_effective_final_rubric,
    save_new_rubric_version,
    save_rubric_score,
    serialize_rubric_definition,
)


RUBRIC_ACCESS_ROLES = ['teacher', 'course-teacher', 'non-editing teacher']
GRADING_ACCESS_ROLES = ['teacher', 'course-teacher', 'non-editing teacher']


def _normalized_graded_stem(current_file_name, submission_id):
    stem = os.path.splitext(os.path.basename(current_file_name or ''))[0]
    patterns = [
        rf"_{submission_id}_graded(?:_[A-Za-z0-9]+)?$",
        r"_graded(?:_[A-Za-z0-9]+)?$",
    ]
    while True:
        new_stem = stem
        for pattern in patterns:
            new_stem = re.sub(pattern, '', new_stem)
        if new_stem == stem:
            break
        stem = new_stem
    return stem or f"submission_{submission_id}"


def _get_rubric_accessible_offerings(user):
    if not hasattr(user, 'userprofile'):
        return CourseOffering.objects.none()
    if user.userprofile.role == 'admin':
        return CourseOffering.objects.select_related('course').all()
    course_ids = (
        Enrollment.objects
        .filter(user=user, role__in=RUBRIC_ACCESS_ROLES)
        .values_list('course_offering__course_id', flat=True)
        .distinct()
    )
    return (
        CourseOffering.objects
        .filter(course_id__in=course_ids)
        .select_related('course')
        .distinct()
    )


def _get_grading_accessible_offerings(user):
    if not hasattr(user, 'userprofile'):
        return CourseOffering.objects.none()
    if user.userprofile.role == 'admin':
        return CourseOffering.objects.select_related('course').all()
    course_ids = (
        Enrollment.objects
        .filter(user=user, role__in=GRADING_ACCESS_ROLES)
        .values_list('course_offering__course_id', flat=True)
        .distinct()
    )
    return (
        CourseOffering.objects
        .filter(course_id__in=course_ids)
        .select_related('course')
        .distinct()
    )


def _resolve_rubric_offering(user, offering_id):
    offerings = list(_get_rubric_accessible_offerings(user))
    if not offerings:
        return None, offerings, JsonResponse(
            {'status': 'error', 'message': 'アクセス可能な科目/年度がありません'},
            status=404,
        )
    allowed_ids = {off.id for off in offerings}
    selected_id = None
    if offering_id:
        try:
            candidate = int(offering_id)
        except (TypeError, ValueError):
            candidate = None
        if candidate and candidate not in allowed_ids:
            return None, offerings, JsonResponse(
                {'status': 'error', 'message': '対象の科目/年度にはアクセスできません'},
                status=403,
            )
        selected_id = candidate
    if not selected_id:
        latest = max(offerings, key=lambda off: (off.year, off.id))
        selected_id = latest.id
    return selected_id, offerings, None


def _serialize_rubric_offerings(offerings):
    return [
        {
            'id': off.id,
            'course_id': off.course_id,
            'course_code': off.course.code,
            'course_name': off.course.name,
            'year': off.year,
            'meeting_days': off.course.meeting_days,
            'experiment_numbers': off.course.experiment_numbers,
        }
        for off in offerings
    ]


def _normalize_experiment_number(offering, experiment_number):
    if not offering:
        return ''
    experiment_numbers = [str(item).strip() for item in (offering.course.experiment_numbers or []) if str(item).strip()]
    requested = str(experiment_number or '').strip()
    if requested and requested in experiment_numbers:
        return requested
    return experiment_numbers[0] if experiment_numbers else ''


def _normalize_rubric_scope(scope):
    if scope == FinalRubric.SCOPE_DEFAULT:
        return FinalRubric.SCOPE_DEFAULT
    return FinalRubric.SCOPE_OFFERING


def _rubric_scope_label(scope):
    return 'デフォルト' if scope == FinalRubric.SCOPE_DEFAULT else '年度個別'


def _build_rubric_export_workbook(offering, experiment_number, scope, payload, active_version=None):
    workbook = Workbook()
    rubric_sheet = workbook.active
    rubric_sheet.title = 'rubric'
    criteria = payload.get('criteria') or []
    max_option_count = max((len(criterion.get('options') or []) for criterion in criteria), default=0)
    headers = ['クライテリア', '満点']
    for option_index in range(max_option_count):
        headers.extend([
            f'評価{option_index + 1}ラベル',
            f'評価{option_index + 1}説明',
        ])
    for col_index, header in enumerate(headers, start=1):
        rubric_sheet.cell(row=1, column=col_index, value=header)

    row_index = 2
    for criterion in criteria:
        rubric_sheet.cell(row=row_index, column=1, value=criterion.get('title', ''))
        rubric_sheet.cell(row=row_index, column=2, value=criterion.get('max_points', 0))
        column_index = 3
        for option in (criterion.get('options') or []):
            rubric_sheet.cell(row=row_index, column=column_index, value=option.get('label', ''))
            rubric_sheet.cell(row=row_index, column=column_index + 1, value=option.get('description', ''))
            column_index += 2
        row_index += 1
    return workbook


def _parse_rubric_import_workbook(uploaded_file):
    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook['rubric'] if 'rubric' in workbook.sheetnames else workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [str(cell or '').strip() for cell in (header_row or [])]

    if len(headers) >= 2 and headers[0] == 'クライテリア' and headers[1] == '満点':
        criteria = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=0):
            if not row or all(value in (None, '') for value in row):
                continue
            title = str(row[0] or '').strip()
            if not title:
                raise ValueError('クライテリア が空の行があります')
            try:
                max_points = int(row[1] or 0)
            except (TypeError, ValueError):
                raise ValueError('満点 は整数で入力してください')
            options = []
            option_order = 0
            for column_index in range(2, len(headers), 2):
                label = str(row[column_index] or '').strip() if column_index < len(row) else ''
                description = str(row[column_index + 1] or '').strip() if column_index + 1 < len(row) else ''
                if not label and not description:
                    continue
                if not label:
                    raise ValueError(f'{title} の評価{option_order + 1}ラベルが空です')
                options.append({
                    'source_option_id': None,
                    'label': label,
                    'description': description,
                    'order': option_order,
                })
                option_order += 1
            if not options:
                raise ValueError(f'{title} に選択肢がありません')
            criteria.append({
                'source_criterion_id': None,
                'title': title,
                'max_points': max_points,
                'order': row_index,
                'options': options,
            })
        if not criteria:
            raise ValueError('rubric シートにクライテリアがありません')
        return {'criteria': criteria}

    if 'criteria' in workbook.sheetnames:
        criteria_sheet = workbook['criteria']
        header_row = next(criteria_sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(cell or '').strip() for cell in (header_row or [])]
        expected = [
            'criterion_order',
            'criterion_title',
            'max_points',
            'option_order',
            'option_label',
            'option_description',
        ]
        if headers[:len(expected)] != expected:
            raise ValueError('Excel のヘッダ形式が不正です')

        criterion_map = {}
        for row in criteria_sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(value in (None, '') for value in row):
                continue
            try:
                criterion_order = int(row[0] or 0)
                max_points = int(row[2] or 0)
                option_order = int(row[3] or 0)
            except (TypeError, ValueError):
                raise ValueError('criterion_order / max_points / option_order は整数で入力してください')
            title = str(row[1] or '').strip()
            label = str(row[4] or '').strip()
            description = str(row[5] or '').strip()
            if not title:
                raise ValueError('criterion_title が空の行があります')
            if not label:
                raise ValueError('option_label が空の行があります')
            criterion = criterion_map.setdefault(criterion_order, {
                'source_criterion_id': None,
                'title': title,
                'max_points': max_points,
                'order': criterion_order,
                'options': [],
            })
            criterion['title'] = title
            criterion['max_points'] = max_points
            criterion['options'].append({
                'source_option_id': None,
                'label': label,
                'description': description,
                'order': option_order,
            })

        if not criterion_map:
            raise ValueError('criteria シートにクライテリアがありません')

        criteria = []
        for criterion_order in sorted(criterion_map.keys()):
            criterion = criterion_map[criterion_order]
            criterion['options'] = sorted(criterion['options'], key=lambda option: option['order'])
            criteria.append(criterion)
        return {'criteria': criteria}

    raise ValueError('rubric シートがありません')

@login_required
@role_required('teacher','admin','course-teacher','non-editing teacher')
def grading_form(request, submission_id):
    submission = get_object_or_404(Submission, pk=submission_id)
    if request.method == 'POST':
        data = json.loads(request.body)
        images = data.get('drawImages')
        old_file_name = submission.file.name
        pdf_path = submission.file.path

        # PyMuPDFでPDF編集
        doc = fitz.open(pdf_path)
        for page_no, img_data in enumerate(images):
            if not img_data:
                continue
            page = doc[page_no]
            header, encoded = img_data.split(",", 1)
            hand_img_bytes = base64.b64decode(encoded)
            # PNGのアルファを保持したまま挿入する
            img_doc = fitz.open("png", hand_img_bytes)
            pix = img_doc[0].get_pixmap(alpha=True)
            page.insert_image(page.rect, pixmap=pix, overlay=True)
            img_doc.close()

        tmp_path = None
        saved_name = None
        try:
            clean_stem = _normalized_graded_stem(old_file_name, submission.id)
            desired_name = f"submissions/{clean_stem}_{submission.id}_graded.pdf"

            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                tmp_path = tmp.name
            doc.save(tmp_path)
            doc.close()

            if default_storage.exists(desired_name):
                default_storage.delete(desired_name)

            with open(tmp_path, 'rb') as graded_fp:
                saved_name = default_storage.save(desired_name, File(graded_fp))

            submission.file.name = saved_name
            submission.graded = True
            submission.score_details = data.get('scoreItems')
            submission.save()
            # デバッグ用にURLをログ出力
            print(f"[graded_pdf] saved: {submission.file.url}")

            # 旧ファイルは保存成功後に削除する
            if old_file_name and old_file_name != saved_name and default_storage.exists(old_file_name):
                default_storage.delete(old_file_name)
        except Exception:
            if saved_name and default_storage.exists(saved_name):
                default_storage.delete(saved_name)
            raise
        finally:
            try:
                doc.close()
            except Exception:
                pass
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)

        return JsonResponse({'status': 'ok', 'new_file_url': submission.file.url})

    # GET時
    score_json = json.dumps(submission.score_details) if submission.score_details else 'null'
    previous = Submission.objects.filter(
        student=submission.student,
        experiment_number=submission.experiment_number,
        report_type=submission.report_type,
        course_offering=submission.course_offering,
        submitted_at__lt=submission.submitted_at
    ).order_by('-submitted_at').first()
    compare_pdf_url = previous.file.url if previous and previous.file else ''
    compare_submitted_at = ''
    if previous and previous.submitted_at:
        compare_submitted_at = timezone.localtime(previous.submitted_at).strftime('%Y-%m-%d %H:%M')
    return render(request, 'submission/grading_form.html', {
        'submission': submission,
        'pdf_url': submission.file.url,
        'score_details': score_json,
        'compare_pdf_url': compare_pdf_url,
        'compare_submitted_at': compare_submitted_at,
    })


@login_required
@role_required('teacher', 'admin', 'course-teacher', 'non-editing teacher')
def graded_pdf(request, submission_id):
    """Ensure graded PDF is served with correct content-type for preview iframe."""
    submission = get_object_or_404(Submission, pk=submission_id)
    if not submission.file:
        raise Http404("file not found")
    try:
        resp = FileResponse(submission.file.open('rb'), content_type='application/pdf')
        resp['Content-Disposition'] = f'inline; filename="{os.path.basename(submission.file.name)}"'
        return resp
    except FileNotFoundError:
        raise Http404("file not found")

@login_required
@role_required('teacher', 'admin', 'course-teacher', 'non-editing teacher')
def scoring_items_api(request):
    offering_id = request.GET.get('offering_id')
    offering_id_int = None
    if offering_id:
        try:
            offering_id_int = int(offering_id)
        except (TypeError, ValueError):
            offering_id_int = None
    if not offering_id_int:
        return JsonResponse({'pre': [], 'main': []})

    accessible_offerings = _get_grading_accessible_offerings(request.user)
    offering = accessible_offerings.filter(id=offering_id_int).first()
    if not offering:
        return JsonResponse({'pre': [], 'main': []})

    def _merge_items(common_items, specific_items):
        merged = []
        index_by_label = {}
        for item in common_items:
            key = item.get('code') or item.get('label') or ''
            if not key or key in index_by_label:
                continue
            index_by_label[key] = len(merged)
            merged.append(item)
        for item in specific_items:
            key = item.get('code') or item.get('label') or ''
            if not key:
                continue
            if key in index_by_label:
                merged[index_by_label[key]] = item
            else:
                index_by_label[key] = len(merged)
                merged.append(item)
        return merged

    def _items(category):
        common_qs = ScoringItem.objects.filter(
            category=category,
            course_id=offering.course_id,
            course_offering__isnull=True
        ).order_by('order')
        common_items = list(common_qs.values('label', 'weight', 'code', 'show_in_grading_form'))
        specific_qs = ScoringItem.objects.filter(
            category=category,
            course_offering_id=offering_id_int
        ).order_by('order')
        specific_items = list(specific_qs.values('label', 'weight', 'code', 'show_in_grading_form'))
        return _merge_items(common_items, specific_items)

    pre = _items('pre')
    main = _items('main')
    for x in pre:
        x['weight'] = int(x['weight'])
        x['code'] = x.get('code') or ''
    for x in main:
        x['weight'] = int(x['weight'])
        x['code'] = x.get('code') or ''
    return JsonResponse({'pre': pre, 'main': main})

@login_required
@role_required('teacher', 'admin', 'course-teacher', 'non-editing teacher')
def stamps_api(request):
    stamps = list(Stamp.objects.all().values('id','text'))
    return JsonResponse({'stamps': stamps})


@login_required
@role_required('non-editing teacher', 'admin')
def final_rubric_settings(request):
    default_offering_id, offerings, error_response = _resolve_rubric_offering(
        request.user,
        request.GET.get('offering_id'),
    )
    if error_response:
        return error_response
    offering_map = {off.id: off for off in offerings}
    selected_offering = offering_map.get(default_offering_id)
    selected_experiment_number = _normalize_experiment_number(
        selected_offering,
        request.GET.get('experiment_number'),
    )
    initial_scope = FinalRubric.SCOPE_OFFERING
    if selected_offering:
        if get_active_default_final_rubric(selected_offering):
            initial_scope = FinalRubric.SCOPE_DEFAULT
    return render(request, 'submission/final_rubric_settings.html', {
        'offerings_json': json.dumps(_serialize_rubric_offerings(offerings), ensure_ascii=False),
        'default_offering_id': default_offering_id,
        'initial_experiment_number_json': json.dumps(selected_experiment_number, ensure_ascii=False),
        'initial_scope_json': json.dumps(initial_scope, ensure_ascii=False),
        'can_edit_default_json': json.dumps(request.user.userprofile.role == 'admin'),
    })


@login_required
@role_required('non-editing teacher', 'admin')
def final_rubric_definition_api(request):
    if request.method == 'GET':
        default_offering_id, offerings, error_response = _resolve_rubric_offering(
            request.user,
            request.GET.get('offering_id'),
        )
        if error_response:
            return error_response
        offering = next((off for off in offerings if off.id == default_offering_id), None)
        scope = _normalize_rubric_scope(request.GET.get('scope'))
        experiment_number = _normalize_experiment_number(offering, request.GET.get('experiment_number'))
        if scope != FinalRubric.SCOPE_DEFAULT and not experiment_number:
            return JsonResponse({
                'status': 'error',
                'message': '対象科目に実験項目が設定されていません',
            }, status=400)
        source_rubric_id = request.GET.get('source_rubric_id')
        if source_rubric_id:
            loaded_from = get_copy_source_rubric(offering, scope, source_rubric_id)
            if loaded_from is None:
                return JsonResponse({
                    'status': 'error',
                    'message': '継承元の評価基準が見つかりません',
                }, status=404)
            payload = serialize_rubric_definition(loaded_from)
            active_rubric = get_active_final_rubric(offering, experiment_number, scope)
        else:
            editor_state = build_rubric_editor_payload(offering, experiment_number, scope)
            payload = editor_state['payload']
            active_rubric = editor_state['active_rubric']
            loaded_from = editor_state['loaded_from']
        return JsonResponse({
            'status': 'ok',
            'rubric_payload': payload,
            'active_version': active_rubric.version if active_rubric else None,
            'loaded_from': {
                'id': loaded_from.id,
                'scope': loaded_from.scope,
                'scope_label': loaded_from.get_scope_display(),
                'year': loaded_from.course_offering.year if loaded_from.course_offering else None,
                'experiment_number': loaded_from.experiment_number,
                'version': loaded_from.version,
            } if loaded_from else None,
            'copy_candidates': get_copy_source_candidates(offering, experiment_number, scope),
        })

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'method not allowed'}, status=405)

    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': '不正なリクエストです'}, status=400)

    default_offering_id, offerings, error_response = _resolve_rubric_offering(
        request.user,
        data.get('offering_id'),
    )
    if error_response:
        return error_response
    offering = next((off for off in offerings if off.id == default_offering_id), None)
    scope = _normalize_rubric_scope(data.get('scope'))
    experiment_number = _normalize_experiment_number(offering, data.get('experiment_number'))
    if scope != FinalRubric.SCOPE_DEFAULT and not experiment_number:
        return JsonResponse({
            'status': 'error',
            'message': '対象科目に実験項目が設定されていません',
        }, status=400)
    if scope == FinalRubric.SCOPE_DEFAULT and request.user.userprofile.role != 'admin':
        return JsonResponse({
            'status': 'error',
            'message': 'デフォルト基準は admin のみ編集できます',
        }, status=403)

    payload = data.get('payload') or {}
    if not (payload.get('criteria') or []):
        return JsonResponse({'status': 'error', 'message': 'クライテリアを1件以上設定してください'}, status=400)

    source_rubric = get_copy_source_rubric(offering, scope, data.get('source_rubric_id'))
    rubric = save_new_rubric_version(
        offering,
        experiment_number,
        scope,
        payload,
        request.user,
        source_rubric=source_rubric,
    )
    return JsonResponse({
        'status': 'ok',
        'message': '最終評価基準を保存しました',
        'scope': rubric.scope,
        'rubric_payload': serialize_rubric_definition(rubric),
        'active_version': rubric.version,
    })


@login_required
@role_required('non-editing teacher', 'admin')
def final_rubric_export_api(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'method not allowed'}, status=405)

    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': '不正なリクエストです'}, status=400)

    default_offering_id, offerings, error_response = _resolve_rubric_offering(
        request.user,
        data.get('offering_id'),
    )
    if error_response:
        return error_response
    offering = next((off for off in offerings if off.id == default_offering_id), None)
    scope = _normalize_rubric_scope(data.get('scope'))
    experiment_number = _normalize_experiment_number(offering, data.get('experiment_number'))
    payload = data.get('payload') or {}
    if scope != FinalRubric.SCOPE_DEFAULT and not experiment_number:
        return JsonResponse({'status': 'error', 'message': '実験項目が不正です'}, status=400)
    if not (payload.get('criteria') or []):
        return JsonResponse({'status': 'error', 'message': 'エクスポートするクライテリアがありません'}, status=400)

    active_rubric = get_active_final_rubric(offering, experiment_number if scope != FinalRubric.SCOPE_DEFAULT else '', scope)
    workbook = _build_rubric_export_workbook(
        offering,
        experiment_number if scope != FinalRubric.SCOPE_DEFAULT else '',
        scope,
        payload,
        active_version=active_rubric.version if active_rubric else None,
    )
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename_suffix = experiment_number if scope != FinalRubric.SCOPE_DEFAULT else 'default'
    filename = f"final_rubric_{offering.course.code}_{offering.year}_{filename_suffix}_{scope}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@role_required('non-editing teacher', 'admin')
def final_rubric_import_api(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'method not allowed'}, status=405)

    default_offering_id, offerings, error_response = _resolve_rubric_offering(
        request.user,
        request.POST.get('offering_id'),
    )
    if error_response:
        return error_response
    offering = next((off for off in offerings if off.id == default_offering_id), None)
    scope = _normalize_rubric_scope(request.POST.get('scope'))
    experiment_number = _normalize_experiment_number(offering, request.POST.get('experiment_number'))
    if scope == FinalRubric.SCOPE_DEFAULT and request.user.userprofile.role != 'admin':
        return JsonResponse({'status': 'error', 'message': 'デフォルト基準は admin のみ編集できます'}, status=403)
    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return JsonResponse({'status': 'error', 'message': 'Excel ファイルを選択してください'}, status=400)

    try:
        payload = _parse_rubric_import_workbook(uploaded_file)
    except ValueError as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'Excel の読み込みに失敗しました'}, status=400)

    return JsonResponse({
        'status': 'ok',
        'message': 'Excel を読み込みました。内容を確認してから保存してください。',
        'rubric_payload': payload,
    })


@login_required
@role_required('non-editing teacher', 'admin')
def final_grading_form(request, submission_id):
    submission = get_object_or_404(Submission, pk=submission_id)
    rubric_error = ''
    submitted_selection = None
    final_comment_value = submission.final_comment or ''
    existing_rubric_score = getattr(submission, 'final_rubric_score', None)
    adjustment_score_value = existing_rubric_score.adjustment_score if existing_rubric_score else Decimal('0')

    def calc_total(sub):
        if not sub or not sub.score_details:
            return 0
        return sum(d.get('value', 0) * d.get('weight', 1) for d in sub.score_details)

    prep_subs = Submission.objects.filter(
        student=submission.student,
        experiment_number=submission.experiment_number,
        report_type='prep',
        course_offering=submission.course_offering
    )

    main_subs = Submission.objects.filter(
        student=submission.student,
        experiment_number=submission.experiment_number,
        report_type='main',
        course_offering=submission.course_offering
    )

    # 得点項目マスターを取得
    def _merge_items(common_items, specific_items):
        merged = []
        index_by_key = {}
        for item in common_items:
            key = item.get('code') or item.get('label') or ''
            if not key or key in index_by_key:
                continue
            index_by_key[key] = len(merged)
            merged.append(item)
        for item in specific_items:
            key = item.get('code') or item.get('label') or ''
            if not key:
                continue
            if key in index_by_key:
                merged[index_by_key[key]] = item
            else:
                index_by_key[key] = len(merged)
                merged.append(item)
        return merged

    def _master_items(category):
        common_items = []
        specific_items = []
        if submission.course_offering:
            common_qs = ScoringItem.objects.filter(
                category=category,
                course_id=submission.course_offering.course_id,
                course_offering__isnull=True
            ).order_by('order')
            common_items = list(common_qs.values('label', 'weight', 'code'))
            specific_qs = ScoringItem.objects.filter(
                category=category,
                course_offering=submission.course_offering
            ).order_by('order')
            specific_items = list(specific_qs.values('label', 'weight', 'code'))
        return _merge_items(common_items, specific_items)

    pre_master = _master_items('pre')
    main_master = _master_items('main')

    def attach_values(master, detail_lists):
        result = []
        for m in master:
            val = 0
            if detail_lists:
                for details in detail_lists:
                    for d in details:
                        if d.get('code') and m.get('code') and d.get('code') == m.get('code'):
                            val += d.get('value', 0)
                            break
                        if d.get('label') == m.get('label'):
                            val += d.get('value', 0)
                            break
            result.append({
                'label': m['label'],
                'weight': int(m['weight']),
                'value': val,
                'code': m.get('code') or '',
            })
        return result

    pre_details_list = [s.score_details for s in prep_subs if s.score_details]
    main_details_list = [s.score_details for s in main_subs if s.score_details]

    pre_items = attach_values(pre_master, pre_details_list)
    main_items = attach_values(main_master, main_details_list)

    total_score = (
        sum(i['value'] * i.get('weight', 1) for i in pre_items)
        + sum(i['value'] * i.get('weight', 1) for i in main_items)
    )

    active_rubric = get_effective_final_rubric(submission.course_offering, submission.experiment_number)

    if request.method == 'POST':
        final_comment = request.POST.get('final_comment', '').strip()
        final_comment_value = final_comment
        try:
            adjustment_score_value = Decimal(str(request.POST.get('adjustment_score', '0') or '0'))
        except Exception:
            adjustment_score_value = Decimal('0')
        if active_rubric is None:
            rubric_error = '最終評価基準が未設定です。先に評価基準を作成してください。'
        else:
            try:
                submitted_selection = json.loads(request.POST.get('rubric_selection_json') or '{}')
            except json.JSONDecodeError:
                submitted_selection = {}
            _, missing = save_rubric_score(
                submission,
                active_rubric,
                submitted_selection,
                final_comment,
                adjustment_score_value,
            )
            if not missing:
                if request.user.userprofile.role == 'admin':
                    return redirect('/submission/admin_dashboard/')
                return redirect('/submission/non_editing_teacher_dashboard/')
            rubric_error = '未選択のクライテリアがあります: ' + ' / '.join(missing)

    candidates = []
    candidate_qs = Submission.objects.filter(
        experiment_number=submission.experiment_number,
        report_type=submission.report_type,
        course_offering=submission.course_offering,
    ).exclude(student=submission.student).select_related('student__userprofile').order_by(
        'student__userprofile__full_name', '-submitted_at'
    )
    seen_user_ids = set()
    for cand in candidate_qs:
        if cand.student_id in seen_user_ids:
            continue
        seen_user_ids.add(cand.student_id)
        up = getattr(cand.student, 'userprofile', None)
        candidates.append({
            'user_id': cand.student_id,
            'full_name': up.full_name if up else cand.student.username,
            'student_id': up.student_id if up else '',
        })

    completed_tasks = set(
        str(task_no)
        for task_no in ExperimentProgress.objects.filter(
            student=submission.student,
            course_offering=submission.course_offering,
            experiment_number=submission.experiment_number,
        ).values_list('task_no', flat=True)
    )
    task_config = ExperimentTaskConfig.objects.filter(
        course_offering=submission.course_offering,
        experiment_number=submission.experiment_number,
    ).first()
    configured_tasks = []
    if task_config and isinstance(task_config.task_list, list):
        configured_tasks = [str(task).strip() for task in task_config.task_list if str(task).strip()]
    ordered_completed_tasks = [task for task in configured_tasks if task in completed_tasks]
    ordered_completed_tasks.extend(sorted(completed_tasks - set(ordered_completed_tasks)))

    rubric_state = build_rubric_state_for_submission(submission)
    if isinstance(submitted_selection, dict) and rubric_state.get('exists'):
        rubric_state['selected_option_ids'] = {
            str(key): value for key, value in submitted_selection.items()
        }
        rubric_state['total_score'] = None
    rubric_state['adjustment_score'] = float(adjustment_score_value)

    rubric_total = rubric_state.get('total_score')
    if rubric_total is None and submission.final_score is not None:
        rubric_total = float(submission.final_score)

    return render(request, 'submission/final_grading_form.html', {
        'submission': submission,
        'total_score': total_score,
        'pre_items': pre_items,
        'main_items': main_items,
        'final_comment': final_comment_value,
        'completed_tasks_text': ', '.join(ordered_completed_tasks) if ordered_completed_tasks else '-',
        'compare_candidates': json.dumps(candidates, ensure_ascii=False),
        'rubric_state_json': json.dumps(rubric_state, ensure_ascii=False),
        'rubric_error': rubric_error,
        'rubric_error_json': json.dumps(rubric_error, ensure_ascii=False),
        'rubric_total': rubric_total if rubric_total is not None else '',
        'adjustment_score': adjustment_score_value,
    })


@login_required
@role_required('non-editing teacher', 'teacher', 'admin', 'course-teacher')
def compare_user_submission(request):
    submission_id = request.GET.get('submission_id')
    user_id = request.GET.get('user_id')
    if not submission_id or not user_id:
        return JsonResponse({'status': 'error', 'message': 'submission_id and user_id are required'}, status=400)
    submission = get_object_or_404(Submission, pk=submission_id)
    candidate = Submission.objects.filter(
        student_id=user_id,
        experiment_number=submission.experiment_number,
        report_type=submission.report_type,
        course_offering=submission.course_offering,
    ).exclude(id=submission.id).select_related('student__userprofile').order_by('-submitted_at').first()
    if not candidate or not candidate.file:
        return JsonResponse({'status': 'not_found'})
    up = getattr(candidate.student, 'userprofile', None)
    name = up.full_name if up else candidate.student.username
    submitted_at = timezone.localtime(candidate.submitted_at).strftime('%Y-%m-%d %H:%M')
    return JsonResponse({
        'status': 'ok',
        'pdf_url': candidate.file.url,
        'submitted_at': submitted_at,
        'full_name': name,
    })


def _normalize_pdf_text(text):
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKC", text)
    normalized = re.sub(r"\s+", "", normalized)
    return normalized


def _is_page_number_footer_line(text):
    raw = unicodedata.normalize("NFKC", (text or "")).strip()
    if not raw:
        return False
    compact = re.sub(r"\s+", "", raw)
    patterns = [
        r"^\d+$",                       # 1
        r"^\d+/\d+$",                   # 1/10
        r"^-\d+-$",                     # - 1 -
        r"^page\d+$",                   # Page 1
        r"^p\.\d+$",                    # p.1
        r"^\d+ページ$",                  # 1ページ
        r"^第?\d+頁$",                  # 第1頁
    ]
    lower_compact = compact.lower()
    return any(re.match(p, lower_compact) for p in patterns)


def _extract_page_lines_without_footer(page):
    # レイアウト情報を使い、ページ下端のページ番号行を除去する。
    try:
        page_dict = page.get_text("dict")
    except Exception:
        text = page.get_text("text") or ""
        return [line for line in text.splitlines() if line.strip()]

    footer_start_y = page.rect.height * 0.92
    line_items = []
    for block in page_dict.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            spans = line.get("spans", [])
            if not spans:
                continue
            text = "".join((span.get("text") or "") for span in spans).strip()
            if not text:
                continue
            bbox = line.get("bbox") or [0, 0, 0, 0]
            y_mid = (float(bbox[1]) + float(bbox[3])) / 2 if len(bbox) >= 4 else 0.0
            if y_mid >= footer_start_y and _is_page_number_footer_line(text):
                continue
            x0 = float(bbox[0]) if len(bbox) >= 1 else 0.0
            y0 = float(bbox[1]) if len(bbox) >= 2 else 0.0
            line_items.append((y0, x0, text))

    line_items.sort(key=lambda item: (item[0], item[1]))
    return [item[2] for item in line_items]


@lru_cache(maxsize=256)
def _extract_pdf_lines_cached(file_path, mtime):
    try:
        with fitz.open(file_path) as doc:
            lines = []
            for page in doc:
                lines.extend(_extract_page_lines_without_footer(page))
        return tuple(lines)
    except Exception:
        return tuple()


@lru_cache(maxsize=256)
def _extract_pdf_text_cached(file_path, mtime):
    try:
        lines = _extract_pdf_lines_cached(file_path, mtime)
        return _normalize_pdf_text("\n".join(lines))
    except Exception:
        return ""


def _extract_pdf_text(file_path):
    try:
        mtime = os.path.getmtime(file_path)
    except OSError:
        return ""
    return _extract_pdf_text_cached(file_path, mtime)


def _extract_pdf_lines(file_path):
    try:
        mtime = os.path.getmtime(file_path)
    except OSError:
        return []
    return list(_extract_pdf_lines_cached(file_path, mtime))


def _hash_file_sha256(file_path):
    digest = hashlib.sha256()
    with open(file_path, 'rb') as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b''):
            digest.update(chunk)
    return digest.hexdigest()


def _build_minhash_signature(text, n=8, size=96):
    if not text:
        return []
    hashes = set()
    if len(text) <= n:
        chunk = text
        hv = int.from_bytes(hashlib.blake2b(chunk.encode('utf-8'), digest_size=8).digest(), 'big')
        hashes.add(hv)
    else:
        for i in range(len(text) - n + 1):
            chunk = text[i:i + n]
            hv = int.from_bytes(hashlib.blake2b(chunk.encode('utf-8'), digest_size=8).digest(), 'big')
            hashes.add(hv)
    return sorted(hashes)[:size]


def _signature_overlap_score(sig_a, sig_b):
    if not sig_a or not sig_b:
        return 0.0
    set_a = set(sig_a)
    set_b = set(sig_b)
    union = set_a | set_b
    if not union:
        return 0.0
    inter = set_a & set_b
    return round((len(inter) / len(union)) * 100, 1)


def _serialize_sections_for_cache(sections):
    serialized = []
    for section in sections:
        text_norm = _normalize_pdf_text(section.get('text_norm', ''))
        if not text_norm:
            continue
        minor_spans = []
        for span in section.get('minor_spans', []) or []:
            if not isinstance(span, dict):
                continue
            start = int(span.get('start', 0) or 0)
            end = int(span.get('end', 0) or 0)
            if end <= start:
                continue
            minor_heading = str(span.get('minor_heading') or '').strip()
            if not minor_heading:
                minor_heading = '小見出しなし'
            minor_spans.append({
                'start': start,
                'end': end,
                'minor_heading': minor_heading,
            })
        compare_unit_key = str(section.get('compare_unit_key') or '').strip()
        if not compare_unit_key:
            compare_unit_key = section.get('title', '') or ''
        compare_unit_label = str(section.get('compare_unit_label') or '').strip()
        if not compare_unit_label:
            compare_unit_label = section.get('title', '') or ''
        serialized.append({
            'title': section.get('title', ''),
            'heading': section.get('heading', ''),
            'section_key': section.get('section_key', ''),
            'compare_unit_key': compare_unit_key,
            'compare_unit_label': compare_unit_label,
            'text_norm': text_norm,
            'minor_spans': minor_spans,
        })
    return serialized


def _load_sections_from_cache(data):
    if not isinstance(data, list):
        return []
    sections = []
    for item in data:
        if not isinstance(item, dict):
            continue
        text_norm = _normalize_pdf_text(item.get('text_norm', ''))
        if not text_norm:
            continue
        minor_spans = []
        for span in item.get('minor_spans', []) or []:
            if not isinstance(span, dict):
                continue
            start = int(span.get('start', 0) or 0)
            end = int(span.get('end', 0) or 0)
            if end <= start:
                continue
            minor_heading = str(span.get('minor_heading') or '').strip()
            if not minor_heading:
                minor_heading = '小見出しなし'
            minor_spans.append({
                'start': start,
                'end': end,
                'minor_heading': minor_heading,
            })
        compare_unit_key = str(item.get('compare_unit_key') or '').strip()
        if not compare_unit_key:
            compare_unit_key = item.get('title', '') or ''
        compare_unit_label = str(item.get('compare_unit_label') or '').strip()
        if not compare_unit_label:
            compare_unit_label = item.get('title', '') or ''
        sections.append({
            'title': item.get('title', ''),
            'heading': item.get('heading', ''),
            'section_key': item.get('section_key', ''),
            'compare_unit_key': compare_unit_key,
            'compare_unit_label': compare_unit_label,
            'text_norm': text_norm,
            'minor_spans': minor_spans,
        })
    return sections


def _get_submission_text_index(submission):
    if not submission.file:
        return None
    try:
        file_path = submission.file.path
        st = os.stat(file_path)
    except (OSError, ValueError):
        return None

    file_size = int(st.st_size or 0)
    file_mtime = float(st.st_mtime or 0.0)
    idx, _ = SubmissionTextIndex.objects.get_or_create(submission=submission)

    is_fresh = (
        idx.index_version == INDEX_VERSION
        and int(idx.file_size or 0) == file_size
        and abs(float(idx.file_mtime or 0.0) - file_mtime) < 1e-6
        and bool(idx.normalized_text)
    )
    if is_fresh:
        return idx

    raw_lines = [line.strip() for line in _extract_pdf_lines(file_path) if (line or "").strip()]
    sections = _build_sections_from_lines(raw_lines)
    normalized_text = "".join(section.get('text_norm', '') for section in sections if section.get('text_norm'))
    signature = _build_minhash_signature(normalized_text)
    file_hash = _hash_file_sha256(file_path)

    idx.index_version = INDEX_VERSION
    idx.file_size = file_size
    idx.file_mtime = file_mtime
    idx.file_hash = file_hash
    idx.normalized_text = normalized_text
    idx.sections_json = _serialize_sections_for_cache(sections)
    idx.signature_json = signature
    idx.save(update_fields=[
        'index_version',
        'file_size',
        'file_mtime',
        'file_hash',
        'normalized_text',
        'sections_json',
        'signature_json',
        'indexed_at',
    ])
    return idx


SECTION_TITLES = [
    '目的', '原理', '実験方法', '予習課題', '使用器具',
    '実験結果', '考察', '課題', '参考文献',
]
DEFAULT_INFO_MIN_LEN = 20
TOP_SIMILARITY_RESULTS = 10
INDEX_VERSION = 'v3'
SIMILARITY_ALGORITHM_VERSION = 'v6-section-rough-major-lock'
SIMILARITY_CACHE_TTL_HOURS = 24
ROUGH_TOP_CANDIDATES = 40
MINHASH_N = 8
MINHASH_SIZE = 96
SECTION_INFO_MIN_LEN = {
    '考察': 10,
    '予習課題': 10,
}
SECTION_TITLE_ORDER = {title: i for i, title in enumerate(SECTION_TITLES)}
SECTION_HEADING_RE = re.compile(
    r'^(?P<prefix>I{1,2})\s*[-‐‑‒–—―ー－]\s*(?P<major>\d+)\s*[\.．]\s*(?P<minor>\d+)\s*(?P<title>'
    + "|".join(re.escape(t) for t in SECTION_TITLES) +
    r')\s*$',
    flags=re.IGNORECASE
)
COMPARE_UNIT_KEY_RE = re.compile(r'^(?P<prefix>I{1,2})-(?P<major>\d+):(?P<title>.+)$')
SECTION_PREFIX_ORDER = {'I': 0, 'II': 1}
MINOR_SUBHEADING_RE = re.compile(
    r'^(?P<num>('
    r'\d+\s*[\.．]\s*\d+(?:\s*[\.．]\s*\d+)?'
    r'|'
    r'\d+\s*[-‐‑‒–—―ー－]\s*\d+(?:\s*[-‐‑‒–—―ー－]\s*\d+)?'
    r'))\s*(?P<body>.+)$'
)
MINOR_SUBHEADING_MAX_LEN = 48
MINOR_SUBHEADING_TRAILING_PUNCT = ('。', '．', '.', '!', '?', '！', '？', '、', ',', '，', ';', '；', ':', '：')
DEFAULT_MINOR_HEADING_LABEL = '小見出しなし'


def _parse_minor_subheading_line(line):
    normalized = unicodedata.normalize("NFKC", (line or "")).strip()
    if not normalized:
        return None
    normalized = re.sub(r"\s+", " ", normalized)
    if len(normalized) > MINOR_SUBHEADING_MAX_LEN:
        return None

    match = MINOR_SUBHEADING_RE.match(normalized)
    if not match:
        return None

    number_text = re.sub(r"\s+", "", (match.group('num') or ''))
    body = (match.group('body') or '').strip()
    if len(body) < 2:
        return None
    if body.endswith(MINOR_SUBHEADING_TRAILING_PUNCT):
        return None
    if not re.search(r'[A-Za-zぁ-んァ-ヶ一-龥]', body):
        return None
    if re.fullmatch(r'[\d\W_]+', body):
        return None

    compact = re.sub(r"\s+", "", body)
    if not compact:
        return None
    digit_count = sum(ch.isdigit() for ch in compact)
    if digit_count > 0 and (digit_count / len(compact)) >= 0.45:
        return None
    if re.search(r'[=<>±×/]', body) and digit_count > 0:
        return None
    return f"{number_text} {body}"


def _match_section_heading(line):
    normalized = unicodedata.normalize("NFKC", (line or "")).strip()
    normalized = re.sub(r"\s+", " ", normalized)
    match = SECTION_HEADING_RE.match(normalized)
    if not match:
        return None
    prefix = match.group('prefix').upper()
    major = str(match.group('major'))
    minor = str(match.group('minor'))
    title = match.group('title')
    heading = f"{prefix}-{major}.{minor} {title}"
    compare_unit_key = f"{prefix}-{major}:{title}"
    compare_unit_label = f"{prefix}-{major} {title}"
    return {
        'title': title,
        'heading': heading,
        'compare_unit_key': compare_unit_key,
        'compare_unit_label': compare_unit_label,
    }


def _build_sections_from_lines(lines):
    sections = []
    current = None
    section_counts = {}

    def _start_section(title, heading, compare_unit_key="", compare_unit_label=""):
        count = section_counts.get(title, 0) + 1
        section_counts[title] = count
        normalized_compare_unit_key = (compare_unit_key or "").strip() or title
        normalized_compare_unit_label = (compare_unit_label or "").strip() or title
        return {
            'title': title,
            'heading': heading,
            'section_key': f'{title}#{count}',
            'compare_unit_key': normalized_compare_unit_key,
            'compare_unit_label': normalized_compare_unit_label,
            'minor_heading': DEFAULT_MINOR_HEADING_LABEL,
            'content_items': [],
        }

    def _flush():
        nonlocal current
        if not current:
            return
        raw_parts = []
        norm_parts = []
        minor_spans = []
        cursor = 0
        for item in current.get('content_items', []):
            raw_line = (item.get('text') or '').strip()
            if not raw_line:
                continue
            raw_parts.append(raw_line)
            norm_line = _normalize_pdf_text(raw_line)
            if not norm_line:
                continue
            start = cursor
            end = start + len(norm_line)
            minor_heading = (item.get('minor_heading') or DEFAULT_MINOR_HEADING_LABEL).strip() or DEFAULT_MINOR_HEADING_LABEL
            if minor_spans and minor_spans[-1]['minor_heading'] == minor_heading and minor_spans[-1]['end'] == start:
                minor_spans[-1]['end'] = end
            else:
                minor_spans.append({
                    'start': start,
                    'end': end,
                    'minor_heading': minor_heading,
                })
            cursor = end
            norm_parts.append(norm_line)
        raw_text = "\n".join(raw_parts).strip()
        norm_text = "".join(norm_parts)
        if norm_text:
            sections.append({
                'title': current['title'],
                'heading': current['heading'],
                'section_key': current['section_key'],
                'compare_unit_key': current.get('compare_unit_key') or current['title'],
                'compare_unit_label': current.get('compare_unit_label') or current['title'],
                'text_raw': raw_text,
                'text_norm': norm_text,
                'minor_spans': minor_spans,
            })
        current = None

    for line in lines:
        raw_line = (line or "").strip()
        if not raw_line:
            continue
        heading = _match_section_heading(raw_line)
        if heading:
            _flush()
            current = _start_section(
                heading['title'],
                heading['heading'],
                heading.get('compare_unit_key', ''),
                heading.get('compare_unit_label', ''),
            )
            continue
        if current is None:
            current = _start_section('未分類', '', '未分類', '未分類')
        minor_heading = _parse_minor_subheading_line(raw_line)
        if minor_heading:
            current['minor_heading'] = minor_heading
            continue
        current['content_items'].append({
            'text': raw_line,
            'minor_heading': current.get('minor_heading') or DEFAULT_MINOR_HEADING_LABEL,
        })

    _flush()
    return sections


def _group_sections_by_compare_unit(sections):
    grouped = {}
    for section in sections:
        key = str(section.get('compare_unit_key') or section.get('title') or '').strip()
        if not key:
            key = '未分類'
        grouped.setdefault(key, []).append(section)
    return grouped


def _compare_unit_sort_key(unit_key):
    normalized = str(unit_key or '').strip()
    match = COMPARE_UNIT_KEY_RE.match(normalized)
    if match:
        prefix = match.group('prefix').upper()
        major = int(match.group('major'))
        title = match.group('title')
        return (
            SECTION_PREFIX_ORDER.get(prefix, 9),
            major,
            SECTION_TITLE_ORDER.get(title, 999),
            title,
            normalized,
        )
    return (9, 999999, SECTION_TITLE_ORDER.get(normalized, 999), normalized, normalized)


def _risk_level_from_summary(high_count, medium_count, has_info):
    if high_count > 0:
        return 'high'
    if medium_count > 0:
        return 'medium'
    if has_info:
        return 'low'
    return 'none'


def _resolve_minor_heading_labels(spans, start, end):
    if not isinstance(spans, list):
        return [DEFAULT_MINOR_HEADING_LABEL]
    labels = []
    for span in spans:
        if not isinstance(span, dict):
            continue
        span_start = int(span.get('start', 0) or 0)
        span_end = int(span.get('end', 0) or 0)
        if span_end <= start or end <= span_start:
            continue
        label = str(span.get('minor_heading') or '').strip() or DEFAULT_MINOR_HEADING_LABEL
        if label not in labels:
            labels.append(label)
    if not labels:
        return [DEFAULT_MINOR_HEADING_LABEL]
    return labels


def _format_minor_heading_label(labels):
    if not labels:
        return DEFAULT_MINOR_HEADING_LABEL
    if len(labels) == 1:
        return labels[0]
    if len(labels) == 2:
        return f"{labels[0]} ～ {labels[1]}"
    return f"{labels[0]} ～ {labels[-1]}"


def _compare_section_texts(target_text, candidate_text, min_match_len):
    if not target_text or not candidate_text:
        return 0.0, []
    matcher = SequenceMatcher(None, target_text, candidate_text, autojunk=False)
    blocks = matcher.get_matching_blocks()
    total_match = 0
    matches = []
    seen = set()
    for block in blocks:
        size = int(block.size or 0)
        if size <= 0:
            continue
        total_match += size
        if size < min_match_len:
            continue
        snippet = target_text[block.a:block.a + size]
        preview = snippet[:120] + ('...' if len(snippet) > 120 else '')
        key = (size, preview, int(block.a or 0), int(block.b or 0))
        if key in seen:
            continue
        seen.add(key)
        matches.append({
            'length': size,
            'snippet': preview,
            'target_start': int(block.a or 0),
            'target_end': int(block.a or 0) + size,
            'candidate_start': int(block.b or 0),
            'candidate_end': int(block.b or 0) + size,
        })
    similarity = round((2 * total_match / (len(target_text) + len(candidate_text))) * 100, 1)
    matches.sort(key=lambda item: item['length'], reverse=True)
    return similarity, matches


def _compare_sections(target_sections, candidate_sections, min_match_len=20, alert_match_len=30):
    target_map = _group_sections_by_compare_unit(target_sections)
    candidate_map = _group_sections_by_compare_unit(candidate_sections)
    common_units = sorted(set(target_map.keys()) & set(candidate_map.keys()), key=_compare_unit_sort_key)

    section_details = []
    max_similarity = 0.0
    total_alert_matches = 0
    total_info_matches = 0

    for unit_key in common_units:
        target_unit_sections = target_map.get(unit_key, [])
        candidate_unit_sections = candidate_map.get(unit_key, [])
        if not target_unit_sections or not candidate_unit_sections:
            continue
        section_title = target_unit_sections[0].get('title') or candidate_unit_sections[0].get('title') or str(unit_key)
        section_label = (
            target_unit_sections[0].get('compare_unit_label')
            or candidate_unit_sections[0].get('compare_unit_label')
            or section_title
        )
        title_min_match_len = SECTION_INFO_MIN_LEN.get(section_title, min_match_len)
        pair_max_similarity = 0.0
        merged_matches = []
        for target_section in target_unit_sections:
            for candidate_section in candidate_unit_sections:
                similarity, pair_matches = _compare_section_texts(
                    target_section['text_norm'],
                    candidate_section['text_norm'],
                    min_match_len=title_min_match_len,
                )
                pair_max_similarity = max(pair_max_similarity, similarity)
                for item in pair_matches:
                    target_minor_labels = _resolve_minor_heading_labels(
                        target_section.get('minor_spans', []),
                        item.get('target_start', 0),
                        item.get('target_end', 0),
                    )
                    candidate_minor_labels = _resolve_minor_heading_labels(
                        candidate_section.get('minor_spans', []),
                        item.get('candidate_start', 0),
                        item.get('candidate_end', 0),
                    )
                    merged_matches.append({
                        'length': item['length'],
                        'snippet': item['snippet'],
                        'target_section': target_section['section_key'],
                        'candidate_section': candidate_section['section_key'],
                        'target_major_heading': target_section.get('heading') or section_label,
                        'candidate_major_heading': candidate_section.get('heading') or section_label,
                        'target_minor_heading': _format_minor_heading_label(target_minor_labels),
                        'candidate_minor_heading': _format_minor_heading_label(candidate_minor_labels),
                    })

        if not merged_matches and pair_max_similarity <= 0:
            continue

        merged_matches.sort(key=lambda item: item['length'], reverse=True)
        deduped = []
        seen_keys = set()
        for item in merged_matches:
            key = (
                item['length'],
                item['snippet'],
                item.get('target_major_heading', ''),
                item.get('candidate_major_heading', ''),
                item.get('target_minor_heading', ''),
                item.get('candidate_minor_heading', ''),
            )
            if key in seen_keys:
                continue
            seen_keys.add(key)
            deduped.append(item)

        all_matches = deduped[:200]
        top_matches = all_matches[:5]
        info_count = len(all_matches)
        alert_count = sum(1 for item in all_matches if item['length'] >= alert_match_len)

        total_info_matches += info_count
        total_alert_matches += alert_count
        max_similarity = max(max_similarity, pair_max_similarity)

        if alert_count > 0 and pair_max_similarity >= 30.0:
            level = 'high'
        elif info_count > 0 and pair_max_similarity >= 15.0:
            level = 'medium'
        elif info_count > 0:
            level = 'low'
        else:
            level = 'none'

        section_details.append({
            'title': section_label,
            'section_title': section_title,
            'compare_unit_key': str(unit_key),
            'level': level,
            'max_similarity': pair_max_similarity,
            'info_match_count': info_count,
            'alert_match_count': alert_count,
            'info_min_len': title_min_match_len,
            'top_matches': top_matches,
            'all_matches': all_matches,
        })

    section_details.sort(key=lambda item: _compare_unit_sort_key(item.get('compare_unit_key', item.get('title', ''))))
    high_count = sum(1 for item in section_details if item['level'] == 'high')
    medium_count = sum(1 for item in section_details if item['level'] == 'medium')
    has_info = any(item['info_match_count'] > 0 for item in section_details)
    risk_level = _risk_level_from_summary(high_count, medium_count, has_info)
    risk_score = (
        high_count * 100000
        + medium_count * 10000
        + int(max_similarity * 100)
        + total_alert_matches
    )
    badges = [{'title': item['title'], 'level': item['level']} for item in section_details if item['level'] in ('high', 'medium')]

    return {
        'section_details': section_details,
        'section_badges': badges,
        'max_similarity': round(max_similarity, 1),
        'risk_level': risk_level,
        'risk_score': risk_score,
        'total_alert_matches': total_alert_matches,
        'total_info_matches': total_info_matches,
    }


def _char_ngrams(text, n=8):
    if not text:
        return set()
    if len(text) <= n:
        return {text}
    return {text[i:i + n] for i in range(len(text) - n + 1)}


def _jaccard_similarity_percent(text_a, text_b, n=8):
    grams_a = _char_ngrams(text_a, n=n)
    grams_b = _char_ngrams(text_b, n=n)
    if not grams_a or not grams_b:
        return 0.0
    union = grams_a | grams_b
    if not union:
        return 0.0
    inter = grams_a & grams_b
    return round((len(inter) / len(union)) * 100, 1)


@login_required
@role_required('teacher', 'admin', 'course-teacher', 'non-editing teacher')
def submission_similarity_api(request):
    submission_id = request.GET.get('submission_id')
    if not submission_id:
        return JsonResponse({'status': 'error', 'message': 'submission_id is required'}, status=400)

    submission = get_object_or_404(Submission, pk=submission_id)
    if not submission.file:
        return JsonResponse({'status': 'error', 'message': '対象PDFがありません'}, status=404)
    now = timezone.now()
    job, _ = SimilarityJob.objects.get_or_create(
        target_submission=submission,
        algorithm_version=SIMILARITY_ALGORITHM_VERSION,
        defaults={'status': 'queued'}
    )
    if (
        job.status == 'done'
        and job.expires_at
        and job.expires_at > now
        and isinstance(job.result_cache_json, dict)
        and job.result_cache_json.get('status') == 'success'
    ):
        cached_payload = dict(job.result_cache_json)
        cached_payload['cached'] = True
        return JsonResponse(cached_payload)

    job.status = 'running'
    job.error_message = ''
    job.started_at = now
    job.finished_at = None
    job.save(update_fields=['status', 'error_message', 'started_at', 'finished_at', 'updated_at'])

    try:
        target_index = _get_submission_text_index(submission)
        if not target_index or not target_index.normalized_text:
            payload = {
                'status': 'success',
                'message': '対象PDFからテキストを抽出できませんでした。',
                'rules': {
                    'info_min_len_default': DEFAULT_INFO_MIN_LEN,
                    'info_min_len_overrides': SECTION_INFO_MIN_LEN,
                    'alert_min_len': 30,
                    'high_similarity_threshold': 30,
                    'medium_similarity_threshold': 15,
                    'minor_subheading_excluded': True,
                    'compare_unit_policy': 'I/II + major + title',
                },
                'results': [],
                'displayed_count': 0,
                'checked_count': 0,
            }
            job.status = 'done'
            job.checked_count = 0
            job.displayed_count = 0
            job.result_cache_json = payload
            job.finished_at = timezone.now()
            job.expires_at = job.finished_at + timedelta(hours=SIMILARITY_CACHE_TTL_HOURS)
            job.save(update_fields=[
                'status',
                'checked_count',
                'displayed_count',
                'result_cache_json',
                'finished_at',
                'expires_at',
                'updated_at',
            ])
            return JsonResponse(payload)

        target_text = target_index.normalized_text
        target_sections = _load_sections_from_cache(target_index.sections_json)
        target_signature = target_index.signature_json if isinstance(target_index.signature_json, list) else []

        candidate_qs = Submission.objects.filter(
            experiment_number=submission.experiment_number,
            report_type=submission.report_type,
            course_offering=submission.course_offering,
        ).exclude(student=submission.student).exclude(file='').select_related(
            'student__userprofile'
        ).order_by('student_id', '-submitted_at')

        latest_by_student = []
        seen_student_ids = set()
        for cand in candidate_qs:
            if cand.student_id in seen_student_ids:
                continue
            seen_student_ids.add(cand.student_id)
            latest_by_student.append(cand)

        rough_candidates = []
        for cand in latest_by_student:
            if not cand.file:
                continue
            cand_index = _get_submission_text_index(cand)
            if not cand_index or not cand_index.normalized_text:
                continue
            cand_signature = cand_index.signature_json if isinstance(cand_index.signature_json, list) else []
            rough_score = _signature_overlap_score(target_signature, cand_signature)
            rough_candidates.append((rough_score, cand, cand_index))

        rough_candidates.sort(key=lambda item: item[0], reverse=True)
        if len(rough_candidates) > ROUGH_TOP_CANDIDATES:
            detailed_candidates = rough_candidates[:ROUGH_TOP_CANDIDATES]
        else:
            detailed_candidates = rough_candidates

        results = []
        for _, cand, cand_index in detailed_candidates:
            cand_text = cand_index.normalized_text
            overall_similarity = _jaccard_similarity_percent(target_text, cand_text, n=8)
            candidate_sections = _load_sections_from_cache(cand_index.sections_json)
            section_result = _compare_sections(
                target_sections,
                candidate_sections,
                min_match_len=20,
                alert_match_len=30,
            )
            up = getattr(cand.student, 'userprofile', None)
            results.append({
                'submission_id': cand.id,
                'user_id': cand.student_id,
                'student_name': up.full_name if up else cand.student.username,
                'student_id': up.student_id if up else '',
                'submitted_at': timezone.localtime(cand.submitted_at).strftime('%Y-%m-%d %H:%M') if cand.submitted_at else '',
                'pdf_url': cand.file.url if cand.file else '',
                'overall_similarity': overall_similarity,
                'max_similarity': section_result['max_similarity'],
                'risk_level': section_result['risk_level'],
                'risk_score': section_result['risk_score'],
                'section_badges': section_result['section_badges'],
                'section_details': section_result['section_details'],
                'total_alert_matches': section_result['total_alert_matches'],
                'total_info_matches': section_result['total_info_matches'],
            })

        results.sort(
            key=lambda x: (
                x['risk_score'],
                x['max_similarity'],
                x['overall_similarity'],
            ),
            reverse=True
        )
        displayed_results = results[:TOP_SIMILARITY_RESULTS]
        checked_count = len(rough_candidates)
        payload = {
            'status': 'success',
            'message': (
                f'同一科目/年度・同一実験番号・同一レポート種別で、見出し単位の一致箇所を抽出しました'
                f'（比較{checked_count}件 / 詳細計算{len(results)}件 / 表示上位{TOP_SIMILARITY_RESULTS}件）。'
            ),
            'rules': {
                'info_min_len_default': DEFAULT_INFO_MIN_LEN,
                'info_min_len_overrides': SECTION_INFO_MIN_LEN,
                'alert_min_len': 30,
                'high_similarity_threshold': 30,
                'medium_similarity_threshold': 15,
                'rough_top_candidates': ROUGH_TOP_CANDIDATES,
                'minor_subheading_excluded': True,
                'compare_unit_policy': 'I/II + major + title',
            },
            'results': displayed_results,
            'displayed_count': len(displayed_results),
            'checked_count': checked_count,
            'calculated_count': len(results),
            'cached': False,
        }

        job.status = 'done'
        job.checked_count = checked_count
        job.displayed_count = len(displayed_results)
        job.result_cache_json = payload
        job.finished_at = timezone.now()
        job.expires_at = job.finished_at + timedelta(hours=SIMILARITY_CACHE_TTL_HOURS)
        job.save(update_fields=[
            'status',
            'checked_count',
            'displayed_count',
            'result_cache_json',
            'finished_at',
            'expires_at',
            'updated_at',
        ])
        return JsonResponse(payload)
    except Exception as exc:
        job.status = 'failed'
        job.error_message = str(exc)
        job.finished_at = timezone.now()
        job.save(update_fields=['status', 'error_message', 'finished_at', 'updated_at'])
        return JsonResponse({'status': 'error', 'message': f'コピペチェックに失敗しました: {exc}'}, status=500)
